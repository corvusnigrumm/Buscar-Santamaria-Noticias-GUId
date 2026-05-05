import os
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from core.logger import logger


class GeneradorExcelIDEAS:
    FONT_NAME = "Century Gothic"
    NEGRO = "000000"
    BLANCO = "FFFFFF"
    GRIS_CLARO = "F5F5F5"
    GRIS_BORDE = "CCCCCC"
    COLORES_TAB = [
        "FF4444", "FFD700", "00CC66", "00CCCC", "FF66FF",
        "FF8800", "4488FF", "AA44FF", "44DDAA", "DD4488",
        "88CC00", "0088DD", "FF6644", "6644FF", "44FF88",
    ]

    def __init__(self, articulos):
        self.articulos = articulos
        self.wb = Workbook()

    def generar(self, nombre_archivo):
        logger.info("")
        logger.info("=" * 60)
        logger.info("  Generando Excel (una hoja por categoria)...")
        logger.info("=" * 60)

        por_cat = {}
        for art in self.articulos:
            cat = art.get("categoria", "GENERAL")
            por_cat.setdefault(cat, []).append(art)

        cats_ord = sorted(por_cat.keys())

        ws_res = self.wb.active
        ws_res.title = "RESUMEN"
        ws_res.sheet_properties.tabColor = self.NEGRO
        self._hoja_resumen(ws_res, por_cat, cats_ord)

        for i, cat in enumerate(cats_ord):
            nombre_hoja = cat[:31]
            ws = self.wb.create_sheet(title=nombre_hoja)
            ws.sheet_properties.tabColor = self.COLORES_TAB[i % len(self.COLORES_TAB)]
            self._hoja_datos(ws, por_cat[cat])
            logger.info(f"  OK Hoja '{nombre_hoja}': {len(por_cat[cat])} articulos")

        self.wb.save(nombre_archivo)
        logger.info(f"  OK Guardado: {nombre_archivo}")
        logger.info(f"  OK Total: {len(self.articulos)} en {len(cats_ord)} hojas")
        return nombre_archivo

    def _hoja_resumen(self, ws, por_cat, cats_ord):
        headers = [("CATEGORIA", "FFFFFF"), ("ARTICULOS", "FFD700")]
        for ci, (h, c) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(name=self.FONT_NAME, bold=True, size=11, color=c)
            cell.fill = PatternFill(start_color=self.NEGRO, end_color=self.NEGRO, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        borde = Border(
            left=Side(style="hair", color=self.GRIS_BORDE),
            right=Side(style="hair", color=self.GRIS_BORDE),
            top=Side(style="hair", color=self.GRIS_BORDE),
            bottom=Side(style="hair", color=self.GRIS_BORDE),
        )
        fp = PatternFill(start_color=self.GRIS_CLARO, end_color=self.GRIS_CLARO, fill_type="solid")
        fi = PatternFill(start_color=self.BLANCO, end_color=self.BLANCO, fill_type="solid")

        total = 0
        for idx, cat in enumerate(cats_ord):
            row = idx + 2
            cnt = len(por_cat[cat])
            total += cnt
            fill = fp if idx % 2 == 0 else fi

            c = ws.cell(row=row, column=1, value=cat)
            c.font = Font(name=self.FONT_NAME, size=10, bold=True, color="333333")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.fill = fill
            c.border = borde

            c = ws.cell(row=row, column=2, value=cnt)
            c.font = Font(name=self.FONT_NAME, size=10, color="333333")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = fill
            c.border = borde

        rt = len(cats_ord) + 2
        for ci, (v, clr) in enumerate([("TOTAL", "FFFFFF"), (total, "FFD700")], 1):
            c = ws.cell(row=rt, column=ci, value=v)
            c.font = Font(name=self.FONT_NAME, size=11, bold=True, color=clr)
            c.fill = PatternFill(start_color=self.NEGRO, end_color=self.NEGRO, fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borde

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.freeze_panes = "A2"

    def _hoja_datos(self, ws, articulos):
        headers = [
            ("FUENTE", "FF4444"),
            ("TITULO", "FFD700"),
            ("RESUMEN CORTO", "00CC66"),
            ("URL", "00CCCC"),
            ("FECHA", "FF66FF"),
            ("TAGS", "4488FF"),
            ("SCORE", "88CC00"),
            ("KEYWORDS SEO", "FF8800"),
            ("SEO TITLE", "6644FF"),
            ("META DESCRIPTION", "44DDAA"),
            ("SEO ANGLE", "DD4488"),
            ("EVERGREEN", "0088DD"),
            ("INTENCION SEO", "FF6644"),
            ("RAZON INTENCION", "44FF88"),
            ("VISIBILIDAD", "FF9900"),
            ("AUDIENCIA", "0099FF"),
        ]
        for ci, (h, c) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(name=self.FONT_NAME, bold=True, size=10, color=c)
            cell.fill = PatternFill(start_color=self.NEGRO, end_color=self.NEGRO, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        borde = Border(
            left=Side(style="hair", color=self.GRIS_BORDE),
            right=Side(style="hair", color=self.GRIS_BORDE),
            top=Side(style="hair", color=self.GRIS_BORDE),
            bottom=Side(style="hair", color=self.GRIS_BORDE),
        )
        fn = Font(name=self.FONT_NAME, size=9, color="333333")
        fl = Font(name=self.FONT_NAME, size=9, color="0563C1", underline="single")
        al = Alignment(vertical="center", wrap_text=True)
        fp = PatternFill(start_color=self.GRIS_CLARO, end_color=self.GRIS_CLARO, fill_type="solid")
        fi = PatternFill(start_color=self.BLANCO, end_color=self.BLANCO, fill_type="solid")

        fill_alto = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_medio = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        fill_bajo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        font_alto = Font(name=self.FONT_NAME, size=10, bold=True, color="006100")
        font_medio = Font(name=self.FONT_NAME, size=10, bold=True, color="9C6500")
        font_bajo = Font(name=self.FONT_NAME, size=10, bold=True, color="9C0006")

        arts = sorted(
            articulos,
            key=lambda a: a.get("fecha_dt") or datetime.min.replace(tzinfo=timezone.utc),
            reverse=True,
        )
        for idx, art in enumerate(arts):
            row = idx + 2
            fill = fp if idx % 2 == 0 else fi

            c = ws.cell(row=row, column=1, value=art.get("fuente", ""))
            c.font = Font(name=self.FONT_NAME, size=9, bold=True, color="333333")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = fill
            c.border = borde

            c = ws.cell(row=row, column=2, value=art.get("titulo", ""))
            c.font = fn
            c.alignment = al
            c.fill = fill
            c.border = borde

            resumen = art.get("resumen", "")
            if len(resumen) > 150:
                resumen = resumen[:147] + "..."
            c = ws.cell(row=row, column=3, value=resumen)
            c.font = Font(name=self.FONT_NAME, size=8, color="666666")
            c.alignment = al
            c.fill = fill
            c.border = borde

            c = ws.cell(row=row, column=4, value=art.get("url", ""))
            c.font = fl
            c.alignment = al
            c.fill = fill
            c.border = borde
            try:
                c.hyperlink = art["url"]
            except Exception:
                pass

            c = ws.cell(row=row, column=5, value=art.get("fecha_str", ""))
            c.font = fn
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = fill
            c.border = borde

            c = ws.cell(row=row, column=6, value=", ".join(art.get("tags", []) or []))
            c.font = Font(name=self.FONT_NAME, size=8, italic=True, color="2255AA")
            c.alignment = al
            c.fill = fill
            c.border = borde

            self._aplicar_score(
                ws.cell(row=row, column=7, value=art.get("trend_score", 0)),
                art.get("trend_score", 0),
                borde,
                fill_alto,
                fill_medio,
                fill_bajo,
                font_alto,
                font_medio,
                font_bajo,
            )

            c = ws.cell(row=row, column=8, value=", ".join(art.get("focus_keywords", []) or []))
            c.font = Font(name=self.FONT_NAME, size=8, color="444444")
            c.alignment = al
            c.fill = fill
            c.border = borde

            c = ws.cell(row=row, column=9, value=art.get("seo_title", ""))
            c.font = fn
            c.alignment = al
            c.fill = fill
            c.border = borde

            c = ws.cell(row=row, column=10, value=art.get("meta_description", ""))
            c.font = Font(name=self.FONT_NAME, size=8, color="666666")
            c.alignment = al
            c.fill = fill
            c.border = borde

            c = ws.cell(row=row, column=11, value=art.get("seo_angle", ""))
            c.font = Font(name=self.FONT_NAME, size=8, italic=True, color="7A2E7A")
            c.alignment = al
            c.fill = fill
            c.border = borde

            self._aplicar_score(
                ws.cell(row=row, column=12, value=art.get("evergreen_score", 0)),
                art.get("evergreen_score", 0),
                borde,
                fill_alto,
                fill_medio,
                fill_bajo,
                font_alto,
                font_medio,
                font_bajo,
            )

            c = ws.cell(row=row, column=13, value=(art.get("search_intent", "") or "").upper())
            c.font = Font(name=self.FONT_NAME, size=9, bold=True, color="333333")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = fill
            c.border = borde

            c = ws.cell(row=row, column=14, value=art.get("search_intent_reason", ""))
            c.font = Font(name=self.FONT_NAME, size=8, color="555555")
            c.alignment = al
            c.fill = fill
            c.border = borde

            c = ws.cell(row=row, column=15, value=(art.get("visibility_potential", "") or "").upper())
            c.font = Font(name=self.FONT_NAME, size=9, bold=True, color="333333")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = fill
            c.border = borde

            c = ws.cell(row=row, column=16, value=art.get("target_audience", ""))
            c.font = Font(name=self.FONT_NAME, size=9, color="555555")
            c.alignment = al
            c.fill = fill
            c.border = borde

            ws.row_dimensions[row].height = 22

        widths = {
            "A": 22,
            "B": 55,
            "C": 45,
            "D": 50,
            "E": 18,
            "F": 35,
            "G": 10,
            "H": 38,
            "I": 42,
            "J": 58,
            "K": 30,
            "L": 12,
            "M": 18,
            "N": 32,
            "O": 16,
            "P": 25,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        uf = len(arts) + 1
        if uf > 1:
            ws.auto_filter.ref = f"A1:P{uf}"
        ws.freeze_panes = "A2"

    def _aplicar_score(
        self,
        cell,
        score,
        borde,
        fill_alto,
        fill_medio,
        fill_bajo,
        font_alto,
        font_medio,
        font_bajo,
    ):
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borde
        if score >= 80:
            cell.fill = fill_alto
            cell.font = font_alto
        elif score >= 50:
            cell.fill = fill_medio
            cell.font = font_medio
        else:
            cell.fill = fill_bajo
            cell.font = font_bajo
