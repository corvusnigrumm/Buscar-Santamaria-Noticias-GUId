# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════╗
║   BUSCADOR DE NOTICIAS CAPA BRINDADA — V.6                         ║
║   BUGS CORREGIDOS:                                                   ║
║   BUG 1: Whitelist bloqueaba TODAS las URLs de Google News           ║
║          (news.google.com no estaba en DOMINIOS_PERMITIDOS)          ║
║   BUG 2: Filtro de inglés " the "/" of "/" is " bloqueaba            ║
║          artículos legítimos de Infobae, Forbes, El Heraldo          ║
║   BUG 3: "el_tiempo" con guión bajo generaba falsos positivos        ║
║   BUG 4: "ingresa a nuestro grupo de whatsapp" bloqueaba             ║
║          múltiples fuentes legítimas colombianas                     ║
║   NEW 5: Desplazamiento de fechas: selección [ini, fin]              ║
║          → búsqueda en [fin, fin+(fin-ini)] (período siguiente)      ║
║   NEW 6: 25+ nuevas fuentes (ministerios, FMI, BM, BID,             ║
║          La Nota Económica, Raddar, finanzas personales, etc.)       ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import io
import html
import json
import os
import re
import subprocess
import sys
import time
import logging
import threading
import unicodedata
import hashlib
import urllib.request
import urllib.parse
import xml.etree.ElementTree as ET
from datetime import datetime, date, timedelta, timezone
from difflib import SequenceMatcher
from email.utils import parsedate_to_datetime
from typing import Optional

try:
    import tkinter as tk
    from tkinter import messagebox
    import customtkinter as ctk
    GUI_AVAILABLE = True
except ImportError:
    GUI_AVAILABLE = False
    class DummyCTk:
        def __init__(self, *args, **kwargs):
            pass

    class ctk:
        CTk = DummyCTk
        CTkToplevel = DummyCTk
        CTkFrame = DummyCTk
        CTkButton = DummyCTk
        CTkLabel = DummyCTk
        CTkTextbox = DummyCTk
        CTkCheckBox = DummyCTk
        CTkEntry = DummyCTk
        CTkFont = DummyCTk

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ═══════════════════════════════════════════════════════════════
# CONFIGURACIÓN GENERAL
# ═══════════════════════════════════════════════════════════════

ZONA_COLOMBIA = timezone(timedelta(hours=-5))
BASE_APP_DIR = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, "frozen", False) else __file__))
HISTORIAL_ARTICULOS_PATH = os.path.join(BASE_APP_DIR, "historial_articulos.json")
HISTORIAL_MEDIOS_PROHIBIDOS_PATH = os.path.join(BASE_APP_DIR, "historial_medios_prohibidos.json")
OLLAMA_SIMILITUD_CACHE_PATH = os.path.join(BASE_APP_DIR, "ollama_similitud_cache.json")
MAX_HISTORIAL_ARTICULOS = 4000
MAX_HISTORIAL_MEDIOS_PROHIBIDOS = 12000
VENTANA_LISTA_NEGRA_DIAS = 120
VENTANA_LISTA_NEGRA_DESCARGA_DIAS = 7
MAX_VALIDACIONES_OLLAMA_POR_BUSQUEDA = 3
OLLAMA_TIMEOUT_SEGUNDOS = 12
CONTENIDO_PROFUNDO_MAX_CHARS = 1400

class _DummyStream:
    def write(self, *a, **k): pass
    def flush(self, *a, **k): pass

if sys.stdout is None:
    sys.stdout = _DummyStream()
if sys.stderr is None:
    sys.stderr = _DummyStream()

if sys.platform == 'win32' and hasattr(sys.stdout, 'buffer'):
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass
    os.environ['PYTHONIOENCODING'] = 'utf-8'

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("BuscadorNoticias")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "application/rss+xml, application/xml, text/xml, */*",
    "Accept-Language": "es-CO,es;q=0.9,en;q=0.8",
}

TIMEOUT = 12


def _normalize_domain(url):
    try:
        domain = urllib.parse.urlparse(url).netloc.lower()
    except Exception:
        return ""

    for prefix in ("www.", "m.", "amp."):
        if domain.startswith(prefix):
            domain = domain[len(prefix):]
    return domain

# ═══════════════════════════════════════════════════════════════
# FUENTES RSS — MULTI-CATEGORÍA
# ═══════════════════════════════════════════════════════════════

FUENTES_RSS = [
    # ══════════════════════════════════════════════════════════════
    #  FUENTES NACIONALES (Colombia) — ORIGINALES
    # ══════════════════════════════════════════════════════════════

    {"nombre": "El Heraldo",
     "url": "https://www.elheraldo.co/arc/outboundfeeds/rss/",
     "categorias": ["general", "barranquilla", "judicial", "politica", "deportes", "economia", "cultura", "salud", "tecnologia", "internacional"],
     "tipo": "nacional"},

    {"nombre": "El Colombiano — Medellín",
     "url": "https://www.elcolombiano.com/rss/medellin.xml",
     "categorias": ["general", "medellin"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Antioquia",
     "url": "https://www.elcolombiano.com/rss/antioquia.xml",
     "categorias": ["general", "antioquia"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Colombia",
     "url": "https://www.elcolombiano.com/rss/colombia.xml",
     "categorias": ["general", "politica", "colombia"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Negocios",
     "url": "https://www.elcolombiano.com/rss/negocios.xml",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Deportes",
     "url": "https://www.elcolombiano.com/rss/deportes.xml",
     "categorias": ["deportes"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Internacional",
     "url": "https://www.elcolombiano.com/rss/internacional.xml",
     "categorias": ["internacional"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Cultura",
     "url": "https://www.elcolombiano.com/rss/cultura.xml",
     "categorias": ["cultura"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Tendencias",
     "url": "https://www.elcolombiano.com/rss/tendencias.xml",
     "categorias": ["tendencias", "salud", "vida"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Tecnología",
     "url": "https://www.elcolombiano.com/rss/tecnologia.xml",
     "categorias": ["tecnologia"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Política",
     "url": "https://www.elcolombiano.com/rss/colombia/politica.xml",
     "categorias": ["politica"], "tipo": "nacional"},

    {"nombre": "El Espectador",
     "url": "https://www.elespectador.com/arc/outboundfeeds/discover/?outputType=xml",
     "categorias": ["general", "politica", "judicial", "justicia", "cultura", "colombia"],
     "tipo": "nacional"},
    {"nombre": "El Espectador — Economía",
     "url": "https://www.elespectador.com/arc/outboundfeeds/rss/?outputType=xml&_website=el-espectador&section=/economia",
     "categorias": ["economia"], "tipo": "nacional"},

    {"nombre": "La República",
     "url": "https://www.larepublica.co/rss/",
     "categorias": ["economia", "negocios", "finanzas"],
     "tipo": "nacional"},

    {"nombre": "Semana",
     "url": "https://www.semana.com/arc/outboundfeeds/rss/?outputType=xml",
     "categorias": ["general", "politica", "economia", "cultura", "deportes", "tecnologia", "salud", "internacional", "judicial", "justicia", "vida", "virales"],
     "tipo": "nacional"},

    {"nombre": "La FM — Actualidad",
     "url": "https://www.lafm.com.co/rss/actualidad.xml",
     "categorias": ["general", "politica", "judicial", "justicia", "colombia"],
     "tipo": "nacional"},

    {"nombre": "Caracol Radio",
     "url": "https://caracol.com.co/arc/outboundfeeds/google-news-feed/?outputType=xml",
     "categorias": ["general", "politica", "deportes", "economia", "judicial", "justicia"],
     "tipo": "nacional"},

    {"nombre": "W Radio",
     "url": "https://news.google.com/rss/search?q=site:wradio.com.co&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["general", "politica", "deportes", "economia"],
     "tipo": "nacional"},

    {"nombre": "Pulzo",
     "url": "https://news.google.com/rss/search?q=site:pulzo.com&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["general", "colombia", "virales"], "tipo": "nacional"},

    {"nombre": "El Universal Cartagena",
     "url": "https://news.google.com/rss/search?q=site:eluniversal.com.co&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["general", "colombia"], "tipo": "nacional"},

    {"nombre": "Vanguardia",
     "url": "https://news.google.com/rss/search?q=site:vanguardia.com&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["general", "colombia"], "tipo": "nacional"},

    {"nombre": "Kienyke",
     "url": "https://www.kienyke.com/feed",
     "categorias": ["general", "virales", "cultura"], "tipo": "nacional"},

    {"nombre": "Asuntos Legales",
     "url": "https://www.asuntoslegales.com.co/rss/",
     "categorias": ["justicia", "judicial", "economia"], "tipo": "nacional"},

    {"nombre": "Valora Analitik",
     "url": "https://www.valoraanalitik.com/feed/",
     "categorias": ["economia", "finanzas", "negocios"], "tipo": "nacional"},

    {"nombre": "Forbes Colombia",
     "url": "https://forbes.co/seccion/economia-y-finanzas/",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},

    {"nombre": "Fútbol Red",
     "url": "https://news.google.com/rss/search?q=site:futbolred.com&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["deportes"], "tipo": "nacional"},

    {"nombre": "Conexión Capital",
     "url": "https://news.google.com/rss/search?q=site:canalcapital.gov.co&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["bogota", "general", "colombia"], "tipo": "nacional"},

    {"nombre": "El Carro Colombiano",
     "url": "https://news.google.com/rss/search?q=site:elcarrocolombiano.com&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["motor"], "tipo": "nacional"},

    {"nombre": "Blu Radio",
     "url": "https://news.google.com/rss/search?q=site:bluradio.com&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["general", "colombia", "tecnologia"], "tipo": "nacional"},

    {"nombre": "Radionica",
     "url": "https://news.google.com/rss/search?q=site:radionica.rocks&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["cultura", "vida"], "tipo": "nacional"},

    {"nombre": "Minuto 30",
     "url": "https://www.minuto30.com/feed/",
     "categorias": ["general", "colombia"], "tipo": "nacional"},

    {"nombre": "Infobae Colombia",
     "url": "https://www.infobae.com/arc/outboundfeeds/rss/category/colombia/",
     "categorias": ["general", "colombia", "politica", "judicial", "justicia", "economia", "tecnologia"],
     "tipo": "nacional"},

    {"nombre": "Publimetro",
     "url": "https://www.publimetro.co/arc/outboundfeeds/rss/?outputType=xml",
     "categorias": ["general", "colombia", "virales"], "tipo": "nacional"},

    {"nombre": "RTVC Noticias",
     "url": "https://www.rtvcnoticias.com/rss.xml",
     "categorias": ["general", "colombia", "politica", "cultura", "deportes"],
     "tipo": "nacional"},

    # ──────────────────────────────────────────────────────────────
    #  NUEVAS FUENTES — ECONOMÍA ESPECIALIZADA (solicitadas)
    # ──────────────────────────────────────────────────────────────

    # NEW: La Nota Económica
    {"nombre": "La Nota Económica",
     "url": "https://lanotaeconomica.com.co/feed/",
     "categorias": ["economia", "negocios", "mis finanzas"], "tipo": "nacional"},

    # NEW: Agro Negocios
    {"nombre": "Agro Negocios",
     "url": "https://news.google.com/rss/search?q=site:agronegocios.co&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},

    # NEW: Mi Bolsillo
    {"nombre": "Mi Bolsillo",
     "url": "https://news.google.com/rss/search?q=site:mibolsillo.co+finanzas&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["mis finanzas", "economia"], "tipo": "nacional"},

    # NEW: Expertos del Ahorro
    {"nombre": "Expertos del Ahorro",
     "url": "https://expertosdelahorro.com.co/feed/",
     "categorias": ["mis finanzas"], "tipo": "nacional"},

    # NEW: Mis Finanzas Personales
    {"nombre": "Mis Finanzas Personales",
     "url": "https://misfinanzaspersonales.co/feed/",
     "categorias": ["mis finanzas"], "tipo": "nacional"},

    # NEW: Raddar (consumo y mercado)
    {"nombre": "Raddar",
     "url": "https://raddar.net/feed/",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},

    # NEW: Ahorro Capital
    {"nombre": "Ahorro Capital",
     "url": "https://news.google.com/rss/search?q=site:ahorrocapital.com+finanzas&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["mis finanzas", "economia"], "tipo": "nacional"},

    # NEW: CESLA (vía GNews — sin RSS propio)
    # (se cubre más abajo con Google News)

    # ──────────────────────────────────────────────────────────────
    #  NUEVAS FUENTES — ORGANISMOS INTERNACIONALES (solicitadas)
    # ──────────────────────────────────────────────────────────────

    # NEW: FMI en español
    {"nombre": "FMI — Artículos",
     "url": "https://www.imf.org/es/news/rss",
     "categorias": ["economia", "internacional", "finanzas"], "tipo": "internacional"},

    # NEW: Banco Mundial noticias LAC
    {"nombre": "Banco Mundial — LAC",
     "url": "https://feeds.worldbank.org/world-bank/rss/press-releases/es",
     "categorias": ["economia", "internacional"], "tipo": "internacional"},

    # NEW: BID
    {"nombre": "BID — Mejorando Vidas",
     "url": "https://blogs.iadb.org/feed/",
     "categorias": ["economia", "internacional"], "tipo": "internacional"},

    # ══════════════════════════════════════════════════════════════
    #  FUENTES INTERNACIONALES — ORIGINALES
    # ══════════════════════════════════════════════════════════════

    {"nombre": "Bloomberg Línea",
     "url": "https://www.bloomberglinea.com/arc/outboundfeeds/rss/?outputType=xml",
     "categorias": ["internacional"], "tipo": "internacional"},

    {"nombre": "BBC Mundo",
     "url": "https://feeds.bbci.co.uk/mundo/rss.xml",
     "categorias": ["general", "internacional", "politica", "cultura", "tecnologia", "deportes", "salud", "vida"],
     "tipo": "internacional"},

    {"nombre": "DW Español",
     "url": "https://rss.dw.com/xml/rss-es-all",
     "categorias": ["general", "internacional", "politica", "cultura", "tecnologia", "vida"],
     "tipo": "internacional"},

    {"nombre": "France 24 Español",
     "url": "https://www.france24.com/es/rss",
     "categorias": ["general", "internacional", "politica"], "tipo": "internacional"},

    {"nombre": "CNN Español",
     "url": "https://cnnespanol.cnn.com/feed/",
     "categorias": ["general", "internacional", "politica", "tecnologia", "deportes", "salud", "vida"],
     "tipo": "internacional"},

    {"nombre": "EFE",
     "url": "https://www.efe.com/efe/rss",
     "categorias": ["general", "internacional"], "tipo": "internacional"},

    {"nombre": "Europa Press",
     "url": "https://www.europapress.es/rss/rss.aspx",
     "categorias": ["general", "internacional", "virales"], "tipo": "internacional"},

    {"nombre": "Newsweek",
     "url": "https://www.newsweek.com/rss",
     "categorias": ["general", "internacional", "tecnologia"], "tipo": "internacional"},

    {"nombre": "NPR",
     "url": "https://feeds.npr.org/1001/rss.xml",
     "categorias": ["general", "internacional"], "tipo": "internacional"},

    {"nombre": "The Guardian World",
     "url": "https://www.theguardian.com/world/rss",
     "categorias": ["internacional"], "tipo": "internacional"},

    {"nombre": "NY Times World",
     "url": "https://rss.nytimes.com/services/xml/rss/nyt/World.xml",
     "categorias": ["internacional"], "tipo": "internacional"},

    {"nombre": "El País América",
     "url": "https://feeds.elpais.com/mrss-s/pages/ep/site/elpais.com/portada",
     "categorias": ["general", "internacional", "politica", "cultura"],
     "tipo": "internacional"},

    {"nombre": "RPP Perú",
     "url": "https://rpp.pe/feed",
     "categorias": ["general", "internacional"], "tipo": "internacional"},

    {"nombre": "La Tercera",
     "url": "https://www.latercera.com/feed/",
     "categorias": ["general", "internacional"], "tipo": "internacional"},

    {"nombre": "20 Minutos",
     "url": "https://www.20minutos.es/rss/",
     "categorias": ["general", "internacional", "cultura", "virales"],
     "tipo": "internacional"},

    {"nombre": "HuffPost",
     "url": "https://www.huffpost.com/section/world-news/feed",
     "categorias": ["general", "internacional"], "tipo": "internacional"},

    {"nombre": "ESPN Deportes",
     "url": "https://espndeportes.espn.com/espn/rss/news",
     "categorias": ["deportes"], "tipo": "internacional"},

    {"nombre": "Mundo Deportivo",
     "url": "https://www.mundodeportivo.com/rss/",
     "categorias": ["deportes"], "tipo": "internacional"},

    {"nombre": "Rolling Stone",
     "url": "https://www.rollingstone.com/feed/",
     "categorias": ["cultura"], "tipo": "internacional"},

    {"nombre": "Billboard",
     "url": "https://www.billboard.com/feed/",
     "categorias": ["cultura"], "tipo": "internacional"},

    {"nombre": "Muy Interesante",
     "url": "https://www.muyinteresante.es/feed/",
     "categorias": ["vida", "salud", "tecnologia"], "tipo": "internacional"},

    {"nombre": "National Geographic LA",
     "url": "https://www.nationalgeographicla.com/feed",
     "categorias": ["vida", "cultura"], "tipo": "internacional"},

    {"nombre": "Digital Trends Español",
     "url": "https://es.digitaltrends.com/feed/",
     "categorias": ["tecnologia"], "tipo": "internacional"},

    {"nombre": "Andro4all",
     "url": "https://andro4all.com/feed",
     "categorias": ["tecnologia"], "tipo": "internacional"},

    {"nombre": "Car and Driver",
     "url": "https://www.caranddriver.com/es/rss/",
     "categorias": ["motor"], "tipo": "internacional"},

    {"nombre": "Autobild",
     "url": "https://www.autobild.es/rss/",
     "categorias": ["motor"], "tipo": "internacional"},

    # ══════════════════════════════════════════════════════════════
    #  GOOGLE NEWS — CATEGORÍAS GENERALES (Colombia)
    # ══════════════════════════════════════════════════════════════

    {"nombre": "Google News — Colombia",
     "url": "https://news.google.com/rss/search?q=colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["general", "colombia"], "tipo": "nacional"},
    {"nombre": "Google News — Economía",
     "url": "https://news.google.com/rss/search?q=economia+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia"], "tipo": "nacional"},
    {"nombre": "Google News — Deportes",
     "url": "https://news.google.com/rss/search?q=deportes+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["deportes"], "tipo": "nacional"},
    {"nombre": "Google News — Política",
     "url": "https://news.google.com/rss/search?q=politica+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["politica"], "tipo": "nacional"},
    {"nombre": "Google News — Tecnología",
     "url": "https://news.google.com/rss/search?q=tecnologia+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["tecnologia"], "tipo": "nacional"},
    {"nombre": "Google News — Salud",
     "url": "https://news.google.com/rss/search?q=salud+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["salud"], "tipo": "nacional"},
    {"nombre": "Google News — Cultura",
     "url": "https://news.google.com/rss/search?q=cultura+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["cultura"], "tipo": "nacional"},
    {"nombre": "Google News — Judicial",
     "url": "https://news.google.com/rss/search?q=judicial+justicia+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["judicial", "justicia"], "tipo": "nacional"},
    {"nombre": "Google News — Motor Colombia",
     "url": "https://news.google.com/rss/search?q=carros+autos+motor+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["motor"], "tipo": "nacional"},
    {"nombre": "Google News — Bogotá",
     "url": "https://news.google.com/rss/search?q=bogota+noticias&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["bogota"], "tipo": "nacional"},
    {"nombre": "Google News — Vida Colombia",
     "url": "https://news.google.com/rss/search?q=ciencia+medio+ambiente+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["vida"], "tipo": "nacional"},
    {"nombre": "Google News — Virales",
     "url": "https://news.google.com/rss/search?q=viral+tendencia+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["virales"], "tipo": "nacional"},
    {"nombre": "Google News - Tendencias Colombia",
     "url": "https://news.google.com/rss/search?q=tendencias+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["tendencias"], "tipo": "nacional"},
    {"nombre": "Google News - Tendencias Lifestyle",
     "url": "https://news.google.com/rss/search?q=tendencia+redes+sociales+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["tendencias"], "tipo": "nacional"},
    {"nombre": "Google News - Tendencias Globales",
     "url": "https://news.google.com/rss/search?q=viral+tendencia+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["tendencias"], "tipo": "internacional"},
    {"nombre": "Google News — Mis Finanzas (Servicios)",
     "url": "https://news.google.com/rss/search?q=servicio+de+la+gente+OR+como+pagar+OR+plazo+para+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["mis finanzas"], "tipo": "nacional"},
    {"nombre": "Google News — Mis Finanzas (Trámites/Subsidios)",
     "url": "https://news.google.com/rss/search?q=devolucion+del+iva+OR+renta+ciudadana+OR+subsidio+OR+dian+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["mis finanzas"], "tipo": "nacional"},

    # ──────────────────────────────────────────────────────────────
    #  NUEVAS CONSULTAS GOOGLE NEWS — MINISTERIOS Y GOB. (solicitadas)
    # ──────────────────────────────────────────────────────────────

    # NEW: DIAN
    {"nombre": "Google News — DIAN",
     "url": "https://news.google.com/rss/search?q=dian+colombia+impuesto+tributario&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["mis finanzas", "economia"], "tipo": "nacional"},

    # NEW: Ministerio de Agricultura
    {"nombre": "Google News — Ministerio Agricultura",
     "url": "https://news.google.com/rss/search?q=ministerio+agricultura+colombia+campo&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia"], "tipo": "nacional"},

    # NEW: Ministerio de Trabajo
    {"nombre": "Google News — Ministerio Trabajo",
     "url": "https://news.google.com/rss/search?q=ministerio+trabajo+colombia+empleo+laboral&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "general"], "tipo": "nacional"},

    # NEW: Ministerio de Energía
    {"nombre": "Google News — MinEnergía",
     "url": "https://news.google.com/rss/search?q=ministerio+minas+energia+colombia+petroleo+gas&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia"], "tipo": "nacional"},

    # NEW: Ministerio de Comercio (MinCIT)
    {"nombre": "Google News — MinCIT",
     "url": "https://news.google.com/rss/search?q=mincit+ministerio+comercio+industria+turismo+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},

    # NEW: Ministerio de Vivienda
    {"nombre": "Google News — MinVivienda",
     "url": "https://news.google.com/rss/search?q=minvivienda+ministerio+vivienda+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["general", "economia"], "tipo": "nacional"},

    # NEW: Banco de la República
    {"nombre": "Google News — BanRep",
     "url": "https://news.google.com/rss/search?q=%22banco+de+la+republica%22+colombia+tasas+inflacion&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "finanzas"], "tipo": "nacional"},

    # NEW: FMI / Colombia (vía GNews también)
    {"nombre": "Google News — FMI Colombia",
     "url": "https://news.google.com/rss/search?q=fmi+colombia+economia+perspectivas&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "internacional"], "tipo": "nacional"},

    # NEW: Banco Mundial Colombia
    {"nombre": "Google News — Banco Mundial Colombia",
     "url": "https://news.google.com/rss/search?q=%22banco+mundial%22+colombia+informe&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "internacional"], "tipo": "nacional"},

    # NEW: BID Colombia
    {"nombre": "Google News — BID Colombia",
     "url": "https://news.google.com/rss/search?q=%22banco+interamericano%22+bid+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "internacional"], "tipo": "nacional"},

    # NEW: La Nota Económica (vía GNews — por si el RSS falla)
    {"nombre": "Google News — La Nota Econ.",
     "url": "https://news.google.com/rss/search?q=site:lanotaeconomica.com.co&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},

    # NEW: Valora Analitik (vía GNews — respaldo)
    {"nombre": "Google News — Valora Analitik",
     "url": "https://news.google.com/rss/search?q=site:valoraanalitik.com&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "finanzas", "negocios"], "tipo": "nacional"},

    # NEW: Agronegocios Colombia
    {"nombre": "Google News — Agronegocios",
     "url": "https://news.google.com/rss/search?q=agronegocios+colombia+campo+agro&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia"], "tipo": "nacional"},

    # NEW: Cámara de Comercio Bogotá
    {"nombre": "Google News — CCB",
     "url": "https://news.google.com/rss/search?q=%22camara+de+comercio%22+bogota+empresa&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},
]

# ═══════════════════════════════════════════════════════════════
# DOMINIOS ARGENTINOS — para filtro opcional
# ═══════════════════════════════════════════════════════════════

DOMINIOS_ARGENTINA = [
    "clarin.com", "lanacion.com.ar", "perfil.com",
    "lavoz.com.ar", "ole.com.ar", "mdzol.com",
    "la100.cienradios.com", "elle.clarin.com",
]

# ═══════════════════════════════════════════════════════════════
# DOMINIOS BLOQUEADOS
# ═══════════════════════════════════════════════════════════════

import pandas as pd

_DIR_ACTUAL = os.path.dirname(os.path.abspath(__file__))
_EXCEL_FUENTES_PATH = os.path.join(_DIR_ACTUAL, 'Fuentes_SEO.xlsx')

DOMINIOS_PERMITIDOS = set()
try:
    _dfs = pd.read_excel(
        _EXCEL_FUENTES_PATH,
        sheet_name=None, header=None
    )
    for sheet, _df in _dfs.items():
        for r in range(_df.shape[0]):
            for c in range(_df.shape[1]):
                val = str(_df.iloc[r, c]).strip()
                if "http" in val:
                    for part in val.split():
                        if "http" in part:
                            try:
                                netloc = _normalize_domain(part)
                                DOMINIOS_PERMITIDOS.add(netloc)
                            except: pass
    if "news.google.com" in DOMINIOS_PERMITIDOS:
        DOMINIOS_PERMITIDOS.remove("news.google.com")
except Exception as e:
    DOMINIOS_PERMITIDOS = set()
    log.warning(f"Fuentes_SEO.xlsx no disponible ({e}) — usando fallback de FUENTES_RSS")

# FALLBACK: construir whitelist desde FUENTES_RSS + nuevas fuentes solicitadas
if len(DOMINIOS_PERMITIDOS) < 10:
    for f in FUENTES_RSS:
        try:
            nl = _normalize_domain(f["url"])
            if "news.google" not in nl:
                DOMINIOS_PERMITIDOS.add(nl)
        except: pass

    # NEW: agregar explícitamente los dominios nuevos solicitados
    _dominios_extra = [
        "lanotaeconomica.com.co", "raddar.net", "agronegocios.co",
        "mibolsillo.com", "valoraanalitik.com", "expertosdelahorro.com.co",
        "misfinanzaspersonales.co", "ahorrocapital.com", "cesla.com",
        "dian.gov.co", "minagricultura.gov.co", "mintrabajo.gov.co",
        "minenergia.gov.co", "mincit.gov.co", "minvivienda.gov.co",
        "iadb.org", "bancomundial.org", "imf.org",
        "tuinterfaz.mx", "paat.mx",
    ]
    for d in _dominios_extra:
        DOMINIOS_PERMITIDOS.add(d)

DOMINIOS_BLOQUEADOS = [
    "portafolio.co", "eltiempo.com",
    "blogs.portafolio.co", "amp.portafolio.co", "m.portafolio.co",
    "amp.eltiempo.com", "m.eltiempo.com", "especiales.eltiempo.com",
    "enter.co", "citytv.com.co", "cronista.com",
]

MEDIOS_PROHIBIDOS = {
    "el_tiempo": {
        "label": "El Tiempo",
        "dominios": [
            "eltiempo.com", "m.eltiempo.com", "amp.eltiempo.com",
            "especiales.eltiempo.com",
        ],
        "source_aliases": [
            "el tiempo", "eltiempo.com", "el tiempo play",
        ],
        "title_aliases": [
            "el tiempo", "eltiempo.com", "el tiempo play",
        ],
        "signatures": [
            "casa editorial el tiempo",
            "el tiempo play",
            "el tiempoplay",
            "sigue a el tiempo en whatsapp",
            "suscribete a el tiempo",
            "suscríbete a el tiempo",
            "fuente: el tiempo",
        ],
        "lista_negra_url": "https://news.google.com/rss/search?q=site:eltiempo.com&hl=es-419&gl=CO&ceid=CO:es-419",
    },
    "portafolio": {
        "label": "Portafolio",
        "dominios": [
            "portafolio.co", "blogs.portafolio.co", "amp.portafolio.co",
            "m.portafolio.co",
        ],
        "source_aliases": [
            "portafolio", "portafolio.co",
        ],
        "title_aliases": [
            "portafolio", "portafolio.co",
        ],
        "signatures": [
            "portafolio digital",
            "suscribete a portafolio",
            "suscríbete a portafolio",
            "portafolio.co",
        ],
        "lista_negra_url": "https://news.google.com/rss/search?q=site:portafolio.co&hl=es-419&gl=CO&ceid=CO:es-419",
    },
    "citytv": {
        "label": "CityTV",
        "dominios": [
            "citytv.com.co",
        ],
        "source_aliases": [
            "citytv", "city tv", "citytv bogota",
        ],
        "title_aliases": [
            "citytv", "city tv",
        ],
        "signatures": [
            "citynoticias",
        ],
        "lista_negra_url": "https://news.google.com/rss/search?q=site:citytv.com.co&hl=es-419&gl=CO&ceid=CO:es-419",
    },
    "cronista": {
        "label": "El Cronista",
        "dominios": [
            "cronista.com", "www.cronista.com",
        ],
        "source_aliases": [
            "el cronista", "cronista", "cronista.com",
        ],
        "title_aliases": [
            "el cronista", "cronista", "cronista.com",
        ],
        "signatures": [
            "cronista.com",
            "el cronista",
            "cronista mexico",
        ],
        "lista_negra_url": "https://news.google.com/rss/search?q=site:cronista.com&hl=es-419&gl=CO&ceid=CO:es-419",
    },
}

LISTA_NEGRA_MEDIOS = []
LISTA_NEGRA_EL_TIEMPO = LISTA_NEGRA_MEDIOS
LISTA_NEGRA_MEDIOS_HASHES = set()
LISTA_NEGRA_MEDIOS_POR_ANCLA = {}
OLLAMA_SIMILITUD_CACHE = None
OLLAMA_CLI_PATH = None
OLLAMA_VALIDACIONES_RESTANTES = 0
CONTENIDO_ARTICULOS_CACHE = {}
CONTENIDO_ARTICULOS_LOCK = threading.Lock()

CATEGORIAS_RELACIONADAS = {
    "vida": {"vida", "salud", "cultura", "tendencias", "virales"},
    "salud": {"salud", "vida", "tendencias"},
    "negocios": {"negocios", "economia", "finanzas", "mis finanzas"},
    "finanzas": {"finanzas", "mis finanzas", "economia", "negocios"},
    "mis finanzas": {"mis finanzas", "finanzas", "economia"},
    "tecnologia": {"tecnologia", "tendencias"},
}

MARCADORES_COLOMBIA = (
    "colombia", "colombiano", "colombiana", "colombianos", "colombianas",
    "bogota", "medellin", "cali", "barranquilla", "cartagena", "bucaramanga",
    "cucuta", "pereira", "manizales", "armenia", "ibague", "villavicencio",
    "neiva", "pasto", "popayan", "monteria", "sincelejo", "riohacha",
    "valledupar", "santa marta", "tunja", "soacha", "bello", "itagui",
    "envigado", "rionegro", "apartado", "antioquia", "atlantico", "bolivar",
    "boyaca", "caldas", "caqueta", "cauca", "cesar", "choco", "cordoba",
    "cundinamarca", "guainia", "guaviare", "huila", "guajira", "magdalena",
    "meta", "narino", "nariño", "norte de santander", "quindio", "quindío",
    "risaralda", "san andres", "san andrés", "santander", "sucre", "tolima",
    "valle del cauca", "vaupes", "vaupés", "vichada", "dian", "minsalud",
    "mineducacion", "mineducación", "minvivienda", "mintrabajo", "mincit",
    "minenergia", "minenergía", "fiscalia", "fiscalía", "procuraduria",
    "procuraduría", "contraloria", "contraloría", "registraduria",
    "registraduría", "superfinanciera", "banco de la republica",
    "banco de la república", "casa de narino", "casa de nariño",
    "alcaldia de bogota", "alcaldía de bogotá", "gobierno nacional",
)

MARCADORES_EXTRANJERO = (
    "estados unidos", "ee uu", "ee. uu", "eeuu", "usa", "u s a",
    "washington", "nueva york", "new york", "los angeles", "texas",
    "peru", "perú", "ecuador", "mexico", "méxico", "argentina", "chile",
    "brasil", "brasilia", "brazil", "venezuela", "panama", "panamá",
    "uruguay", "paraguay", "bolivia", "españa", "espana", "francia",
    "alemania", "italia", "reino unido", "londres", "china", "rusia",
    "ucrania", "israel", "gaza", "canada", "canadá", "japon", "japón",
)

MARCADORES_FUENTE_LOCAL = {
    "colombia", "bogota", "medellin", "medellín", "antioquia",
    "barranquilla", "cali",
}


def _expandir_categorias_solicitadas(categorias):
    if not categorias:
        return set()
    expandidas = set(categorias)
    for cat in list(expandidas):
        expandidas.update(CATEGORIAS_RELACIONADAS.get(cat, set()))
    return expandidas


def _resolver_categoria_solicitada(categorias_fuente, categorias_solicitadas):
    if not categorias_solicitadas:
        return None
    fuente_set = set(categorias_fuente or [])
    for cat in categorias_solicitadas:
        if cat in fuente_set:
            return cat
        relacionadas = CATEGORIAS_RELACIONADAS.get(cat, set())
        if relacionadas and fuente_set.intersection(relacionadas):
            return cat
    return None


PALABRAS_TENDENCIA_VALIDAS = (
    "tendencia", "tendencias", "viral", "virales", "redes", "sociales",
    "entretenimiento", "famosos", "streaming", "moda", "belleza",
    "lifestyle", "hogar", "bienestar", "apps", "tecnologia",
)

PALABRAS_TENDENCIA_EXCLUIDAS = (
    "dolar", "euro", "trm", "cotizacion", "apertura", "inflacion", "tasas",
    "petro", "elecciones", "votacion", "congreso", "judicial", "captur",
    "asesin", "accidente", "balacera", "partido", "liga", "vs ",
    "previa", "marcador", "gol", "clima", "temperaturas",
)

PALABRAS_RUIDO_GENERAL = (
    "partido", "liga", "vs ", "previa", "marcador", "gol", "penal",
    "captur", "captura", "asesin", "balacera", "fiscalia", "juez",
    "tribunal", "accidente", "choque", "hurto",
)

CATEGORIA_REGLAS = {
    "salud": {
        "include": ("salud", "medico", "medica", "medicina", "hospital", "clinica", "vacuna", "tratamiento", "sintomas",
                    "bienestar", "nutricion", "ejercicio", "salud mental", "psicologia", "cancer", "diabetes",
                    "obesidad", "paciente", "pacientes", "sindrome", "enfermedad", "enfermedades", "virus", "epidemia"),
        "exclude": PALABRAS_RUIDO_GENERAL + ("dolar", "euro", "gasolina", "supermercados"),
    },
    "vida": {
        "include": ("vida", "hogar", "familia", "pareja", "viaje", "viajes", "vacaciones", "mascota", "mascotas",
                    "cocina", "receta", "descanso", "hotel", "estilo de vida", "habitos", "bienestar", "convivencia"),
        "exclude": PALABRAS_RUIDO_GENERAL + ("dolar", "euro", "inflacion", "elecciones"),
    },
    "tendencias": {
        "include": ("tendencia", "tendencias", "viral", "virales", "redes", "streaming", "famos", "moda",
                    "belleza", "lifestyle", "hogar", "consumo", "tecnologia", "app", "apps", "entretenimiento"),
        "exclude": PALABRAS_TENDENCIA_EXCLUIDAS,
    },
    "negocios": {
        "include": ("empresa", "empresas", "negocio", "negocios", "mercado", "industria", "comercio", "startup",
                    "startups", "emprendimiento", "inversion", "alianza", "ventas", "consumo", "supermercados",
                    "clientes", "empresarial"),
        "exclude": PALABRAS_RUIDO_GENERAL,
    },
    "finanzas": {
        "include": ("finanzas", "banco", "banca", "credito", "deuda", "ahorro", "inversion", "bolsa",
                    "dolar", "euro", "inflacion", "tasa", "tasas", "trm", "divisa", "cotizacion",
                    "mercado", "hacienda", "tributaria", "tributario"),
        "exclude": PALABRAS_RUIDO_GENERAL,
    },
    "mis finanzas": {
        "include": ("finanzas personales", "ahorro", "ahorrar", "tarjeta", "credito", "subsidio", "dian", "impuesto",
                    "renta", "pension", "cesant", "presupuesto", "factura", "devolucion", "iva", "bolsillo", "pagar"),
        "exclude": PALABRAS_RUIDO_GENERAL,
    },
    "tecnologia": {
        "include": ("tecnologia", "ia", "inteligencia artificial", "app", "apps", "celular", "iphone", "android",
                    "samsung", "internet", "starlink", "robotica", "software", "digital", "startup", "ciberseguridad", "datos"),
        "exclude": PALABRAS_RUIDO_GENERAL,
    },
    "cultura": {
        "include": ("cultura", "cine", "musica", "libro", "libros", "teatro", "arte", "artista", "festival",
                    "museo", "concierto", "pelicula", "peliculas", "serie", "series", "literatura", "danza", "netflix"),
        "exclude": PALABRAS_RUIDO_GENERAL + ("dolar", "inflacion", "gasolina"),
    },
}


def _texto_categoria_norm(texto):
    texto = unicodedata.normalize("NFKD", texto or "").encode("ascii", "ignore").decode("ascii")
    return _normalizar_texto_medio(texto)


def _contar_patrones_texto(texto, patrones):
    return sum(1 for patron in patrones if _texto_contiene_patron(texto, patron))


def _articulo_es_nacional_colombia(articulo, fuente):
    texto = _texto_categoria_norm(
        f"{articulo.get('titulo', '')} {articulo.get('resumen', '')} "
        f"{articulo.get('url', '')} {articulo.get('fuente', '')}"
    )
    if not texto:
        return False

    score_colombia_texto = _contar_patrones_texto(texto, MARCADORES_COLOMBIA)
    score_extranjero = _contar_patrones_texto(texto, MARCADORES_EXTRANJERO)

    if score_colombia_texto == 0:
        return False

    score_colombia = score_colombia_texto

    categorias_fuente = set(fuente.get("categorias", []))
    if categorias_fuente.intersection(MARCADORES_FUENTE_LOCAL):
        score_colombia += 1

    fuente_norm = _texto_categoria_norm(fuente.get("nombre", ""))
    if any(_texto_contiene_patron(fuente_norm, marcador) for marcador in MARCADORES_FUENTE_LOCAL):
        score_colombia += 1

    if score_extranjero >= 2 and score_colombia <= 1:
        return False
    if score_extranjero > score_colombia + 1:
        return False
    return True


def _parece_ingles_filtro_internacional(titulo, descripcion):
    texto = _normalizar_para_repeticion(f"{titulo} {descripcion}")
    if len(texto) < 35:
        return False

    tokens = set(texto.split())
    english_tokens = {
        "the", "and", "with", "from", "after", "about", "officials",
        "warn", "warns", "warning", "latest", "report", "reports",
        "market", "markets", "global", "health", "social", "mental",
        "risk", "risks", "said", "says", "outlook", "government",
        "economy", "economic", "official", "officially", "city",
    }
    spanish_tokens = {
        "el", "la", "los", "las", "de", "del", "y", "para", "con",
        "por", "una", "uno", "unos", "unas", "salud", "riesgos",
        "bogota", "colombia", "alcaldia", "gobierno", "mental",
    }
    hits_en = len(tokens.intersection(english_tokens))
    hits_es = len(tokens.intersection(spanish_tokens))
    return hits_en >= 3 and hits_es <= 1


def _articulo_cumple_filtro_internacional(articulo):
    return not (
        _parece_ingles_puro(
            articulo.get("titulo", ""),
            articulo.get("resumen", ""),
        ) or _parece_ingles_filtro_internacional(
            articulo.get("titulo", ""),
            articulo.get("resumen", ""),
        )
    )


def _texto_contiene_patron(texto, patron):
    patron = _texto_categoria_norm(patron)
    if not patron:
        return False
    if " " in patron:
        return patron in texto
    return re.search(r"\b" + re.escape(patron) + r"\b", texto) is not None


def _articulo_coincide_categoria(categoria, titulo="", resumen="", fuente="", categorias_fuente=None):
    categorias_fuente = set(categorias_fuente or [])
    texto = _texto_categoria_norm(f"{titulo} {resumen} {fuente}")
    if not texto:
        return False

    regla = CATEGORIA_REGLAS.get(categoria)
    if not regla:
        return categoria in categorias_fuente

    include = regla.get("include", ())
    exclude = regla.get("exclude", ())
    include_hits = sum(1 for palabra in include if _texto_contiene_patron(texto, palabra))
    exclude_hits = sum(1 for palabra in exclude if _texto_contiene_patron(texto, palabra))

    if categoria in categorias_fuente and include_hits >= 1 and exclude_hits == 0:
        return True

    if include_hits >= 2 and exclude_hits == 0:
        return True

    if include_hits >= 1 and categoria in categorias_fuente and exclude_hits <= 1:
        return True

    return False


def _es_tendencia_valida(titulo="", resumen="", fuente=""):
    texto = unicodedata.normalize("NFKD", f"{titulo} {resumen} {fuente}").encode("ascii", "ignore").decode("ascii")
    texto = _normalizar_texto_medio(texto)
    if not texto:
        return False
    if any(p in texto for p in PALABRAS_TENDENCIA_EXCLUIDAS):
        return False
    if any(p in texto for p in PALABRAS_TENDENCIA_VALIDAS):
        return True
    fuente_norm = unicodedata.normalize("NFKD", fuente).encode("ascii", "ignore").decode("ascii")
    return "tendencias" in _normalizar_texto_medio(fuente_norm)

def _extraer_palabras_clave(texto):
    if not texto: return set()
    cifras = set(re.findall(r'\b\d+(?:[\.,]\d+)?\b', texto))
    palabras = set(re.findall(r'\b[a-záéíóúñ]{7,}\b', texto.lower()))
    comunes = {"colombia", "nacional", "gobierno", "general", "informes", "durante", "noticias", "presenta", "también", "nuestro", "después", "algunos"}
    return cifras.union(palabras - comunes)

def _normalizar_texto_medio(texto):
    if not texto:
        return ""
    return re.sub(r"[^a-z0-9]+", " ", texto.lower()).strip()


def _extraer_tokens_relevantes(texto):
    if not texto:
        return set()
    tokens = set(re.findall(r"\b[a-zÃ¡Ã©Ã­Ã³ÃºÃ±]{5,}\b", texto.lower()))
    stop = {
        "colombia", "noticias", "general", "mundo", "ultima", "ultimas",
        "sobre", "desde", "hasta", "entre", "contra", "porque", "donde",
        "cuando", "todos", "todas", "tras", "segun", "nuevo", "nueva",
        "este", "esta", "estos", "estas",
    }
    return tokens - stop


def _normalizar_para_repeticion(texto):
    texto = _limpiar_html(texto or "")
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    texto = texto.lower()
    texto = re.sub(r"https?://\S+", " ", texto)
    texto = re.sub(r"[^a-z0-9\s]+", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def _raiz_simple_token(token):
    token = (token or "").strip().lower()
    sufijos = (
        "amientos", "imiento", "imientos", "aciones", "adoras", "adores",
        "adoras", "adora", "adores", "antes", "ancia", "encias", "encia",
        "idades", "idad", "mente", "ciones", "cion", "siones", "sion",
        "arios", "arias", "ario", "aria", "logias", "logia", "ismos",
        "istas", "ismos", "istas", "ables", "ible", "ibles", "ables",
        "anzas", "anza", "icos", "icas", "ico", "ica", "ales", "ados",
        "adas", "idos", "idas", "ando", "iendo", "oras", "ores", "osa",
        "oso", "ivas", "ivos", "iva", "ivo", "es", "s",
    )
    for sufijo in sufijos:
        if len(token) - len(sufijo) >= 5 and token.endswith(sufijo):
            return token[:-len(sufijo)]
    return token


def _extraer_tokens_repeticion(texto):
    texto = _normalizar_para_repeticion(texto)
    tokens = {
        _raiz_simple_token(token)
        for token in re.findall(r"\b[a-z0-9]{5,}\b", texto)
    }
    stop = {
        "colombia", "noticias", "general", "ultima", "ultimas", "mundo",
        "fuente", "video", "fotos", "tras", "sobre", "desde", "hasta",
        "entre", "porque", "nuevo", "nueva", "este", "esta", "estos",
        "estas", "hoy", "ayer",
    }
    stop_raices = {_raiz_simple_token(token) for token in stop}
    return {token for token in tokens if token and token not in stop_raices}


def _tokens_repeticion_desde_url(url):
    if not url:
        return []
    try:
        parsed = urllib.parse.urlparse(url)
    except Exception:
        return []
    slug = urllib.parse.unquote(parsed.path or "")
    slug = slug.replace("-", " ").replace("_", " ").replace("/", " ")
    slug = _normalizar_para_repeticion(slug)
    tokens = []
    for token in slug.split():
        if len(token) < 5 or token.isdigit():
            continue
        raiz = _raiz_simple_token(token)
        if raiz:
            tokens.append(raiz)
    return tokens


def _extraer_frases_repeticion(titulo, url=""):
    texto = _normalizar_para_repeticion(titulo)
    tokens_titulo = [_raiz_simple_token(t) for t in texto.split() if len(t) >= 4 and not t.isdigit()]
    tokens_slug = _tokens_repeticion_desde_url(url)
    stop = {
        "colombia", "santander", "bogota", "medellin", "cartagena", "semana",
        "santa", "tiempo", "portafolio", "citytv", "visit", "visitant",
        "articul", "notici", "colomb", "ubic", "hor", "debe", "saber",
        "recorr", "miles", "mismo", "mism", "larg", "largo",
    }

    frases = []
    for fuente_tokens in (tokens_titulo, tokens_slug):
        n_tokens = len(fuente_tokens)
        for size in (5, 4, 3):
            if n_tokens < size:
                continue
            for i in range(0, n_tokens - size + 1):
                ventana = fuente_tokens[i:i + size]
                if sum(1 for t in ventana if t and t not in stop) < max(2, size - 1):
                    continue
                frase = " ".join(ventana).strip()
                if len(frase) >= 18:
                    frases.append(frase)
    frases = sorted(set(frases), key=lambda x: (-len(x), x))
    return frases[:10]


def _seleccionar_claves_repeticion(titulo, resumen_limpio, url, tokens_repeticion):
    titulo_tokens = set(_extraer_tokens_repeticion(titulo))
    slug_tokens = set(_tokens_repeticion_desde_url(url))
    resumen_tokens = set(_extraer_tokens_repeticion(resumen_limpio))
    tokens_base = set(tokens_repeticion or set())

    score = {}
    for token in tokens_base:
        valor = len(token)
        if token in titulo_tokens:
            valor += 5
        if token in slug_tokens:
            valor += 4
        if token in resumen_tokens:
            valor += 2
        if token in MARCADORES_COLOMBIA or token in MARCADORES_EXTRANJERO:
            valor -= 2
        score[token] = valor

    claves = sorted(tokens_base, key=lambda t: (-score.get(t, 0), -len(t), t))
    return claves[:12]


def _construir_huella_repeticion(titulo, resumen, url=""):
    resumen_limpio = _limpiar_html(resumen) or titulo or ""
    titulo_norm = _normalizar_para_repeticion(titulo)[:180]
    resumen_norm = _normalizar_para_repeticion(resumen_limpio)[:420]
    slug_tokens = _tokens_repeticion_desde_url(url)
    slug_norm = " ".join(slug_tokens)[:180]
    texto_repeticion = f"{titulo_norm} {resumen_norm} {slug_norm}".strip()
    tokens_repeticion = _extraer_tokens_repeticion(f"{titulo} {resumen_limpio} {slug_norm}")
    claves_repeticion = _seleccionar_claves_repeticion(titulo, resumen_limpio, url, tokens_repeticion)
    firma_tokens_repeticion = " ".join(sorted(tokens_repeticion))[:420]
    firma_claves_repeticion = " ".join(claves_repeticion)[:240]
    frases_repeticion = _extraer_frases_repeticion(titulo, url)
    anclas = sorted(tokens_repeticion, key=lambda t: (-len(t), t))[:6]
    hash_repeticion = hashlib.sha1(texto_repeticion.encode("utf-8", errors="ignore")).hexdigest()
    return {
        "resumen_limpio": resumen_limpio,
        "texto_repeticion": texto_repeticion,
        "tokens_repeticion": tokens_repeticion,
        "claves_repeticion": claves_repeticion,
        "slug_repeticion": slug_norm,
        "firma_tokens_repeticion": firma_tokens_repeticion,
        "firma_claves_repeticion": firma_claves_repeticion,
        "frases_repeticion": frases_repeticion,
        "anclas_repeticion": anclas,
        "hash_repeticion": hash_repeticion,
    }


def _registro_historial_desde_articulo(articulo):
    return {
        "titulo": articulo.get("titulo", ""),
        "fuente": articulo.get("fuente", ""),
        "url": articulo.get("url", ""),
        "fecha": articulo.get("fecha_str", ""),
        "t_norm": articulo.get("t_norm", ""),
        "texto_repeticion": articulo.get("texto_repeticion", ""),
        "hash_repeticion": articulo.get("hash_repeticion", ""),
        "tokens_repeticion": sorted(articulo.get("tokens_repeticion", set())),
        "claves_repeticion": list(articulo.get("claves_repeticion", [])),
        "slug_repeticion": articulo.get("slug_repeticion", ""),
        "firma_tokens_repeticion": articulo.get("firma_tokens_repeticion", ""),
        "firma_claves_repeticion": articulo.get("firma_claves_repeticion", ""),
        "frases_repeticion": list(articulo.get("frases_repeticion", [])),
        "anclas_repeticion": list(articulo.get("anclas_repeticion", [])),
    }


def _cargar_historial_articulos():
    if not os.path.exists(HISTORIAL_ARTICULOS_PATH):
        return []
    try:
        with open(HISTORIAL_ARTICULOS_PATH, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        if not isinstance(data, list):
            return []
        historial = []
        for item in data:
            if not isinstance(item, dict) or not item.get("hash_repeticion"):
                continue
            item["tokens_repeticion"] = set(item.get("tokens_repeticion", []))
            item["claves_repeticion"] = list(item.get("claves_repeticion", []))
            item["frases_repeticion"] = list(item.get("frases_repeticion", []))
            item["anclas_repeticion"] = list(item.get("anclas_repeticion", []))
            historial.append(item)
        return historial
    except Exception:
        return []


def _cargar_historial_medios_prohibidos():
    if not os.path.exists(HISTORIAL_MEDIOS_PROHIBIDOS_PATH):
        return []
    try:
        with open(HISTORIAL_MEDIOS_PROHIBIDOS_PATH, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        if not isinstance(data, list):
            return []
        historial = []
        for item in data:
            if not isinstance(item, dict) or not item.get("hash_repeticion"):
                continue
            item["tokens_repeticion"] = set(item.get("tokens_repeticion", []))
            item["claves_repeticion"] = list(item.get("claves_repeticion", []))
            item["frases_repeticion"] = list(item.get("frases_repeticion", []))
            item["anclas_repeticion"] = list(item.get("anclas_repeticion", []))
            historial.append(item)
        return historial
    except Exception:
        return []


def _guardar_historial_articulos(historial):
    try:
        serializable = []
        for item in historial[-MAX_HISTORIAL_ARTICULOS:]:
            row = dict(item)
            row["tokens_repeticion"] = sorted(row.get("tokens_repeticion", set()))
            row["claves_repeticion"] = list(row.get("claves_repeticion", []))
            row["frases_repeticion"] = list(row.get("frases_repeticion", []))
            row["anclas_repeticion"] = list(row.get("anclas_repeticion", []))
            serializable.append(row)
        with open(HISTORIAL_ARTICULOS_PATH, "w", encoding="utf-8") as fh:
            json.dump(serializable, fh, ensure_ascii=False, indent=2)
    except Exception as exc:
        log.warning(f"No se pudo guardar historial de articulos: {exc}")


def _guardar_historial_medios_prohibidos(historial):
    try:
        serializable = []
        for item in historial[-MAX_HISTORIAL_MEDIOS_PROHIBIDOS:]:
            row = dict(item)
            row["tokens_repeticion"] = sorted(row.get("tokens_repeticion", set()))
            row["claves_repeticion"] = list(row.get("claves_repeticion", []))
            row["frases_repeticion"] = list(row.get("frases_repeticion", []))
            row["anclas_repeticion"] = list(row.get("anclas_repeticion", []))
            serializable.append(row)
        with open(HISTORIAL_MEDIOS_PROHIBIDOS_PATH, "w", encoding="utf-8") as fh:
            json.dump(serializable, fh, ensure_ascii=False, indent=2)
    except Exception as exc:
        log.warning(f"No se pudo guardar historial de medios prohibidos: {exc}")


def _agregar_articulo_a_indice(articulo, hashes, por_ancla):
    hash_rep = articulo.get("hash_repeticion", "")
    if hash_rep:
        hashes.add(hash_rep)
    for clave in articulo.get("claves_repeticion", [])[:10]:
        if clave and len(clave) >= 5:
            por_ancla.setdefault(f"clave::{clave}", []).append(articulo)
    for ancla in articulo.get("anclas_repeticion", []):
        por_ancla.setdefault(ancla, []).append(articulo)
    for frase in articulo.get("frases_repeticion", [])[:6]:
        if len(frase) >= 18:
            por_ancla.setdefault(frase, []).append(articulo)


def _indexar_articulos_repeticion(articulos):
    hashes = set()
    por_ancla = {}
    for item in articulos:
        _agregar_articulo_a_indice(item, hashes, por_ancla)
    return hashes, por_ancla


def _indexar_historial_articulos(historial):
    return _indexar_articulos_repeticion(historial)


def _obtener_candidatos_repeticion(articulo, por_ancla, max_candidatos=80):
    candidatos = {}
    claves_busqueda = list(articulo.get("anclas_repeticion", []))
    claves_busqueda.extend(
        f"clave::{clave}"
        for clave in articulo.get("claves_repeticion", [])[:10]
        if clave and len(clave) >= 5
    )
    claves_busqueda.extend(frase for frase in articulo.get("frases_repeticion", [])[:6] if len(frase) >= 18)
    for ancla in claves_busqueda:
        for item in por_ancla.get(ancla, []):
            clave = item.get("hash_repeticion", f"id-{id(item)}")
            candidatos[clave] = item
            if len(candidatos) >= max_candidatos:
                return list(candidatos.values())
    return list(candidatos.values())


def _metricas_similitud_articulos(articulo, referencia):
    art_t_norm = articulo.get("t_norm", "")
    ref_t_norm = referencia.get("t_norm", "")
    titulo_ratio = 0.0
    if art_t_norm and ref_t_norm and abs(len(art_t_norm) - len(ref_t_norm)) <= 60:
        titulo_ratio = SequenceMatcher(None, art_t_norm, ref_t_norm).ratio()

    art_texto = articulo.get("texto_repeticion", "")
    ref_texto = referencia.get("texto_repeticion", "")
    texto_ratio = 0.0
    if art_texto and ref_texto:
        texto_ratio = SequenceMatcher(None, art_texto[:720], ref_texto[:720]).ratio()

    art_firma = articulo.get("firma_tokens_repeticion", "")
    ref_firma = referencia.get("firma_tokens_repeticion", "")
    firma_ratio = 0.0
    if art_firma and ref_firma:
        firma_ratio = SequenceMatcher(None, art_firma, ref_firma).ratio()

    art_firma_claves = articulo.get("firma_claves_repeticion", "")
    ref_firma_claves = referencia.get("firma_claves_repeticion", "")
    firma_claves_ratio = 0.0
    if art_firma_claves and ref_firma_claves:
        firma_claves_ratio = SequenceMatcher(None, art_firma_claves, ref_firma_claves).ratio()

    art_tokens = set(articulo.get("tokens_repeticion", set())) or set(articulo.get("tokens_relevantes", set()))
    ref_tokens = set(referencia.get("tokens_repeticion", set())) or set(referencia.get("tokens_relevantes", set()))
    inter = art_tokens.intersection(ref_tokens)
    union = art_tokens.union(ref_tokens)
    jaccard = (len(inter) / len(union)) if union else 0.0

    art_claves = set(articulo.get("claves_repeticion", []))
    ref_claves = set(referencia.get("claves_repeticion", []))
    claves_inter = art_claves.intersection(ref_claves)
    claves_union = art_claves.union(ref_claves)
    claves_jaccard = (len(claves_inter) / len(claves_union)) if claves_union else 0.0

    art_frases = {
        frase for frase in articulo.get("frases_repeticion", [])
        if frase and len(frase) >= 18
    }
    ref_frases = {
        frase for frase in referencia.get("frases_repeticion", [])
        if frase and len(frase) >= 18
    }
    frases_comunes = art_frases.intersection(ref_frases)

    art_slug_tokens = set((articulo.get("slug_repeticion", "") or "").split())
    ref_slug_tokens = set((referencia.get("slug_repeticion", "") or "").split())
    slug_inter = art_slug_tokens.intersection(ref_slug_tokens)

    return {
        "titulo_ratio": titulo_ratio,
        "texto_ratio": texto_ratio,
        "firma_ratio": firma_ratio,
        "firma_claves_ratio": firma_claves_ratio,
        "inter": inter,
        "union": union,
        "jaccard": jaccard,
        "claves_inter": claves_inter,
        "claves_jaccard": claves_jaccard,
        "frases_comunes": frases_comunes,
        "slug_inter": slug_inter,
    }


def _es_articulo_muy_parecido(articulo, referencia):
    hash_art = articulo.get("hash_repeticion", "")
    hash_ref = referencia.get("hash_repeticion", "")
    if hash_art and hash_ref and hash_art == hash_ref:
        return True

    metricas = _metricas_similitud_articulos(articulo, referencia)
    titulo_ratio = metricas["titulo_ratio"]
    texto_ratio = metricas["texto_ratio"]
    firma_ratio = metricas["firma_ratio"]
    firma_claves_ratio = metricas["firma_claves_ratio"]
    inter = metricas["inter"]
    jaccard = metricas["jaccard"]
    claves_inter = metricas["claves_inter"]
    claves_jaccard = metricas["claves_jaccard"]
    frases_comunes = metricas["frases_comunes"]
    slug_inter = metricas["slug_inter"]

    if titulo_ratio >= 0.95:
        return True
    if texto_ratio >= 0.92:
        return True
    if any(len(frase) >= 22 for frase in frases_comunes):
        return True
    if len(frases_comunes) >= 2:
        return True
    if len(frases_comunes) >= 1 and len(inter) >= 4:
        return True
    if len(slug_inter) >= 4 and len(inter) >= 5 and titulo_ratio >= 0.50:
        return True
    if len(slug_inter) >= 5 and len(inter) >= 5:
        return True
    if len(claves_inter) >= 4:
        return True
    if len(claves_inter) >= 3 and (firma_claves_ratio >= 0.62 or claves_jaccard >= 0.34):
        return True
    if titulo_ratio >= 0.88 and texto_ratio >= 0.84 and len(inter) >= 4:
        return True
    if len(inter) >= 7 and jaccard >= 0.30:
        return True
    if len(inter) >= 5 and (firma_ratio >= 0.70 or jaccard >= 0.38):
        return True
    if len(inter) >= 4 and titulo_ratio >= 0.70 and (texto_ratio >= 0.66 or firma_ratio >= 0.68):
        return True
    if len(inter) >= 6 and titulo_ratio >= 0.55 and firma_ratio >= 0.58:
        return True
    if len(inter) >= 5 and titulo_ratio >= 0.45 and (firma_ratio >= 0.48 or firma_claves_ratio >= 0.58):
        return True

    return False


def _es_coincidencia_prohibida_extrema(articulo, referencia):
    metricas = _metricas_similitud_articulos(articulo, referencia)
    inter = metricas["inter"]
    claves_inter = metricas["claves_inter"]

    if len(claves_inter) >= 3 and len(inter) >= 4:
        return True
    if len(claves_inter) >= 2 and len(metricas["slug_inter"]) >= 3:
        return True
    if len(inter) >= 5 and metricas["jaccard"] >= 0.20:
        return True
    if len(inter) >= 4 and metricas["firma_ratio"] >= 0.46:
        return True
    if len(inter) >= 4 and metricas["firma_claves_ratio"] >= 0.52:
        return True
    if len(metricas["frases_comunes"]) >= 1 and len(inter) >= 3:
        return True
    if len(inter) >= 6:
        return True
    return False


def _es_caso_borde_repeticion(articulo, referencia):
    metricas = _metricas_similitud_articulos(articulo, referencia)
    inter = metricas["inter"]
    if metricas["frases_comunes"]:
        return True
    if len(metricas["claves_inter"]) >= 2:
        return True
    if len(metricas["slug_inter"]) >= 4 and len(inter) >= 4:
        return True
    if len(inter) >= 5 and (metricas["firma_ratio"] >= 0.48 or metricas["jaccard"] >= 0.22):
        return True
    if metricas["titulo_ratio"] >= 0.42 and len(inter) >= 4:
        return True
    return False


def _es_coincidencia_indice_repeticion(articulo, hashes, por_ancla, max_candidatos=80):
    hash_rep = articulo.get("hash_repeticion", "")
    if hash_rep and hash_rep in hashes:
        return True

    for referencia in _obtener_candidatos_repeticion(articulo, por_ancla, max_candidatos=max_candidatos):
        if _es_articulo_muy_parecido(articulo, referencia):
            return True

    return False


def _es_coincidencia_historial(articulo, historial_hashes, historial_por_ancla):
    return _es_coincidencia_indice_repeticion(
        articulo,
        historial_hashes,
        historial_por_ancla,
        max_candidatos=60,
    )


def _cargar_cache_ollama_similitud():
    global OLLAMA_SIMILITUD_CACHE
    if OLLAMA_SIMILITUD_CACHE is not None:
        return OLLAMA_SIMILITUD_CACHE
    if not os.path.exists(OLLAMA_SIMILITUD_CACHE_PATH):
        OLLAMA_SIMILITUD_CACHE = {}
        return OLLAMA_SIMILITUD_CACHE
    try:
        with open(OLLAMA_SIMILITUD_CACHE_PATH, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        OLLAMA_SIMILITUD_CACHE = data if isinstance(data, dict) else {}
    except Exception:
        OLLAMA_SIMILITUD_CACHE = {}
    return OLLAMA_SIMILITUD_CACHE


def _guardar_cache_ollama_similitud():
    if OLLAMA_SIMILITUD_CACHE is None:
        return
    try:
        with open(OLLAMA_SIMILITUD_CACHE_PATH, "w", encoding="utf-8") as fh:
            json.dump(OLLAMA_SIMILITUD_CACHE, fh, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _resolver_ollama_cli():
    global OLLAMA_CLI_PATH
    if OLLAMA_CLI_PATH is not None:
        return OLLAMA_CLI_PATH
    candidatos = [
        "ollama",
        r"C:\Users\photo\AppData\Local\Programs\Ollama\ollama.exe",
        r"C:\Program Files\Ollama\ollama.exe",
    ]
    for candidato in candidatos:
        try:
            proc = subprocess.run(
                [candidato, "--version"],
                capture_output=True,
                text=True,
                timeout=4,
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
            )
            if proc.returncode == 0:
                OLLAMA_CLI_PATH = candidato
                return OLLAMA_CLI_PATH
        except Exception:
            continue
    OLLAMA_CLI_PATH = ""
    return OLLAMA_CLI_PATH


def _fetch_texto_url(url, timeout=8, accept_html=False):
    import urllib.error
    headers = dict(HEADERS)
    if accept_html:
        headers["Accept"] = "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    try:
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            contenido = resp.read()
            encoding = "utf-8"
            ct = resp.headers.get("Content-Type", "")
            if "charset=" in ct:
                encoding = ct.split("charset=")[-1].strip().split(";")[0]
            try:
                return contenido.decode(encoding)
            except Exception:
                return contenido.decode("utf-8", errors="replace")
    except Exception:
        return ""


def _extraer_texto_html_articulo(html_str):
    if not html_str:
        return ""
    texto = re.sub(r"(?is)<script.*?>.*?</script>", " ", html_str)
    texto = re.sub(r"(?is)<style.*?>.*?</style>", " ", texto)
    texto = re.sub(r"(?is)<noscript.*?>.*?</noscript>", " ", texto)
    texto = re.sub(r"(?is)<!--.*?-->", " ", texto)
    texto = re.sub(r"(?is)<(br|p|div|li|h1|h2|h3|article|section)[^>]*>", "\n", texto)
    texto = re.sub(r"(?is)<[^>]+>", " ", texto)
    texto = html.unescape(texto)
    lineas = []
    vistos = set()
    for linea in texto.splitlines():
        limpia = re.sub(r"\s+", " ", linea).strip()
        if len(limpia) < 60:
            continue
        limpia_norm = _normalizar_para_repeticion(limpia)
        if not limpia_norm or limpia_norm in vistos:
            continue
        if "cookies" in limpia_norm or "suscrib" in limpia_norm or "whatsapp" in limpia_norm:
            continue
        vistos.add(limpia_norm)
        lineas.append(limpia)
        if len(" ".join(lineas)) >= CONTENIDO_PROFUNDO_MAX_CHARS:
            break
    return " ".join(lineas)[:CONTENIDO_PROFUNDO_MAX_CHARS]


def _obtener_contexto_profundo_articulo(articulo):
    url = articulo.get("url", "")
    if not url:
        return articulo.get("resumen", "")
    with CONTENIDO_ARTICULOS_LOCK:
        if url in CONTENIDO_ARTICULOS_CACHE:
            return CONTENIDO_ARTICULOS_CACHE[url]
    html_str = _fetch_texto_url(url, timeout=7, accept_html=True)
    contenido = _extraer_texto_html_articulo(html_str)
    if not contenido:
        contenido = articulo.get("resumen_limpio") or articulo.get("resumen") or articulo.get("titulo", "")
    with CONTENIDO_ARTICULOS_LOCK:
        CONTENIDO_ARTICULOS_CACHE[url] = contenido[:CONTENIDO_PROFUNDO_MAX_CHARS]
    return CONTENIDO_ARTICULOS_CACHE[url]


def _dictamen_ollama_mismo_tema(articulo, referencia):
    global OLLAMA_VALIDACIONES_RESTANTES
    ollama_cli = _resolver_ollama_cli()
    if not ollama_cli or OLLAMA_VALIDACIONES_RESTANTES <= 0:
        return False

    cache = _cargar_cache_ollama_similitud()
    hash_a = articulo.get("hash_repeticion", "")
    hash_b = referencia.get("hash_repeticion", "")
    cache_key = "|".join(sorted([hash_a or articulo.get("url", ""), hash_b or referencia.get("url", "")]))
    if cache_key in cache:
        return bool(cache[cache_key])

    contexto_a = _obtener_contexto_profundo_articulo(articulo)
    contexto_b = _obtener_contexto_profundo_articulo(referencia)
    prompt = (
        "Responde solo SI o NO.\n"
        "Determina si estas dos noticias tratan del mismo hecho, tema central o contenido editorial claramente derivado, "
        "aunque el titular haya sido reescrito.\n\n"
        f"NOTICIA A\nTitulo: {articulo.get('titulo', '')}\nResumen: {articulo.get('resumen_limpio', articulo.get('resumen', ''))}\n"
        f"URL: {articulo.get('url', '')}\nContenido: {contexto_a}\n\n"
        f"NOTICIA B\nTitulo: {referencia.get('titulo', '')}\nResumen: {referencia.get('resumen_limpio', referencia.get('resumen', ''))}\n"
        f"URL: {referencia.get('url', '')}\nContenido: {contexto_b}\n"
    )

    try:
        proc = subprocess.run(
            [ollama_cli, "run", "llama3.2", prompt],
            capture_output=True,
            text=True,
            timeout=OLLAMA_TIMEOUT_SEGUNDOS,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        salida = (proc.stdout or "").strip().lower()
        verdict = salida.startswith("si") or salida.startswith("sí")
    except Exception:
        verdict = False

    OLLAMA_VALIDACIONES_RESTANTES -= 1
    cache[cache_key] = bool(verdict)
    _guardar_cache_ollama_similitud()
    return verdict


def _medio_prohibido_por_texto(url="", titulo="", descripcion="", fuente_rss=""):
    netloc = _normalize_domain(url)
    fuente_norm = _normalizar_texto_medio(fuente_rss)
    titulo_lower = (titulo or "").lower().strip()
    texto_norm = _normalizar_texto_medio(f"{titulo} {descripcion} {fuente_rss}")
    separadores = (" - ", " — ", " | ", " – ", " :: ")

    for medio_key, cfg in MEDIOS_PROHIBIDOS.items():
        for dominio in cfg["dominios"]:
            if netloc == dominio or netloc.endswith(f".{dominio}"):
                return medio_key, f"Dominio bloqueado: {dominio}"

        for alias in cfg["source_aliases"]:
            if fuente_norm == _normalizar_texto_medio(alias):
                return medio_key, f"Feed RSS Fuente exacta: {fuente_rss}"

        for alias in cfg["title_aliases"]:
            alias_lower = alias.lower()
            if any(titulo_lower.endswith(f"{sep}{alias_lower}") for sep in separadores):
                return medio_key, f"Filtro en TÃ­tulo: {cfg['label']}"

        for firma in cfg["signatures"]:
            firma_norm = _normalizar_texto_medio(firma)
            if firma_norm and firma_norm in texto_norm:
                return medio_key, f"Firma especÃ­fica: '{cfg['label']}'"

    return None, ""


def _es_razon_medio_prohibido(razon):
    razon_norm = _normalizar_texto_medio(razon)
    if not razon_norm:
        return False
    for cfg in MEDIOS_PROHIBIDOS.values():
        if _normalizar_texto_medio(cfg["label"]) in razon_norm:
            return True
        for dominio in cfg["dominios"]:
            if dominio in razon_norm:
                return True
    return False


FIRMAS_BLOQUEADAS = [
    # FIX BUG 3: "El Tiempo" (periódico) se quitó de la lista de texto porque es
    # indistinguible de la palabra española "el tiempo" (the weather / the time).
    # El bloqueo de El Tiempo se hace por dominio (eltiempo.com) y por frases
    # específicas del periódico, NO por la cadena genérica "el tiempo".
    "portafolio.co", "Portafolio.co",
    "eltiempo.com",
    "Casa Editorial El Tiempo",
    "ELTIEMPO.COM",
]

# ═══════════════════════════════════════════════════════════════
# UTILIDADES DE FECHA
# ═══════════════════════════════════════════════════════════════

def _parsear_fecha_rss(fecha_str):
    if not fecha_str or not str(fecha_str).strip():
        return None
    fecha_str = str(fecha_str).strip()
    try:
        dt = parsedate_to_datetime(fecha_str)
        return dt.astimezone(timezone.utc)
    except Exception:
        pass
    formatos = [
        "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S.%fZ",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
        "%a, %d %b %Y %H:%M:%S %Z",
    ]
    for fmt in formatos:
        try:
            dt = datetime.strptime(fecha_str, fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except (ValueError, TypeError):
            continue
    if hasattr(fecha_str, 'tm_year'):
        try:
            return datetime(*fecha_str[:6], tzinfo=timezone.utc)
        except Exception:
            pass
    return None


def _es_fecha_confiable(dt):
    if dt is None:
        return False
    ahora = datetime.now(timezone.utc)
    return (ahora - timedelta(days=730)) <= dt <= (ahora + timedelta(hours=2))


def _fecha_a_date_colombia(dt):
    return (dt + timedelta(hours=-5)).date()


def _fecha_display(dt):
    dt_col = dt + timedelta(hours=-5)
    return dt_col.strftime("%Y-%m-%d %H:%M")


# ═══════════════════════════════════════════════════════════════
# FILTRO DE FUENTES BLOQUEADAS
# ═══════════════════════════════════════════════════════════════

import base64

def _esta_bloqueado(url, titulo="", descripcion="", fuente_rss="", filtrar_argentina=True):
    """
    Retorna (bloqueado, razón).

    FIX 1: Las URLs de Google News (news.google.com/rss/articles/...) son
            redirecciones a fuentes externas. Se excluyen del check de
            DOMINIOS_PERMITIDOS porque su dominio es siempre news.google.com,
            que intencionalmente no está en la whitelist.

    FIX 2: Se eliminó el "el_tiempo" con guión bajo (false positive).

    FIX 3: "ingresa a nuestro grupo de whatsapp" eliminado — phrase demasiado
            genérica que bloqueaba fuentes legítimas.
    """
    url_lower = url.lower()
    fuente_lower = fuente_rss.lower()

    medio_key, razon_medio = _medio_prohibido_por_texto(url, titulo, descripcion, fuente_rss)
    if medio_key:
        return True, razon_medio

    # ── Bloqueo explícito por nombre de fuente RSS ──────────────────────
    # Sólo bloqueamos si la fuente dice EXACTAMENTE "el tiempo" o "portafolio"
    # como nombre de medio, no si aparece en cualquier contexto.
    if fuente_lower in ("el tiempo", "portafolio", "portafolio.co",
                        "eltiempo.com", "el tiempo play", "citytv", "city tv"):
        return True, f"Feed RSS Fuente exacta: {fuente_rss}"

    t_lower = titulo.lower()
    if t_lower.endswith(" - el tiempo") or t_lower.endswith(" — el tiempo") or \
       t_lower.endswith(" | el tiempo") or t_lower.endswith("- eltiempo.com"):
        return True, "Filtro en Título: El Tiempo"

    if t_lower.endswith(" - portafolio") or t_lower.endswith(" — portafolio") or \
       t_lower.endswith(" | portafolio") or t_lower.endswith("- portafolio.co"):
        return True, "Filtro en Título: Portafolio"

    # ── Frases EXCLUSIVAS de El Tiempo y Portafolio ─────────────────────
    if t_lower.endswith(" - citytv") or t_lower.endswith(" â€” citytv") or \
       t_lower.endswith(" | citytv") or t_lower.endswith(" - city tv"):
        return True, "Filtro en TÃ­tulo: CityTV"

    texto_lower = f"{titulo} {descripcion} {fuente_rss}".lower()
    frases_prohibidas = [
        "artículo exclusivo para suscriptores de el tiempo",
        "sigue a el tiempo en whatsapp",
        "el tiempo play",
        "casa editorial el tiempo",
        "portafolio digital",
        "suscríbete a portafolio",
        "suscríbete a el tiempo",
        "el tiempoplay",
        "fuente: el tiempo",
        "por el tiempo",
        # FIX: "el_tiempo" con guión bajo ELIMINADO (generaba falsos positivos)
        # FIX: "ingresa a nuestro grupo de whatsapp" ELIMINADO (demasiado genérico)
    ]
    for frase in frases_prohibidas:
        if frase in texto_lower:
            return True, f"Frase bloqueada (El Tiempo/Portafolio): {frase}"

    # ── Decodificar URL base64 de Google News ───────────────────────────
    if "news.google.com/rss/articles/" in url_lower:
        try:
            b64_part = url_lower.split("articles/")[-1].split("?")[0]
            b64_part += "=" * ((4 - len(b64_part) % 4) % 4)
            decodificado = base64.urlsafe_b64decode(b64_part).decode('utf-8', errors='ignore').lower()
            dominios_prohibidos = []
            for cfg in MEDIOS_PROHIBIDOS.values():
                dominios_prohibidos.extend(cfg["dominios"])
            for dom_bloq in dominios_prohibidos:
                if dom_bloq in decodificado:
                    return True, f"Google News redirige a dominio bloqueado: {dom_bloq}"
        except Exception:
            pass

    # ── FIX 1: Whitelist (DOMINIOS_PERMITIDOS) ──────────────────────────
    # Las URLs de Google News redirect (news.google.com/rss/articles/...)
    # NO se pasan por la whitelist porque su netloc siempre es news.google.com,
    # que fue excluido intencionalmente. El destino real ya fue verificado arriba.
    try:
        netloc = _normalize_domain(url)

        es_gnews_redirect = "news.google.com" in netloc

        if not es_gnews_redirect:
            # Para fuentes directas, aplicar whitelist normalmente
            if DOMINIOS_PERMITIDOS and netloc not in DOMINIOS_PERMITIDOS:
                return True, f"Dominio no está en BD permitida: {netloc}"
    except Exception:
        pass

    if "redmas.com.co" in url_lower or "cronista.com" in url_lower:
        return True, "Filtro manual redmas/cronista"

    for dominio in DOMINIOS_BLOQUEADOS:
        if dominio in url_lower:
            return True, f"Dominio bloqueado: {dominio}"

    # ── Filtro de Argentina (opcional) ──────────────────────────────────
    if filtrar_argentina:
        for dominio in DOMINIOS_ARGENTINA:
            if dominio in url_lower:
                return True, f"Dominio Argentina: {dominio}"

    # ── Firmas textuales ─────────────────────────────────────────────────
    # FIX BUG 3: Solo usamos patrones muy específicos que no existen en español normal.
    # "el tiempo" genérico fue ELIMINADO (falso positivo con "el tiempo libre", etc.)
    # El bloqueo principal es por dominio (DOMINIOS_BLOQUEADOS / whitelist).
    firmas_especificas = [
        "eltiempo.com", "portafolio.co", "portafolio.co",
        "casa editorial el tiempo",
        "el tiempo play", "eltiempoplay",
        "suscríbete a el tiempo", "suscríbete a portafolio",
    ]
    for firma in firmas_especificas:
        if firma in texto_lower:
            return True, f"Firma específica: '{firma}'"

    if 'source="el tiempo"' in texto_lower or '>el tiempo</' in texto_lower:
        return True, "Firma oculta RSS: El Tiempo"
    if 'source="citytv"' in texto_lower or '>citytv</' in texto_lower:
        return True, "Firma oculta RSS: CityTV"

    return False, ""


# ═══════════════════════════════════════════════════════════════
# UTILIDADES DE TEXTO
# ═══════════════════════════════════════════════════════════════

def _limpiar_html(texto):
    if not texto:
        return ""
    limpio = re.sub(r'<[^>]+>', '', texto)
    return re.sub(r'\s+', ' ', limpio).strip()[:300]


def _normalizar(texto):
    if not texto:
        return ""
    texto = texto.lower().strip()
    texto = re.sub(r'[^\w\s]', '', texto)
    return re.sub(r'\s+', ' ', texto)


def _parece_ingles_puro(titulo, descripcion):
    """
    FIX 2: Reemplaza el filtro anterior " the "/" of "/" is " que era
    demasiado agresivo y bloqueaba artículos legítimos de Infobae, Forbes,
    El Heraldo, etc. que mencionan marcas o medios en inglés.

    Ahora sólo descarta un artículo si:
    - Tiene ≥4 marcadores lingüísticos del inglés en el título, Y
    - No tiene ninguna vocal española (á,é,í,ó,ú,ñ,ü,¿,¡), Y
    - Tiene longitud suficiente para el análisis.
    """
    MARCADORES_EN = [" the ", " is ", " are ", " was ", " were ",
                     " has ", " have ", " with ", " and ", " that ",
                     " this ", " from ", " their "]
    texto = f"{titulo} {descripcion}".lower()
    conteo = sum(1 for m in MARCADORES_EN if m in texto)
    tiene_espanol = any(c in texto for c in 'áéíóúñü¿¡')
    return conteo >= 4 and not tiene_espanol and len(texto) > 60


# ═══════════════════════════════════════════════════════════════
# FETCH + PARSE RSS
# ═══════════════════════════════════════════════════════════════

def _fetch_rss(url, timeout=TIMEOUT, max_retries=3):
    import urllib.error
    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                contenido = resp.read()
                encoding = "utf-8"
                ct = resp.headers.get("Content-Type", "")
                if "charset=" in ct:
                    encoding = ct.split("charset=")[-1].strip().split(";")[0]
                try:
                    return contenido.decode(encoding)
                except Exception:
                    return contenido.decode("utf-8", errors="replace")
        except urllib.error.HTTPError as e:
            if e.code == 429:
                time.sleep(1.5 ** attempt)
                continue
            return None
        except Exception:
            return None
    return None


def _parsear_feed(xml_str, nombre_fuente):
    """
    Parsea RSS/Atom. Descarta artículos sin fecha confiable.

    FIX 2: El filtro de idioma inglés ' the '/' of '/' is ' fue ELIMINADO.
    Era la causa principal de que Infobae, Forbes y El Heraldo no aparecieran:
    sus artículos sobre marcas internacionales, Netflix, Apple, FMI, etc.
    contenían estas palabras y se descartaban silenciosamente.

    Ahora se usa _parece_ingles_puro() que requiere ≥4 marcadores Y ausencia
    de caracteres del español — mucho más preciso.
    """
    articulos = []
    try:
        xml_str = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', xml_str)
        root = ET.fromstring(xml_str)
    except ET.ParseError:
        return []

    ns_atom = {"atom": "http://www.w3.org/2005/Atom"}
    es_atom = "feed" in root.tag.lower()

    items = (
        root.findall(".//atom:entry", ns_atom) or root.findall(".//entry")
    ) if es_atom else root.findall(".//item")

    for item in items:
        def _get(tag, ns_tag=None):
            el = item.find(tag)
            if el is None and ns_tag:
                el = item.find(ns_tag, ns_atom)
            return (el.text or "").strip() if el is not None else ""

        if es_atom:
            titulo = _get("title") or _get("atom:title", "atom:title")
            url = ""
            link_el = item.find("link") or item.find("{http://www.w3.org/2005/Atom}link")
            if link_el is not None:
                url = link_el.get("href", "") or link_el.text or ""
            fecha_str = _get("updated") or _get("published") or _get("{http://www.w3.org/2005/Atom}updated")
            descripcion = _get("summary") or _get("content")
            fuente_rss = _get("source") or _get("atom:source", "atom:source") or ""
        else:
            titulo = _get("title")
            url = _get("link") or _get("guid")
            fecha_str = _get("pubDate") or _get("dc:date") or _get("{http://purl.org/dc/elements/1.1/}date")
            descripcion = _get("description") or _get("summary")
            fuente_rss = _get("source") or ""

        if not titulo or not url:
            continue

        fecha_dt = _parsear_fecha_rss(fecha_str)
        if not _es_fecha_confiable(fecha_dt):
            continue

        # FIX 2: Filtro de idioma mejorado (mucho más permisivo con español legítimo)
        if _parece_ingles_puro(titulo, descripcion):
            continue

        t_norm = _normalizar(titulo)[:80]
        resumen_limpio = _limpiar_html(descripcion) or titulo
        art_claves = _extraer_palabras_clave(titulo + " " + descripcion)
        huella_repeticion = _construir_huella_repeticion(titulo, resumen_limpio, url)

        art_data = {
            "titulo":     titulo,
            "resumen":    resumen_limpio,
            "url":        url,
            "fecha_dt":   fecha_dt,
            "fecha_date": _fecha_a_date_colombia(fecha_dt),
            "fecha_str":  _fecha_display(fecha_dt),
            "fuente":     nombre_fuente,
            "t_norm":     t_norm,
            "claves":     art_claves,
            "tokens_relevantes": _extraer_tokens_relevantes(f"{titulo} {descripcion}"),
            **huella_repeticion,
        }

        bloqueado, razon = _esta_bloqueado(url, titulo, descripcion, fuente_rss, filtrar_argentina=False)
        if bloqueado:
            if _es_razon_medio_prohibido(razon):
                LISTA_NEGRA_MEDIOS.append(art_data)
            continue

        articulos.append(art_data)

    return articulos


def _parsear_forbes_economia_html(html_str, nombre_fuente):
    articulos = []
    vistos = set()
    patron = re.compile(
        r'<a[^>]+href="(?P<href>/\d{4}/\d{2}/\d{2}/economia-y-finanzas/[^"]+)"[^>]*>(?P<inner>.*?)</a>',
        re.IGNORECASE | re.DOTALL,
    )

    for match in patron.finditer(html_str):
        href = match.group("href")
        if href in vistos:
            continue

        titulo = re.sub(r"<[^>]+>", " ", match.group("inner"))
        titulo = html.unescape(re.sub(r"\s+", " ", titulo)).strip()
        if not titulo:
            continue

        vistos.add(href)
        url = urllib.parse.urljoin("https://forbes.co", href)

        fecha_match = re.search(r"/(\d{4})/(\d{2})/(\d{2})/economia-y-finanzas/", href)
        if not fecha_match:
            continue

        try:
            y, m, d = map(int, fecha_match.groups())
            fecha_dt = datetime(y, m, d, 12, 0, tzinfo=timezone.utc)
        except ValueError:
            continue

        if not _es_fecha_confiable(fecha_dt):
            continue

        t_norm = _normalizar(titulo)[:80]
        huella_repeticion = _construir_huella_repeticion(titulo, titulo, url)
        art_data = {
            "titulo": titulo,
            "resumen": titulo,
            "url": url,
            "fecha_dt": fecha_dt,
            "fecha_date": _fecha_a_date_colombia(fecha_dt),
            "fecha_str": _fecha_display(fecha_dt),
            "fuente": nombre_fuente,
            "t_norm": t_norm,
            "claves": _extraer_palabras_clave(titulo),
            "tokens_relevantes": _extraer_tokens_relevantes(titulo),
            **huella_repeticion,
        }

        bloqueado, razon = _esta_bloqueado(url, titulo, "", "", filtrar_argentina=False)
        if bloqueado:
            if _es_razon_medio_prohibido(razon):
                LISTA_NEGRA_MEDIOS.append(art_data)
            continue

        articulos.append(art_data)

    return articulos


def _cargar_lista_negra_medios(fecha_inicio=None, fecha_fin=None, verbose=True):
    global LISTA_NEGRA_MEDIOS_HASHES, LISTA_NEGRA_MEDIOS_POR_ANCLA
    LISTA_NEGRA_MEDIOS.clear()
    LISTA_NEGRA_MEDIOS_HASHES = set()
    LISTA_NEGRA_MEDIOS_POR_ANCLA = {}

    historial_medios = _cargar_historial_medios_prohibidos()
    if historial_medios:
        LISTA_NEGRA_MEDIOS.extend(historial_medios)

    fecha_ref = fecha_fin or fecha_inicio or _fecha_a_date_colombia(datetime.now(timezone.utc))
    fecha_negra_fin = fecha_fin or fecha_ref
    fecha_negra_inicio = fecha_inicio or fecha_ref
    fecha_negra_inicio = min(fecha_negra_inicio, fecha_negra_fin) - timedelta(days=VENTANA_LISTA_NEGRA_DIAS)
    fecha_descarga_inicio = max(
        fecha_negra_inicio,
        fecha_negra_fin - timedelta(days=VENTANA_LISTA_NEGRA_DESCARGA_DIAS),
    )

    fuentes_lista_negra = []
    for cfg in MEDIOS_PROHIBIDOS.values():
        fuentes_lista_negra.append({
            "nombre": f"{cfg['label']} Base",
            "url": cfg["lista_negra_url"],
            "categorias": ["general"],
            "tipo": "nacional",
        })

    with ThreadPoolExecutor(max_workers=min(3, len(fuentes_lista_negra))) as executor:
        futuro_a_fuente = {
            executor.submit(_fetch_fuente, f, fecha_descarga_inicio, fecha_negra_fin): f
            for f in fuentes_lista_negra
        }
        for futuro in as_completed(futuro_a_fuente):
            try:
                _, _, articulos = futuro.result()
            except Exception:
                articulos = []
            if articulos:
                LISTA_NEGRA_MEDIOS.extend(articulos)

    LISTA_NEGRA_MEDIOS_HASHES, LISTA_NEGRA_MEDIOS_POR_ANCLA = _indexar_articulos_repeticion(LISTA_NEGRA_MEDIOS)
    if LISTA_NEGRA_MEDIOS:
        historial_actualizado = []
        hashes_historial = set()
        for item in LISTA_NEGRA_MEDIOS:
            hash_rep = item.get("hash_repeticion", "")
            if not hash_rep or hash_rep in hashes_historial:
                continue
            historial_actualizado.append(_registro_historial_desde_articulo(item))
            hashes_historial.add(hash_rep)
        _guardar_historial_medios_prohibidos(historial_actualizado)

    if verbose:
        log.info(
            f"  [✓] {len(LISTA_NEGRA_MEDIOS)} noticias en lista negra de medios bloqueados "
            f"(memoria {VENTANA_LISTA_NEGRA_DIAS}d / descarga {VENTANA_LISTA_NEGRA_DESCARGA_DIAS}d)."
        )


def _es_coincidencia_lista_negra(articulo, lista_negra_hashes, lista_negra_por_ancla):
    if _es_coincidencia_indice_repeticion(
        articulo,
        lista_negra_hashes,
        lista_negra_por_ancla,
        max_candidatos=90,
    ):
        return True

    candidatos = _obtener_candidatos_repeticion(articulo, lista_negra_por_ancla, max_candidatos=28)
    for referencia in candidatos[:12]:
        if _es_coincidencia_prohibida_extrema(articulo, referencia):
            return True

    for referencia in candidatos[:8]:
        if not _es_caso_borde_repeticion(articulo, referencia):
            continue
        if _dictamen_ollama_mismo_tema(articulo, referencia):
            return True
    return False


# ═══════════════════════════════════════════════════════════════
# MOTOR DE BÚSQUEDA
# ═══════════════════════════════════════════════════════════════

CATEGORIAS_DISPONIBLES = [
    "general", "economia", "politica", "deportes", "tecnologia",
    "cultura", "judicial", "justicia", "internacional", "salud",
    "tendencias", "negocios", "finanzas", "colombia",
    "motor", "vida", "virales", "bogota", "mis finanzas",
]

from concurrent.futures import ThreadPoolExecutor, as_completed


def _fetch_fuente(fuente, fecha_inicio, fecha_fin):
    """Descarga y parsea una fuente RSS completa. Diseñado para ejecución paralela."""
    nombre = fuente["nombre"]
    url_base = fuente["url"]
    is_google = "news.google.com/rss/search?q=" in url_base

    urls_to_fetch = []
    if is_google and fecha_inicio and fecha_fin:
        delta = (fecha_fin - fecha_inicio).days
        for d in range(0, delta + 1):
            dia_actual = fecha_inicio + timedelta(days=d)
            dia_siguiente = dia_actual + timedelta(days=1)
            f_ini_str = dia_actual.strftime("%Y-%m-%d")
            f_fin_str = dia_siguiente.strftime("%Y-%m-%d")
            partes = url_base.split("&hl=")
            q_part = partes[0]
            resto = "&hl=" + partes[1] if len(partes) > 1 else ""
            fechas_query = urllib.parse.quote(f" after:{f_ini_str} before:{f_fin_str}")
            urls_to_fetch.append(q_part + fechas_query + resto)
    else:
        urls_to_fetch.append(url_base)

    articulos_fuente = []
    responded = False
    for u in urls_to_fetch:
        xml_str = _fetch_rss(u)
        if xml_str:
            responded = True
            if nombre == "Forbes Colombia" and "economia-y-finanzas" in u:
                arts = _parsear_forbes_economia_html(xml_str, nombre)
            else:
                arts = _parsear_feed(xml_str, nombre)
            articulos_fuente.extend(arts)
        if is_google and len(urls_to_fetch) > 1:
            time.sleep(1.0)

    return nombre, responded, articulos_fuente


def buscar_noticias(categorias_seleccionadas=None, fecha_inicio=None, fecha_fin=None,
                    max_por_fuente=50, max_total=1000, verbose=True,
                    tipo_noticias="ambas", filtrar_argentina=True):
    global OLLAMA_VALIDACIONES_RESTANTES
    OLLAMA_VALIDACIONES_RESTANTES = MAX_VALIDACIONES_OLLAMA_POR_BUSQUEDA

    if categorias_seleccionadas:
        cats_ordenadas = [c.lower() for c in categorias_seleccionadas]
    else:
        cats_ordenadas = []

    cats_filtrado = list(cats_ordenadas)
    cats_filtrado_set = set(cats_filtrado) if cats_filtrado else None
    cats_expandidas = _expandir_categorias_solicitadas(cats_filtrado) if cats_filtrado else set()

    fuentes = []
    for fuente in FUENTES_RSS:
        if tipo_noticias != "ambas" and fuente.get("tipo", "nacional") != tipo_noticias:
            continue
        if not cats_filtrado:
            fuentes.append(fuente)
            continue

        coincide_categoria = any(c in cats_expandidas for c in fuente["categorias"]) if cats_expandidas else False
        if coincide_categoria:
            fuentes.append(fuente)

    if verbose:
        log.info("")
        log.info("=" * 60)
        log.info("  BUSCADOR DE NOTICIAS CAPA BRINDADA - V.6")
        log.info("=" * 60)
        log.info("  [!] Descargando lista negra de El Tiempo, Portafolio y CityTV...")

    _cargar_lista_negra_medios(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, verbose=verbose)

    if fecha_inicio and fecha_fin:
        if fecha_inicio == fecha_fin:
            fecha_info = fecha_inicio.strftime("%Y-%m-%d")
        else:
            fecha_info = f"{fecha_inicio.strftime('%Y-%m-%d')} a {fecha_fin.strftime('%Y-%m-%d')}"
    else:
        fecha_info = "Todas"

    tipo_info = {
        "nacional": "Nacional",
        "internacional": "Internacional",
        "ambas": "Ambas",
    }.get(tipo_noticias, tipo_noticias)

    if verbose:
        log.info(f"  [OK] {len(LISTA_NEGRA_MEDIOS)} noticias en lista negra.")
        log.info(f"  Fuentes seleccionadas: {len(fuentes)}")
        log.info(f"  Tipo de noticias: {tipo_info}")
        log.info(f"  Filtro Argentina: {'Si' if filtrar_argentina else 'No'}")
        log.info(f"  Filtro de fecha: {fecha_info}")
        log.info("=" * 60)

    todas = []
    fuentes_fallidas = []
    conteo_fuentes = {}
    articulos_vistos_hashes = set()
    articulos_vistos_por_ancla = {}
    historial_articulos = _cargar_historial_articulos()
    historial_hashes, historial_por_ancla = _indexar_historial_articulos(historial_articulos)

    fuentes_google = [f for f in fuentes if "news.google.com/rss/search?q=" in f["url"]]
    fuentes_directas = [f for f in fuentes if "news.google.com/rss/search?q=" not in f["url"]]

    resultados_directos = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        futuro_a_fuente = {
            executor.submit(_fetch_fuente, f, fecha_inicio, fecha_fin): f
            for f in fuentes_directas
        }
        for futuro in as_completed(futuro_a_fuente):
            try:
                nombre, responded, articulos = futuro.result()
                resultados_directos[nombre] = (responded, articulos)
            except Exception:
                fuente = futuro_a_fuente[futuro]
                resultados_directos[fuente["nombre"]] = (False, [])

    resultados_google = {}
    for fuente in fuentes_google:
        try:
            nombre, responded, articulos = _fetch_fuente(fuente, fecha_inicio, fecha_fin)
            resultados_google[nombre] = (responded, articulos)
        except Exception:
            resultados_google[fuente["nombre"]] = (False, [])

    todos_resultados = {**resultados_directos, **resultados_google}

    for fuente in fuentes:
        nombre = fuente["nombre"]
        responded, articulos_fuente = todos_resultados.get(nombre, (False, []))

        if not responded:
            fuentes_fallidas.append(nombre)
            if verbose:
                log.info(f"  [...] {nombre}...")
                log.info("    x Sin respuesta")
            continue

        articulos = articulos_fuente

        if filtrar_argentina:
            articulos = [
                art for art in articulos
                if not _esta_bloqueado(
                    art["url"],
                    art["titulo"],
                    art.get("resumen", ""),
                    "",
                    filtrar_argentina=True,
                )[0]
            ]

        if fecha_inicio and fecha_fin:
            articulos = [a for a in articulos if fecha_inicio <= a["fecha_date"] <= fecha_fin]

        if tipo_noticias == "nacional":
            articulos = [a for a in articulos if _articulo_es_nacional_colombia(a, fuente)]
        elif tipo_noticias == "internacional":
            articulos = [a for a in articulos if _articulo_cumple_filtro_internacional(a)]

        articulos_filtrados_categoria = []
        for art in articulos:
            if cats_filtrado_set:
                categoria_articulo = None
                for cat in cats_filtrado:
                    relacionadas = CATEGORIAS_RELACIONADAS.get(cat, set())
                    if cat not in fuente["categorias"] and not set(fuente["categorias"]).intersection(relacionadas):
                        continue
                    if _articulo_coincide_categoria(
                        cat,
                        titulo=art.get("titulo", ""),
                        resumen=art.get("resumen", ""),
                        fuente=art.get("fuente", ""),
                        categorias_fuente=fuente["categorias"],
                    ):
                        categoria_articulo = cat
                        break

                if not categoria_articulo:
                    continue
                art["categoria"] = categoria_articulo.upper()
            else:
                art["categoria"] = fuente["categorias"][0].upper()

            articulos_filtrados_categoria.append(art)

        articulos = articulos_filtrados_categoria

        nuevos = []
        for art in articulos[:max_por_fuente]:
            art_t_norm = art.get("t_norm", "")
            if not art_t_norm:
                continue

            if art.get("categoria") == "TENDENCIAS":
                if not _es_tendencia_valida(
                    art.get("titulo", ""),
                    art.get("resumen", ""),
                    art.get("fuente", ""),
                ):
                    continue

            if _es_coincidencia_lista_negra(
                art,
                LISTA_NEGRA_MEDIOS_HASHES,
                LISTA_NEGRA_MEDIOS_POR_ANCLA,
            ):
                continue

            if _es_coincidencia_historial(art, historial_hashes, historial_por_ancla):
                continue

            if not _es_coincidencia_indice_repeticion(
                art,
                articulos_vistos_hashes,
                articulos_vistos_por_ancla,
                max_candidatos=30,
            ):
                _agregar_articulo_a_indice(art, articulos_vistos_hashes, articulos_vistos_por_ancla)
                nuevos.append(art)

        todas.extend(nuevos)
        conteo_fuentes[nombre] = len(nuevos)
        if verbose:
            log.info(f"  [...] {nombre}...")
            log.info(f"    OK {len(nuevos)} articulos con fecha verificada")

    todas.sort(key=lambda a: a["fecha_dt"], reverse=True)
    resultado = todas[:max_total]

    if resultado:
        historial_actualizado = historial_articulos[:]
        hashes_historial = set(historial_hashes)
        for art in resultado:
            hash_rep = art.get("hash_repeticion", "")
            if not hash_rep or hash_rep in hashes_historial:
                continue
            historial_actualizado.append(_registro_historial_desde_articulo(art))
            hashes_historial.add(hash_rep)
        _guardar_historial_articulos(historial_actualizado)

    notificacion = None
    if len(resultado) == 0:
        if fecha_inicio and fecha_fin:
            if fecha_inicio == fecha_fin:
                fmt = fecha_inicio.strftime("%d/%m/%Y")
            else:
                fmt = f"del {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
            ahora_col = _fecha_a_date_colombia(datetime.now(timezone.utc))
            if fecha_inicio > ahora_col:
                notificacion = f"Rango {fmt} es futuro. No hay noticias aun."
            else:
                notificacion = (
                    f"Sin noticias para el rango {fmt}.\n"
                    f"   Se consultaron {len(fuentes) - len(fuentes_fallidas)} fuentes.\n"
                    f"   Las fuentes pueden no tener articulos de esas fechas."
                )
        else:
            notificacion = "No se encontraron noticias con los criterios seleccionados."
        if verbose:
            log.warning(f"  {notificacion}")

    if verbose:
        log.info(f"\n  TOTAL: {len(resultado)} articulos con fecha verificada")

    return {
        "noticias": resultado,
        "total": len(resultado),
        "fuentes_consultadas": len(fuentes) - len(fuentes_fallidas),
        "fuentes_fallidas": fuentes_fallidas,
        "conteo_fuentes": conteo_fuentes,
        "notificacion": notificacion,
    }

# ═══════════════════════════════════════════════════════════════
# GENERADOR DE EXCEL — UNA HOJA POR CATEGORÍA
# ═══════════════════════════════════════════════════════════════

class GeneradorExcelIDEAS:
    FONT_NAME = "Century Gothic"
    NEGRO = "000000"
    BLANCO = "FFFFFF"
    GRIS_CLARO = "F5F5F5"
    GRIS_BORDE = "CCCCCC"
    COLORES_TAB = [
        "FF4444", "FFD700", "00CC66", "00CCCC", "FF66FF",
        "FF8800", "4488FF", "AA44FF", "44DDAA", "DD4488",
        "88CC00", "0088DD", "FF6644", "6644FF", "44FF88",
    ]

    def __init__(self, articulos):
        self.articulos = articulos
        self.wb = Workbook()

    def generar(self, nombre_archivo):
        log.info("")
        log.info("=" * 60)
        log.info("  Generando Excel (una hoja por categoría)...")
        log.info("=" * 60)

        por_cat = {}
        for art in self.articulos:
            cat = art.get("categoria", "GENERAL")
            por_cat.setdefault(cat, []).append(art)

        cats_ord = sorted(por_cat.keys())

        ws_res = self.wb.active
        ws_res.title = "RESUMEN"
        ws_res.sheet_properties.tabColor = self.NEGRO
        self._hoja_resumen(ws_res, por_cat, cats_ord)

        for i, cat in enumerate(cats_ord):
            nombre_hoja = cat[:31]
            ws = self.wb.create_sheet(title=nombre_hoja)
            ws.sheet_properties.tabColor = self.COLORES_TAB[i % len(self.COLORES_TAB)]
            self._hoja_datos(ws, por_cat[cat])
            log.info(f"  ✓ Hoja '{nombre_hoja}': {len(por_cat[cat])} artículos")

        self.wb.save(nombre_archivo)
        log.info(f"  ✓ Guardado: {nombre_archivo}")
        log.info(f"  ✓ Total: {len(self.articulos)} en {len(cats_ord)} hojas")
        return nombre_archivo

    def _hoja_resumen(self, ws, por_cat, cats_ord):
        headers = [("CATEGORÍA", "FFFFFF"), ("ARTÍCULOS", "FFD700")]
        for ci, (h, c) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(name=self.FONT_NAME, bold=True, size=11, color=c)
            cell.fill = PatternFill(start_color=self.NEGRO, end_color=self.NEGRO, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        borde = Border(
            left=Side(style='hair', color=self.GRIS_BORDE),
            right=Side(style='hair', color=self.GRIS_BORDE),
            top=Side(style='hair', color=self.GRIS_BORDE),
            bottom=Side(style='hair', color=self.GRIS_BORDE))
        fp = PatternFill(start_color=self.GRIS_CLARO, end_color=self.GRIS_CLARO, fill_type='solid')
        fi = PatternFill(start_color=self.BLANCO, end_color=self.BLANCO, fill_type='solid')
        total = 0
        for idx, cat in enumerate(cats_ord):
            row = idx + 2
            cnt = len(por_cat[cat])
            total += cnt
            fill = fp if idx % 2 == 0 else fi
            c = ws.cell(row=row, column=1, value=cat)
            c.font = Font(name=self.FONT_NAME, size=10, bold=True, color="333333")
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.fill = fill; c.border = borde
            c = ws.cell(row=row, column=2, value=cnt)
            c.font = Font(name=self.FONT_NAME, size=10, color="333333")
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = fill; c.border = borde
        rt = len(cats_ord) + 2
        for ci, (v, clr) in enumerate([("TOTAL", "FFFFFF"), (total, "FFD700")], 1):
            c = ws.cell(row=rt, column=ci, value=v)
            c.font = Font(name=self.FONT_NAME, size=11, bold=True, color=clr)
            c.fill = PatternFill(start_color=self.NEGRO, end_color=self.NEGRO, fill_type='solid')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = borde
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.freeze_panes = "A2"

    def _hoja_datos(self, ws, articulos):
        headers = [("FUENTE","FF4444"),("TÍTULO","FFD700"),("RESUMEN CORTO","00CC66"),
                   ("URL","00CCCC"),("FECHA","FF66FF")]
        for ci, (h, c) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(name=self.FONT_NAME, bold=True, size=10, color=c)
            cell.fill = PatternFill(start_color=self.NEGRO, end_color=self.NEGRO, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28
        borde = Border(
            left=Side(style='hair', color=self.GRIS_BORDE),
            right=Side(style='hair', color=self.GRIS_BORDE),
            top=Side(style='hair', color=self.GRIS_BORDE),
            bottom=Side(style='hair', color=self.GRIS_BORDE))
        fn = Font(name=self.FONT_NAME, size=9, color="333333")
        fl = Font(name=self.FONT_NAME, size=9, color="0563C1", underline='single')
        al = Alignment(vertical='center', wrap_text=True)
        fp = PatternFill(start_color=self.GRIS_CLARO, end_color=self.GRIS_CLARO, fill_type='solid')
        fi = PatternFill(start_color=self.BLANCO, end_color=self.BLANCO, fill_type='solid')
        arts = sorted(articulos,
                      key=lambda a: a.get("fecha_dt") or datetime.min.replace(tzinfo=timezone.utc),
                      reverse=True)
        for idx, art in enumerate(arts):
            row = idx + 2
            fill = fp if idx % 2 == 0 else fi
            c = ws.cell(row=row, column=1, value=art.get("fuente", ""))
            c.font = Font(name=self.FONT_NAME, size=9, bold=True, color="333333")
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = fill; c.border = borde
            c = ws.cell(row=row, column=2, value=art.get("titulo", ""))
            c.font = fn; c.alignment = al; c.fill = fill; c.border = borde
            resumen = art.get("resumen", "")
            if len(resumen) > 150: resumen = resumen[:147] + "..."
            c = ws.cell(row=row, column=3, value=resumen)
            c.font = Font(name=self.FONT_NAME, size=8, color="666666")
            c.alignment = al; c.fill = fill; c.border = borde
            c = ws.cell(row=row, column=4, value=art.get("url", ""))
            c.font = fl; c.alignment = al; c.fill = fill; c.border = borde
            try: c.hyperlink = art["url"]
            except Exception: pass
            c = ws.cell(row=row, column=5, value=art.get("fecha_str", ""))
            c.font = fn
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = fill; c.border = borde
            ws.row_dimensions[row].height = 22
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 18
        uf = len(arts) + 1
        if uf > 1: ws.auto_filter.ref = f"A1:E{uf}"
        ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════
# GUI — CustomTkinter
# ═══════════════════════════════════════════════════════════════

def _siguiente_nombre_tabla():
    n = 1
    while True:
        nombre = f"TABLA DE NOTICIAS {n}.xlsx"
        if not os.path.exists(nombre):
            return nombre
        n += 1


CATEGORIAS_GUI = [
    "General", "Economía", "Política", "Deportes", "Tecnología",
    "Cultura", "Judicial", "Internacional", "Salud", "Tendencias",
    "Negocios", "Finanzas", "Colombia",
    "Motor", "Vida", "Virales", "Bogotá", "Mis Finanzas",
]

MAPA_CATEGORIAS = {
    "General": ["general"],
    "Economía": ["economia"],
    "Política": ["politica"],
    "Deportes": ["deportes"],
    "Tecnología": ["tecnologia"],
    "Cultura": ["cultura"],
    "Judicial": ["judicial", "justicia"],
    "Internacional": ["internacional"],
    "Salud": ["salud"],
    "Tendencias": ["tendencias"],
    "Negocios": ["negocios"],
    "Finanzas": ["finanzas"],
    "Colombia": ["colombia"],
    "Motor": ["motor"],
    "Vida": ["vida"],
    "Virales": ["virales"],
    "Bogotá": ["bogota"],
    "Mis Finanzas": ["mis finanzas"],
}

if GUI_AVAILABLE:
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")


class RedirectText:
    def __init__(self, text_ctrl):
        self.output = text_ctrl
        self.after = text_ctrl.after
    def write(self, string):
        self.after(0, self._escribir, string)
    def _escribir(self, string):
        self.output.insert("end", string)
        self.output.see("end")
        self.output.update_idletasks()
    def flush(self): pass


import calendar

class CTkCalendar(ctk.CTkToplevel):
    def __init__(self, master, current_date=None, command=None, **kwargs):
        super().__init__(master, **kwargs)
        self.title("Seleccionar Fecha")
        self.geometry("320x340")
        self.resizable(False, False)
        self.attributes("-topmost", True)
        self.grab_set()
        self.command = command
        self.today = date.today()
        self.current_date = current_date if current_date else self.today
        self.display_year = self.current_date.year
        self.display_month = self.current_date.month
        self._build_header()
        self._build_body()

    def _build_header(self):
        self.header_fr = ctk.CTkFrame(self, fg_color="transparent")
        self.header_fr.pack(fill="x", pady=10)
        ctk.CTkButton(self.header_fr, text="<", width=30, hover_color="#217346",
                      fg_color="#005931", command=self._prev_month).pack(side="left", padx=10)
        self.lbl_month = ctk.CTkLabel(self.header_fr,
                                       font=ctk.CTkFont(weight="bold", size=14), text="")
        self.lbl_month.pack(side="left", expand=True)
        ctk.CTkButton(self.header_fr, text=">", width=30, hover_color="#217346",
                      fg_color="#005931", command=self._next_month).pack(side="right", padx=10)

    def _build_body(self):
        if hasattr(self, 'body_fr'):
            self.body_fr.destroy()
        self.body_fr = ctk.CTkFrame(self, fg_color="transparent")
        self.body_fr.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        days = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
        for c, d in enumerate(days):
            ctk.CTkLabel(self.body_fr, text=d,
                         font=ctk.CTkFont(weight="bold")).grid(row=0, column=c, padx=5, pady=5)
        cal = calendar.monthcalendar(self.display_year, self.display_month)
        meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        self.lbl_month.configure(text=f"{meses[self.display_month-1]} {self.display_year}")
        for r, week in enumerate(cal):
            for c, day in enumerate(week):
                if day != 0:
                    es_hoy = (day == self.today.day and
                              self.display_month == self.today.month and
                              self.display_year == self.today.year)
                    btn = ctk.CTkButton(self.body_fr, text=str(day), width=35, height=35,
                        fg_color="#005931" if es_hoy else "transparent",
                        text_color="white" if es_hoy else "black",
                        hover_color="#217346",
                        command=lambda d=day: self._select_date(d))
                    btn.grid(row=r+1, column=c, padx=2, pady=2)

    def _prev_month(self):
        if self.display_month == 1:
            self.display_month = 12; self.display_year -= 1
        else:
            self.display_month -= 1
        self._build_body()

    def _next_month(self):
        if self.display_month == 12:
            self.display_month = 1; self.display_year += 1
        else:
            self.display_month += 1
        self._build_body()

    def _select_date(self, day):
        selected = date(self.display_year, self.display_month, day)
        if self.command: self.command(selected)
        self.destroy()


class CTkDateEntry(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.entry = ctk.CTkEntry(self, width=110)
        self.entry.pack(side="left", padx=(0, 5))
        self.btn = ctk.CTkButton(self, text="📅", width=30,
                                  fg_color="#005931", hover_color="#217346",
                                  command=self._open_cal)
        self.btn.pack(side="left")
        self.date = date.today()
        self.delete(0, "end")
        self.insert(0, self.date.strftime("%Y-%m-%d"))

    def _open_cal(self):
        CTkCalendar(self.winfo_toplevel(), current_date=self.date, command=self._on_select)

    def _on_select(self, sel_date):
        self.date = sel_date
        self.delete(0, "end")
        self.insert(0, self.date.strftime("%Y-%m-%d"))

    def get(self): return self.entry.get()
    def delete(self, first, last=None): self.entry.delete(first, last)
    def insert(self, index, string):
        self.entry.insert(index, string)
        try:
            self.date = datetime.strptime(string, "%Y-%m-%d").date()
        except ValueError:
            pass


class AppNoticiasIDEAS(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("Light")
        self.title("Buscador de Noticias CAPA BRINDADA V.6")
        self.geometry("1200x870")
        self.minsize(1100, 750)
        self.configure(fg_color="#f8f9fa")
        try:
            base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
            icon_path = os.path.join(base_path, "blindado_icon.ico")
            self.iconbitmap(icon_path)
        except Exception:
            pass
        self.vars_categorias = {}
        self.create_widgets()
        sys.stdout = RedirectText(self.consola)
        sys.stderr = RedirectText(self.consola)
        import logging as _logging
        logger = _logging.getLogger("BuscadorNoticias")
        logger.handlers = []
        logger.addHandler(_logging.StreamHandler(sys.stdout))
        self.mostrar_bienvenida()

    def mostrar_bienvenida(self):
        print("[READY] System kernel initialized. Core v6.0 — CAPA BRINDADA")
        print("[SYNC] Connected to RSS/News XML Feeds ... OK")
        print("[SYNC] Local node 'Colombia-Main' active.")
        print("--------------------------------------------------")
        print("  Programa diseñado por Sebastian Rozo.")
        print("  Todos los derechos reservados.")
        print("  Utilizar con responsabilidad.")
        print("--------------------------------------------------")
        print("[FIX] Whitelist Google News: CORREGIDA")
        print("[FIX] Filtro inglés agresivo: CORREGIDO")
        print("[FIX] False positives El Tiempo: CORREGIDOS")
        print("[NEW] Fuentes agregadas: La Nota Econ., Raddar,")
        print("      Mi Bolsillo, Agro Negocios, DIAN, Ministerios,")
        print("      FMI, Banco Mundial, BID, CCB y más.")
        print("[NEW] Desplazamiento de fechas: ACTIVO")
        print("      (selección [ini,fin] → busca [fin, fin+delta])")
        print("[IDLE] Awaiting search dispatch...")
        print("")

    def create_widgets(self):
        font_family = "Helvetica"
        header_fr = ctk.CTkFrame(self, fg_color="transparent")
        header_fr.pack(fill="x", padx=30, pady=(25, 10))
        ctk.CTkLabel(header_fr, text="Buscador de Noticias CAPA BRINDADA V.6",
                     font=ctk.CTkFont(family=font_family, size=28, weight="bold"),
                     text_color="#191c1d").pack(anchor="w")
        ctk.CTkLabel(header_fr,
                     text="Configure los parámetros de búsqueda para la extracción y análisis de prensa.",
                     font=ctk.CTkFont(family=font_family, size=15),
                     text_color="#3f4941").pack(anchor="w")

        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=30, pady=10)
        main_frame.columnconfigure(0, weight=7)
        main_frame.columnconfigure(1, weight=5)

        panel_izq = ctk.CTkFrame(main_frame, fg_color="transparent")
        panel_izq.grid(row=0, column=0, sticky="nsew", padx=(0, 20))

        # Categorías
        fr_cat = ctk.CTkFrame(panel_izq, fg_color="#ffffff",
                               border_color="#e1e3e4", border_width=1, corner_radius=12)
        fr_cat.pack(fill="x", pady=(0, 15), ipadx=5, ipady=5)
        ctk.CTkLabel(fr_cat, text="CATEGORÍAS DE BÚSQUEDA",
                     font=ctk.CTkFont(family=font_family, size=11, weight="bold"),
                     text_color="#6f7a70").pack(anchor="w", padx=15, pady=(10, 0))
        self.scroll_cat = ctk.CTkScrollableFrame(fr_cat, height=140, fg_color="transparent")
        self.scroll_cat.pack(fill="x", padx=10, pady=5)
        col, row = 0, 0
        for cat in CATEGORIAS_GUI:
            var = ctk.BooleanVar(value=True)
            chk = ctk.CTkCheckBox(self.scroll_cat, text=cat, variable=var,
                                  font=ctk.CTkFont(family=font_family, size=12, weight="bold"),
                                  fg_color="#005931", hover_color="#217346",
                                  text_color="#191c1d", border_color="#bfc9be")
            chk.grid(row=row, column=col, padx=8, pady=6, sticky="w")
            self.vars_categorias[cat] = var
            col += 1
            if col > 2:
                col = 0; row += 1
        fr_btn_cat = ctk.CTkFrame(fr_cat, fg_color="transparent")
        fr_btn_cat.pack(fill="x", padx=15, pady=(0, 10))
        ctk.CTkButton(fr_btn_cat, text="Todas", width=80, height=26,
                       fg_color="#005931", hover_color="#217346", text_color="white",
                       font=ctk.CTkFont(family=font_family, size=11, weight="bold"),
                       command=lambda: self._marcar(True)).pack(side="left", padx=(0, 10))
        ctk.CTkButton(fr_btn_cat, text="Ninguna", width=80, height=26,
                       fg_color="#e7e8e9", text_color="#3f4941", hover_color="#d9dadb",
                       font=ctk.CTkFont(family=font_family, size=11, weight="bold"),
                       command=lambda: self._marcar(False)).pack(side="left")

        # Fechas — NEW: label explicativo del desplazamiento
        fr_fecha = ctk.CTkFrame(panel_izq, fg_color="#ffffff",
                                 border_color="#e1e3e4", border_width=1, corner_radius=12)
        fr_fecha.pack(fill="x", pady=(0, 15), ipadx=5, ipady=5)
        c_sw = ctk.CTkFrame(fr_fecha, fg_color="transparent")
        c_sw.pack(fill="x", padx=15, pady=(10, 5))
        self.var_usar_fecha = ctk.BooleanVar(value=True)
        ctk.CTkSwitch(c_sw, text="FILTRAR POR RANGO DE FECHAS", variable=self.var_usar_fecha,
                      font=ctk.CTkFont(family=font_family, size=11, weight="bold"),
                      text_color="#6f7a70", progress_color="#005931").pack(side="left")
        fr_inp = ctk.CTkFrame(fr_fecha, fg_color="transparent")
        fr_inp.pack(fill="x", padx=15, pady=(5, 5))
        ctk.CTkLabel(fr_inp, text="Período:",
                     font=ctk.CTkFont(family=font_family, size=12, weight="bold"),
                     text_color="#191c1d").pack(side="left", padx=(0, 5))
        self.entry_fecha_ini = CTkDateEntry(fr_inp)
        self.entry_fecha_ini.pack(side="left", padx=(0, 5))
        ctk.CTkLabel(fr_inp, text="→",
                     font=ctk.CTkFont(family=font_family, size=12),
                     text_color="#6f7a70").pack(side="left", padx=(0, 5))
        self.entry_fecha_fin = CTkDateEntry(fr_inp)
        self.entry_fecha_fin.pack(side="left", padx=(0, 10))
        ctk.CTkButton(fr_inp, text="Hoy", width=50, fg_color="#005931",
                       hover_color="#217346", text_color="white",
                       command=self._poner_hoy).pack(side="left", padx=(5, 5))
        ctk.CTkButton(fr_inp, text="✕", width=30, fg_color="#ba1a1a",
                       hover_color="#93000a", text_color="white",
                       command=lambda: [self.entry_fecha_ini.delete(0, "end"),
                                        self.entry_fecha_fin.delete(0, "end")]).pack(side="left")

        # NEW: Label informativo sobre el desplazamiento de fechas
        self.lbl_fecha_info = ctk.CTkLabel(
            fr_fecha,
            text="📅 Las noticias se buscarán en el período SIGUIENTE de igual duración.",
            font=ctk.CTkFont(family=font_family, size=10),
            text_color="#005931"
        )
        self.lbl_fecha_info.pack(anchor="w", padx=15, pady=(0, 8))

        # Filtros adicionales
        fr_filtros = ctk.CTkFrame(panel_izq, fg_color="transparent")
        fr_filtros.pack(fill="x", pady=(0, 15))
        fr_arg = ctk.CTkFrame(fr_filtros, fg_color="#ffffff",
                               border_color="#e1e3e4", border_width=1, corner_radius=12)
        fr_arg.pack(fill="x", pady=(0, 10))
        self.var_filtrar_argentina = ctk.BooleanVar(value=True)
        ctk.CTkSwitch(fr_arg, text="Omitir Noticias de Argentina",
                      variable=self.var_filtrar_argentina,
                      font=ctk.CTkFont(family=font_family, size=13, weight="bold"),
                      text_color="#191c1d", progress_color="#005931").pack(side="left", padx=20, pady=15)
        fr_scope = ctk.CTkFrame(fr_filtros, fg_color="#ffffff",
                                 border_color="#e1e3e4", border_width=1, corner_radius=12)
        fr_scope.pack(fill="x", ipady=3)
        self._tipo_noticias = "ambas"
        self.seg_tipo = ctk.CTkSegmentedButton(
            fr_scope, values=["Nacional", "Internacional", "Ambas"],
            command=self._on_tipo_change,
            font=ctk.CTkFont(family=font_family, size=13, weight="bold"),
            selected_color="#005931", selected_hover_color="#217346",
            unselected_color="#343a40", unselected_hover_color="#495057",
            text_color="white"
        )
        self.seg_tipo.set("Ambas")
        self.seg_tipo.pack(fill="x", padx=10, pady=10)

        self.btn_ejecutar = ctk.CTkButton(
            panel_izq, text="INICIAR BÚSQUEDA Y GENERAR EXCEL",
            font=ctk.CTkFont(family=font_family, size=16, weight="bold"), height=55,
            fg_color="#005931", hover_color="#217346", text_color="#ffffff", corner_radius=12,
            command=self.ejecutar_scraper)
        self.btn_ejecutar.pack(fill="x", pady=(10, 0))
        self.btn_limpieza = ctk.CTkButton(
            panel_izq, text="Limpieza Pendeja",
            font=ctk.CTkFont(family=font_family, size=14, weight="bold"), height=42,
            fg_color="#ba1a1a", hover_color="#93000a", text_color="#ffffff", corner_radius=12,
            command=self._limpieza_pendeja)
        self.btn_limpieza.pack(fill="x", pady=(10, 0))

        panel_der = ctk.CTkFrame(main_frame, fg_color="transparent")
        panel_der.grid(row=0, column=1, sticky="nsew")
        fr_stats = ctk.CTkFrame(panel_der, fg_color="#ffffff",
                                 border_color="#e1e3e4", border_width=1, corner_radius=12)
        fr_stats.pack(fill="x", pady=(0, 15), ipadx=10, ipady=10)
        ctk.CTkLabel(fr_stats, text="ESTADO DEL SISTEMA",
                     font=ctk.CTkFont(family=font_family, size=11, weight="bold"),
                     text_color="#6f7a70").pack(anchor="w", padx=15, pady=(5, 0))
        ctk.CTkLabel(fr_stats, text="En Espera",
                     font=ctk.CTkFont(family=font_family, size=24, weight="bold"),
                     text_color="#005931").pack(anchor="w", padx=15, pady=(0, 5))
        fr_term = ctk.CTkFrame(panel_der, fg_color="#d9dadb", corner_radius=12)
        fr_term.pack(fill="both", expand=True)
        fr_term_head = ctk.CTkFrame(fr_term, fg_color="#ffffff", corner_radius=12)
        fr_term_head.pack(fill="x", padx=2, pady=(2, 0))
        ctk.CTkLabel(fr_term_head, text="TERMINAL DE PROCESO",
                     font=ctk.CTkFont(family="Consolas", size=11, weight="bold"),
                     text_color="#191c1d").pack(side="left", padx=15, pady=8)
        self.consola = ctk.CTkTextbox(fr_term, font=ctk.CTkFont(family="Consolas", size=12),
                                      wrap="word", fg_color="#ffffff", text_color="#3f4941",
                                      corner_radius=0)
        self.consola.pack(fill="both", expand=True, padx=2, pady=(0, 2))

    def _on_tipo_change(self, valor):
        mapa = {"Nacional": "nacional", "Internacional": "internacional", "Ambas": "ambas"}
        self._tipo_noticias = mapa.get(valor, "ambas")

    def _marcar(self, estado):
        for v in self.vars_categorias.values():
            v.set(estado)

    def _poner_hoy(self):
        hoy = datetime.now(ZONA_COLOMBIA).strftime("%Y-%m-%d")
        self.entry_fecha_ini.delete(0, "end")
        self.entry_fecha_ini.insert(0, hoy)
        self.entry_fecha_fin.delete(0, "end")
        self.entry_fecha_fin.insert(0, hoy)

    def _limpieza_pendeja(self):
        try:
            self.btn_ejecutar.configure(state="disabled")
            self.btn_limpieza.configure(state="disabled")
            self.consola.delete("0.0", "end")
            print("[RESET] Limpieza Pendeja activada. Reiniciando la aplicaciÃ³n...")

            if getattr(sys, "frozen", False):
                comando = [sys.executable, *sys.argv[1:]]
                cwd = os.path.dirname(sys.executable)
            else:
                comando = [sys.executable, os.path.abspath(__file__), *sys.argv[1:]]
                cwd = os.path.dirname(os.path.abspath(__file__))

            subprocess.Popen(comando, cwd=cwd)
            self.after(250, self.destroy)
        except Exception as e:
            self.btn_ejecutar.configure(state="normal")
            self.btn_limpieza.configure(state="normal")
            messagebox.showerror("Reinicio fallido", f"No se pudo reiniciar la aplicaciÃ³n:\n{e}")

    def ejecutar_scraper(self):
        seleccionadas = [cat for cat, var in self.vars_categorias.items() if var.get()]
        if not seleccionadas:
            messagebox.showwarning("Atención", "Debes seleccionar al menos una categoría.")
            return

        fecha_ini_obj = None
        fecha_fin_obj = None
        if self.var_usar_fecha.get():
            fecha_ini_txt = self.entry_fecha_ini.get().strip()
            fecha_fin_txt = self.entry_fecha_fin.get().strip()
            if fecha_ini_txt and fecha_fin_txt:
                try:
                    fecha_ini_obj = datetime.strptime(fecha_ini_txt, "%Y-%m-%d").date()
                    fecha_fin_obj = datetime.strptime(fecha_fin_txt, "%Y-%m-%d").date()
                    if fecha_ini_obj > fecha_fin_obj:
                        messagebox.showwarning("Rango inválido",
                                               "La fecha de inicio debe ser menor o igual a la de fin.")
                        return
                except ValueError:
                    messagebox.showwarning("Fecha inválida",
                                           "Formato: YYYY-MM-DD\nEjemplo: 2026-03-23")
                    return
            else:
                messagebox.showwarning("Rango incompleto",
                                       "Debes ingresar tanto fecha de inicio como fecha de fin.")
                return

        self.btn_ejecutar.configure(state="disabled")
        self.consola.delete("0.0", "end")
        self.mostrar_bienvenida()

        tipo = self._tipo_noticias
        filtrar_arg = self.var_filtrar_argentina.get()
        hilo = threading.Thread(
            target=self._proceso,
            args=(seleccionadas, fecha_ini_obj, fecha_fin_obj, tipo, filtrar_arg),
            daemon=True
        )
        hilo.start()

    def _proceso(self, seleccionadas, fecha_inicio, fecha_fin, tipo_noticias="ambas", filtrar_argentina=True):
        try:
            cats_internas = set()
            for cat_gui in seleccionadas:
                for c in MAPA_CATEGORIAS.get(cat_gui, [cat_gui.lower()]):
                    cats_internas.add(c)

            nombre_archivo = _siguiente_nombre_tabla()

            # ═══════════════════════════════════════════════════
            # Fechas: usar directamente las fechas seleccionadas
            # ═══════════════════════════════════════════════════
            fecha_busqueda_inicio = fecha_inicio
            fecha_busqueda_fin = fecha_fin

            tipo_display = {
                "nacional": "🇨🇴 Nacional",
                "internacional": "🌍 Internacional",
                "ambas": "🌐 Ambas"
            }.get(tipo_noticias, tipo_noticias)

            print()
            print("  ═" * 30)
            print(f"  Categorías seleccionadas: {len(seleccionadas)}")
            print(f"  Categorías internas: {', '.join(sorted(cats_internas))}")
            print(f"  Tipo de noticias: {tipo_display}")
            print(f"  Filtro Argentina: {'Sí' if filtrar_argentina else 'No'}")
            if fecha_inicio and fecha_fin:
                print(f"  Período seleccionado: {fecha_inicio} → {fecha_fin}")
                if fecha_inicio != fecha_fin:
                    print(f"  📰 Buscando publicaciones: {fecha_busqueda_inicio} → {fecha_busqueda_fin}")
                else:
                    print(f"  📰 Buscando publicaciones: {fecha_busqueda_inicio}")
            else:
                print(f"  Filtro de fecha: Todas (más recientes)")
            print(f"  Archivo de salida: {nombre_archivo}")
            print("  ═" * 30)
            print()

            resultado = buscar_noticias(
                categorias_seleccionadas=list(cats_internas),
                fecha_inicio=fecha_busqueda_inicio,
                fecha_fin=fecha_busqueda_fin,
                verbose=True,
                tipo_noticias=tipo_noticias,
                filtrar_argentina=filtrar_argentina,
            )

            noticias = resultado["noticias"]

            if noticias:
                generador = GeneradorExcelIDEAS(noticias)
                generador.generar(nombre_archivo)

                print()
                print("  ═" * 30)
                print("  RESUMEN FINAL")
                print("  ═" * 30)

                por_cat = {}
                for art in noticias:
                    cat = art["categoria"]
                    por_cat[cat] = por_cat.get(cat, 0) + 1
                for cat, cnt in sorted(por_cat.items(), key=lambda x: -x[1]):
                    bar = "█" * min(cnt, 30)
                    print(f"  {cat:<20} {cnt:>4}  {bar}")

                print()
                for fuente, cnt in sorted(resultado["conteo_fuentes"].items(),
                                          key=lambda x: -x[1]):
                    if cnt > 0:
                        print(f"  {fuente:<35} {cnt:>4} noticias")

                print(f"\n  TOTAL: {len(noticias)} artículos con fecha verificada")
                print(f"  Archivo: {nombre_archivo}")
                if resultado["fuentes_fallidas"]:
                    print(f"  Feeds sin respuesta: {', '.join(resultado['fuentes_fallidas'])}")
                print()

                total_f = len(noticias)
                self.after(0, lambda: self._msg(
                    "Proceso Terminado",
                    f"Listo. Se encontraron {total_f} resultados.\n\nGuardado en: {nombre_archivo}",
                    "info"))
            else:
                msg = resultado.get("notificacion", "No se encontraron noticias.")
                log.warning(msg)
                self.after(0, lambda: self._msg("Sin resultados", msg, "warning"))

        except Exception as e:
            log.error(f"Error: {e}")
            err = str(e)
            self.after(0, lambda: self._msg("Error", f"Error inesperado:\n{err}", "error"))
        finally:
            self.after(0, lambda: self.btn_ejecutar.configure(state="normal"))

    def _msg(self, titulo, mensaje, tipo):
        self.bell()
        if tipo == "info": messagebox.showinfo(titulo, mensaje)
        elif tipo == "warning": messagebox.showwarning(titulo, mensaje)
        elif tipo == "error": messagebox.showerror(titulo, mensaje)


if __name__ == "__main__":
    if GUI_AVAILABLE:
        app = AppNoticiasIDEAS()
        app.mainloop()
    else:
        print("La interfaz gráfica (CustomTkinter) no está disponible en este entorno.")
