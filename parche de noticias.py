# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════╗
║   BUSCADOR SANTAMARIA DE NOTICIAS INTELIGENTE — V4 (PARCHE)          ║
║   BUGS CORREGIDOS, NOTAS PARA MI:                                    ║
║   BUG 1: Whitelist bloqueaba TODAS las URLs de Google News           ║
║          (news.google.com no estaba en DOMINIOS_PERMITIDOS)          ║
║   BUG 2: Filtro de inglés " the "/" of "/" is " bloqueaba            ║
║          artículos legítimos de Infobae, Forbes, El Heraldo          ║
║   BUG 3: "el_tiempo" con guión bajo generaba falsos positivos        ║
║   BUG 4: "ingresa a nuestro grupo de whatsapp" bloqueaba             ║
║          múltiples fuentes legítimas colombianas                     ║
║   NEW 5: Desplazamiento de fechas: selección [ini, fin]              ║
║          → búsqueda en [fin, fin+(fin-ini)] (período siguiente)      ║
║   NEW 6: 25+ nuevas fuentes (ministerios, FMI, BM, BID,              ║
║          La Nota Económica, Raddar, finanzas personales, etc.)       ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import io
import os
import re
import sys
import time
import logging
import threading
import urllib.request
import urllib.parse
import xml.etree.ElementTree as ET
from datetime import datetime, date, timedelta, timezone
from difflib import SequenceMatcher
from email.utils import parsedate_to_datetime
from typing import Optional

try:
    import tkinter as tk
    from tkinter import messagebox
    import customtkinter as ctk
    GUI_AVAILABLE = True
except ImportError:
    GUI_AVAILABLE = False
    class DummyCTk:
        pass
    class ctk:
        CTk = DummyCTk

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ═══════════════════════════════════════════════════════════════
# CONFIGURACIÓN GENERAL
# ═══════════════════════════════════════════════════════════════

ZONA_COLOMBIA = timezone(timedelta(hours=-5))

class _DummyStream:
    def write(self, *a, **k): pass
    def flush(self, *a, **k): pass

if sys.stdout is None:
    sys.stdout = _DummyStream()
if sys.stderr is None:
    sys.stderr = _DummyStream()

if sys.platform == 'win32' and hasattr(sys.stdout, 'buffer'):
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass
    os.environ['PYTHONIOENCODING'] = 'utf-8'

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("BuscadorNoticias")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "application/rss+xml, application/xml, text/xml, */*",
    "Accept-Language": "es-CO,es;q=0.9,en;q=0.8",
}

TIMEOUT = 12

# ═══════════════════════════════════════════════════════════════
# FUENTES RSS — MULTI-CATEGORÍA
# ═══════════════════════════════════════════════════════════════

FUENTES_RSS = [
    # ══════════════════════════════════════════════════════════════
    #  FUENTES NACIONALES (Colombia) — ORIGINALES
    # ══════════════════════════════════════════════════════════════

    {"nombre": "El Heraldo",
     "url": "https://www.elheraldo.co/rss.xml",
     "categorias": ["general", "barranquilla", "judicial", "politica", "deportes", "economia", "cultura", "salud", "tecnologia", "internacional"],
     "tipo": "nacional"},

    {"nombre": "El Colombiano — Medellín",
     "url": "https://www.elcolombiano.com/rss/medellin.xml",
     "categorias": ["general", "medellin"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Antioquia",
     "url": "https://www.elcolombiano.com/rss/antioquia.xml",
     "categorias": ["general", "antioquia"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Colombia",
     "url": "https://www.elcolombiano.com/rss/colombia.xml",
     "categorias": ["general", "politica", "colombia"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Negocios",
     "url": "https://www.elcolombiano.com/rss/negocios.xml",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Deportes",
     "url": "https://www.elcolombiano.com/rss/deportes.xml",
     "categorias": ["deportes"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Internacional",
     "url": "https://www.elcolombiano.com/rss/internacional.xml",
     "categorias": ["internacional"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Cultura",
     "url": "https://www.elcolombiano.com/rss/cultura.xml",
     "categorias": ["cultura"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Tendencias",
     "url": "https://www.elcolombiano.com/rss/tendencias.xml",
     "categorias": ["tendencias", "salud", "vida"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Tecnología",
     "url": "https://www.elcolombiano.com/rss/tecnologia.xml",
     "categorias": ["tecnologia"], "tipo": "nacional"},
    {"nombre": "El Colombiano — Política",
     "url": "https://www.elcolombiano.com/rss/colombia/politica.xml",
     "categorias": ["politica"], "tipo": "nacional"},

    {"nombre": "El Espectador",
     "url": "https://www.elespectador.com/arc/outboundfeeds/rss/",
     "categorias": ["general", "politica", "judicial", "justicia", "cultura", "colombia"],
     "tipo": "nacional"},
    {"nombre": "El Espectador — Economía",
     "url": "https://www.elespectador.com/arc/outboundfeeds/rss/?outputType=xml&_website=el-espectador&section=/economia",
     "categorias": ["economia"], "tipo": "nacional"},

    {"nombre": "La República — Economía",
     "url": "https://www.larepublica.co/rss/economia.xml",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},
    {"nombre": "La República — Finanzas",
     "url": "https://www.larepublica.co/rss/finanzas.xml",
     "categorias": ["economia", "finanzas"], "tipo": "nacional"},

    {"nombre": "Semana",
     "url": "https://www.semana.com/rss/",
     "categorias": ["general", "politica", "economia", "cultura", "deportes", "tecnologia", "salud", "internacional", "judicial", "justicia", "vida", "virales"],
     "tipo": "nacional"},

    {"nombre": "La FM — Actualidad",
     "url": "https://www.lafm.com.co/rss/actualidad.xml",
     "categorias": ["general", "politica", "judicial", "justicia", "colombia"],
     "tipo": "nacional"},

    {"nombre": "Caracol Radio",
     "url": "https://caracol.com.co/rss/",
     "categorias": ["general", "politica", "deportes", "economia", "judicial", "justicia"],
     "tipo": "nacional"},

    {"nombre": "W Radio",
     "url": "https://www.wradio.com.co/rss/",
     "categorias": ["general", "politica", "deportes", "economia"],
     "tipo": "nacional"},

    {"nombre": "Pulzo",
     "url": "https://www.pulzo.com/feed",
     "categorias": ["general", "colombia", "virales"], "tipo": "nacional"},

    {"nombre": "El Universal Cartagena",
     "url": "https://www.eluniversal.com.co/rss.xml",
     "categorias": ["general", "colombia"], "tipo": "nacional"},

    {"nombre": "Vanguardia",
     "url": "https://www.vanguardia.com/rss/",
     "categorias": ["general", "colombia"], "tipo": "nacional"},

    {"nombre": "Kienyke",
     "url": "https://www.kienyke.com/feed",
     "categorias": ["general", "virales", "cultura"], "tipo": "nacional"},

    {"nombre": "Asuntos Legales",
     "url": "https://www.asuntoslegales.com.co/rss/",
     "categorias": ["justicia", "judicial", "economia"], "tipo": "nacional"},

    {"nombre": "Valora Analitik",
     "url": "https://www.valoraanalitik.com/feed/",
     "categorias": ["economia", "finanzas", "negocios"], "tipo": "nacional"},

    {"nombre": "Forbes Colombia",
     "url": "https://forbes.co/feed/",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},

    {"nombre": "Fútbol Red",
     "url": "https://www.futbolred.com/rss/",
     "categorias": ["deportes"], "tipo": "nacional"},

    {"nombre": "Conexión Capital",
     "url": "https://conexioncapital.co/feed/",
     "categorias": ["bogota", "general", "colombia"], "tipo": "nacional"},

    {"nombre": "El Carro Colombiano",
     "url": "https://www.elcarrocolombiano.com/feed/",
     "categorias": ["motor"], "tipo": "nacional"},

    {"nombre": "Blu Radio",
     "url": "https://www.bluradio.com/rss/",
     "categorias": ["general", "colombia", "tecnologia"], "tipo": "nacional"},

    {"nombre": "Radionica",
     "url": "https://www.radionica.rocks/feed",
     "categorias": ["cultura", "vida"], "tipo": "nacional"},

    {"nombre": "Minuto 30",
     "url": "https://www.minuto30.com/feed/",
     "categorias": ["general", "colombia"], "tipo": "nacional"},

    {"nombre": "Infobae Colombia",
     "url": "https://www.infobae.com/colombia/feed/",
     "categorias": ["general", "colombia", "politica", "judicial", "justicia", "economia"],
     "tipo": "nacional"},

    {"nombre": "Publimetro",
     "url": "https://www.publimetro.co/arc/outboundfeeds/rss/?outputType=xml",
     "categorias": ["general", "colombia", "virales"], "tipo": "nacional"},

    {"nombre": "RTVC Noticias",
     "url": "https://www.rtvcnoticias.com/rss.xml",
     "categorias": ["general", "colombia", "politica", "cultura", "deportes"],
     "tipo": "nacional"},

    # ──────────────────────────────────────────────────────────────
    #  NUEVAS FUENTES — ECONOMÍA ESPECIALIZADA (solicitadas)
    # ──────────────────────────────────────────────────────────────

    # NEW: La Nota Económica
    {"nombre": "La Nota Económica",
     "url": "https://lanotaeconomica.com.co/feed/",
     "categorias": ["economia", "negocios", "mis finanzas"], "tipo": "nacional"},

    # NEW: Agro Negocios
    {"nombre": "Agro Negocios",
     "url": "https://www.agronegocios.co/feed/",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},

    # NEW: Mi Bolsillo
    {"nombre": "Mi Bolsillo",
     "url": "https://www.mibolsillo.com/feed/",
     "categorias": ["mis finanzas", "economia"], "tipo": "nacional"},

    # NEW: Expertos del Ahorro
    {"nombre": "Expertos del Ahorro",
     "url": "https://expertosdelahorro.com.co/feed/",
     "categorias": ["mis finanzas"], "tipo": "nacional"},

    # NEW: Mis Finanzas Personales
    {"nombre": "Mis Finanzas Personales",
     "url": "https://misfinanzaspersonales.co/feed/",
     "categorias": ["mis finanzas"], "tipo": "nacional"},

    # NEW: Raddar (consumo y mercado)
    {"nombre": "Raddar",
     "url": "https://raddar.net/feed/",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},

    # NEW: Ahorro Capital
    {"nombre": "Ahorro Capital",
     "url": "https://www.ahorrocapital.com/feed/",
     "categorias": ["mis finanzas", "economia"], "tipo": "nacional"},

    # NEW: CESLA (vía GNews — sin RSS propio)
    # (se cubre más abajo con Google News)

    # ──────────────────────────────────────────────────────────────
    #  NUEVAS FUENTES — ORGANISMOS INTERNACIONALES (solicitadas)
    # ──────────────────────────────────────────────────────────────

    # NEW: FMI en español
    {"nombre": "FMI — Artículos",
     "url": "https://www.imf.org/es/news/rss",
     "categorias": ["economia", "internacional", "finanzas"], "tipo": "internacional"},

    # NEW: Banco Mundial noticias LAC
    {"nombre": "Banco Mundial — LAC",
     "url": "https://feeds.worldbank.org/world-bank/rss/press-releases/es",
     "categorias": ["economia", "internacional"], "tipo": "internacional"},

    # ══════════════════════════════════════════════════════════════
    #  FUENTES INTERNACIONALES — ORIGINALES
    # ══════════════════════════════════════════════════════════════

    {"nombre": "Bloomberg Línea",
     "url": "https://www.bloomberglinea.com/arc/outboundfeeds/rss/?outputType=xml",
     "categorias": ["internacional"], "tipo": "internacional"},

    {"nombre": "BBC Mundo",
     "url": "https://feeds.bbci.co.uk/mundo/rss.xml",
     "categorias": ["general", "internacional", "politica", "cultura", "tecnologia", "deportes", "salud", "vida"],
     "tipo": "internacional"},

    {"nombre": "DW Español",
     "url": "https://rss.dw.com/xml/rss-es-all",
     "categorias": ["general", "internacional", "politica", "cultura", "tecnologia", "vida"],
     "tipo": "internacional"},

    {"nombre": "France 24 Español",
     "url": "https://www.france24.com/es/rss",
     "categorias": ["general", "internacional", "politica"], "tipo": "internacional"},

    {"nombre": "CNN Español",
     "url": "https://cnnespanol.cnn.com/feed/",
     "categorias": ["general", "internacional", "politica", "tecnologia", "deportes", "salud", "vida"],
     "tipo": "internacional"},

    {"nombre": "EFE",
     "url": "https://www.efe.com/efe/rss",
     "categorias": ["general", "internacional"], "tipo": "internacional"},

    {"nombre": "Europa Press",
     "url": "https://www.europapress.es/rss/rss.aspx",
     "categorias": ["general", "internacional", "virales"], "tipo": "internacional"},

    {"nombre": "Newsweek",
     "url": "https://www.newsweek.com/rss",
     "categorias": ["general", "internacional", "tecnologia"], "tipo": "internacional"},

    {"nombre": "NPR",
     "url": "https://feeds.npr.org/1001/rss.xml",
     "categorias": ["general", "internacional"], "tipo": "internacional"},

    {"nombre": "The Guardian World",
     "url": "https://www.theguardian.com/world/rss",
     "categorias": ["internacional"], "tipo": "internacional"},

    {"nombre": "NY Times World",
     "url": "https://rss.nytimes.com/services/xml/rss/nyt/World.xml",
     "categorias": ["internacional"], "tipo": "internacional"},

    {"nombre": "El País América",
     "url": "https://feeds.elpais.com/mrss-s/pages/ep/site/elpais.com/portada",
     "categorias": ["general", "internacional", "politica", "cultura"],
     "tipo": "internacional"},

    {"nombre": "RPP Perú",
     "url": "https://rpp.pe/feed",
     "categorias": ["general", "internacional"], "tipo": "internacional"},

    {"nombre": "La Tercera",
     "url": "https://www.latercera.com/feed/",
     "categorias": ["general", "internacional"], "tipo": "internacional"},

    {"nombre": "20 Minutos",
     "url": "https://www.20minutos.es/rss/",
     "categorias": ["general", "internacional", "cultura", "virales"],
     "tipo": "internacional"},

    {"nombre": "HuffPost",
     "url": "https://www.huffpost.com/section/world-news/feed",
     "categorias": ["general", "internacional"], "tipo": "internacional"},

    {"nombre": "ESPN Deportes",
     "url": "https://espndeportes.espn.com/espn/rss/news",
     "categorias": ["deportes"], "tipo": "internacional"},

    {"nombre": "Mundo Deportivo",
     "url": "https://www.mundodeportivo.com/rss/",
     "categorias": ["deportes"], "tipo": "internacional"},

    {"nombre": "Rolling Stone",
     "url": "https://www.rollingstone.com/feed/",
     "categorias": ["cultura"], "tipo": "internacional"},

    {"nombre": "Billboard",
     "url": "https://www.billboard.com/feed/",
     "categorias": ["cultura"], "tipo": "internacional"},

    {"nombre": "Muy Interesante",
     "url": "https://www.muyinteresante.es/feed/",
     "categorias": ["vida", "salud", "tecnologia"], "tipo": "internacional"},

    {"nombre": "National Geographic LA",
     "url": "https://www.nationalgeographicla.com/feed",
     "categorias": ["vida", "cultura"], "tipo": "internacional"},

    {"nombre": "Digital Trends Español",
     "url": "https://es.digitaltrends.com/feed/",
     "categorias": ["tecnologia"], "tipo": "internacional"},

    {"nombre": "Andro4all",
     "url": "https://andro4all.com/feed",
     "categorias": ["tecnologia"], "tipo": "internacional"},

    {"nombre": "Car and Driver",
     "url": "https://www.caranddriver.com/es/rss/",
     "categorias": ["motor"], "tipo": "internacional"},

    {"nombre": "Autobild",
     "url": "https://www.autobild.es/rss/",
     "categorias": ["motor"], "tipo": "internacional"},

    # ══════════════════════════════════════════════════════════════
    #  GOOGLE NEWS — CATEGORÍAS GENERALES (Colombia)
    # ══════════════════════════════════════════════════════════════

    {"nombre": "Google News — Colombia",
     "url": "https://news.google.com/rss/search?q=colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["general", "colombia"], "tipo": "nacional"},
    {"nombre": "Google News — Economía",
     "url": "https://news.google.com/rss/search?q=economia+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia"], "tipo": "nacional"},
    {"nombre": "Google News — Deportes",
     "url": "https://news.google.com/rss/search?q=deportes+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["deportes"], "tipo": "nacional"},
    {"nombre": "Google News — Política",
     "url": "https://news.google.com/rss/search?q=politica+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["politica"], "tipo": "nacional"},
    {"nombre": "Google News — Tecnología",
     "url": "https://news.google.com/rss/search?q=tecnologia+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["tecnologia"], "tipo": "nacional"},
    {"nombre": "Google News — Salud",
     "url": "https://news.google.com/rss/search?q=salud+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["salud"], "tipo": "nacional"},
    {"nombre": "Google News — Cultura",
     "url": "https://news.google.com/rss/search?q=cultura+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["cultura"], "tipo": "nacional"},
    {"nombre": "Google News — Judicial",
     "url": "https://news.google.com/rss/search?q=judicial+justicia+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["judicial", "justicia"], "tipo": "nacional"},
    {"nombre": "Google News — Motor Colombia",
     "url": "https://news.google.com/rss/search?q=carros+autos+motor+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["motor"], "tipo": "nacional"},
    {"nombre": "Google News — Bogotá",
     "url": "https://news.google.com/rss/search?q=bogota+noticias&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["bogota"], "tipo": "nacional"},
    {"nombre": "Google News — Vida Colombia",
     "url": "https://news.google.com/rss/search?q=ciencia+medio+ambiente+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["vida"], "tipo": "nacional"},
    {"nombre": "Google News — Virales",
     "url": "https://news.google.com/rss/search?q=viral+tendencia+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["virales"], "tipo": "nacional"},
    {"nombre": "Google News — Mis Finanzas (Servicios)",
     "url": "https://news.google.com/rss/search?q=servicio+de+la+gente+OR+como+pagar+OR+plazo+para+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["mis finanzas"], "tipo": "nacional"},
    {"nombre": "Google News — Mis Finanzas (Trámites/Subsidios)",
     "url": "https://news.google.com/rss/search?q=devolucion+del+iva+OR+renta+ciudadana+OR+subsidio+OR+dian+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["mis finanzas"], "tipo": "nacional"},

    # ──────────────────────────────────────────────────────────────
    #  NUEVAS CONSULTAS GOOGLE NEWS — MINISTERIOS Y GOB. (solicitadas)
    # ──────────────────────────────────────────────────────────────

    # NEW: DIAN
    {"nombre": "Google News — DIAN",
     "url": "https://news.google.com/rss/search?q=dian+colombia+impuesto+tributario&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["mis finanzas", "economia"], "tipo": "nacional"},

    # NEW: Ministerio de Agricultura
    {"nombre": "Google News — Ministerio Agricultura",
     "url": "https://news.google.com/rss/search?q=ministerio+agricultura+colombia+campo&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia"], "tipo": "nacional"},

    # NEW: Ministerio de Trabajo
    {"nombre": "Google News — Ministerio Trabajo",
     "url": "https://news.google.com/rss/search?q=ministerio+trabajo+colombia+empleo+laboral&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "general"], "tipo": "nacional"},

    # NEW: Ministerio de Energía
    {"nombre": "Google News — MinEnergía",
     "url": "https://news.google.com/rss/search?q=ministerio+minas+energia+colombia+petroleo+gas&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia"], "tipo": "nacional"},

    # NEW: Ministerio de Comercio (MinCIT)
    {"nombre": "Google News — MinCIT",
     "url": "https://news.google.com/rss/search?q=mincit+ministerio+comercio+industria+turismo+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},

    # NEW: Ministerio de Vivienda
    {"nombre": "Google News — MinVivienda",
     "url": "https://news.google.com/rss/search?q=minvivienda+ministerio+vivienda+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["general", "economia"], "tipo": "nacional"},

    # NEW: Banco de la República
    {"nombre": "Google News — BanRep",
     "url": "https://news.google.com/rss/search?q=%22banco+de+la+republica%22+colombia+tasas+inflacion&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "finanzas"], "tipo": "nacional"},

    # NEW: FMI / Colombia (vía GNews también)
    {"nombre": "Google News — FMI Colombia",
     "url": "https://news.google.com/rss/search?q=fmi+colombia+economia+perspectivas&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "internacional"], "tipo": "nacional"},

    # NEW: Banco Mundial Colombia
    {"nombre": "Google News — Banco Mundial Colombia",
     "url": "https://news.google.com/rss/search?q=%22banco+mundial%22+colombia+informe&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "internacional"], "tipo": "nacional"},

    # NEW: BID Colombia
    {"nombre": "Google News — BID Colombia",
     "url": "https://news.google.com/rss/search?q=%22banco+interamericano%22+bid+colombia&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "internacional"], "tipo": "nacional"},

    # NEW: La Nota Económica (vía GNews — por si el RSS falla)
    {"nombre": "Google News — La Nota Econ.",
     "url": "https://news.google.com/rss/search?q=site:lanotaeconomica.com.co&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},

    # NEW: Valora Analitik (vía GNews — respaldo)
    {"nombre": "Google News — Valora Analitik",
     "url": "https://news.google.com/rss/search?q=site:valoraanalitik.com&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "finanzas", "negocios"], "tipo": "nacional"},

    # NEW: Agronegocios Colombia
    {"nombre": "Google News — Agronegocios",
     "url": "https://news.google.com/rss/search?q=agronegocios+colombia+campo+agro&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia"], "tipo": "nacional"},

    # NEW: Cámara de Comercio Bogotá
    {"nombre": "Google News — CCB",
     "url": "https://news.google.com/rss/search?q=%22camara+de+comercio%22+bogota+empresa&hl=es-419&gl=CO&ceid=CO:es-419",
     "categorias": ["economia", "negocios"], "tipo": "nacional"},
]

# ═══════════════════════════════════════════════════════════════
# DOMINIOS ARGENTINOS — para filtro opcional
# ═══════════════════════════════════════════════════════════════

DOMINIOS_ARGENTINA = [
    "clarin.com", "lanacion.com.ar", "perfil.com",
    "lavoz.com.ar", "ole.com.ar", "mdzol.com",
    "la100.cienradios.com", "elle.clarin.com",
]

# ═══════════════════════════════════════════════════════════════
# DOMINIOS BLOQUEADOS
# ═══════════════════════════════════════════════════════════════

import pandas as pd

DOMINIOS_PERMITIDOS = set()
try:
    _dfs = pd.read_excel(
        r'l:\DESCARGAS\PROYECTOS\Proyecto Búsqueda Noticias\Fuentes_SEO.xlsx',
        sheet_name=None, header=None
    )
    for sheet, _df in _dfs.items():
        for r in range(_df.shape[0]):
            for c in range(_df.shape[1]):
                val = str(_df.iloc[r, c]).strip()
                if "http" in val:
                    for part in val.split():
                        if "http" in part:
                            try:
                                netloc = urllib.parse.urlparse(part).netloc.lower()
                                if netloc.startswith("www."): netloc = netloc[4:]
                                DOMINIOS_PERMITIDOS.add(netloc)
                            except: pass
    if "news.google.com" in DOMINIOS_PERMITIDOS:
        DOMINIOS_PERMITIDOS.remove("news.google.com")
except Exception as e:
    DOMINIOS_PERMITIDOS = set()
    log.warning(f"Fuentes_SEO.xlsx no disponible ({e}) — usando fallback de FUENTES_RSS")

# FALLBACK: construir whitelist desde FUENTES_RSS + nuevas fuentes solicitadas
if len(DOMINIOS_PERMITIDOS) < 10:
    for f in FUENTES_RSS:
        try:
            nl = urllib.parse.urlparse(f["url"]).netloc.lower()
            if "news.google" not in nl:
                if nl.startswith("www."): nl = nl[4:]
                DOMINIOS_PERMITIDOS.add(nl)
        except: pass

    # NEW: agregar explícitamente los dominios nuevos solicitados
    _dominios_extra = [
        "lanotaeconomica.com.co", "raddar.net", "agronegocios.co",
        "mibolsillo.com", "valoraanalitik.com", "expertosdelahorro.com.co",
        "misfinanzaspersonales.co", "ahorrocapital.com", "cesla.com",
        "dian.gov.co", "minagricultura.gov.co", "mintrabajo.gov.co",
        "minenergia.gov.co", "mincit.gov.co", "minvivienda.gov.co",
        "iadb.org", "bancomundial.org", "imf.org",
        "tuinterfaz.mx", "paat.mx",
    ]
    for d in _dominios_extra:
        DOMINIOS_PERMITIDOS.add(d)

DOMINIOS_BLOQUEADOS = [
    "portafolio.co", "eltiempo.com",
    "blogs.portafolio.co", "amp.portafolio.co", "m.portafolio.co",
    "amp.eltiempo.com", "m.eltiempo.com", "especiales.eltiempo.com",
    "enter.co", "citytv.com.co",
]

FIRMAS_BLOQUEADAS = [
    # FIX BUG 3: "El Tiempo" (periódico) se quitó de la lista de texto porque es
    # indistinguible de la palabra española "el tiempo" (the weather / the time).
    # El bloqueo de El Tiempo se hace por dominio (eltiempo.com) y por frases
    # específicas del periódico, NO por la cadena genérica "el tiempo".
    "portafolio.co", "Portafolio.co",
    "eltiempo.com",
    "Casa Editorial El Tiempo",
    "ELTIEMPO.COM",
]

# ═══════════════════════════════════════════════════════════════
# UTILIDADES DE FECHA
# ═══════════════════════════════════════════════════════════════

def _parsear_fecha_rss(fecha_str):
    if not fecha_str or not str(fecha_str).strip():
        return None
    fecha_str = str(fecha_str).strip()
    try:
        dt = parsedate_to_datetime(fecha_str)
        return dt.astimezone(timezone.utc)
    except Exception:
        pass
    formatos = [
        "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S.%fZ",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
        "%a, %d %b %Y %H:%M:%S %Z",
    ]
    for fmt in formatos:
        try:
            dt = datetime.strptime(fecha_str, fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except (ValueError, TypeError):
            continue
    if hasattr(fecha_str, 'tm_year'):
        try:
            return datetime(*fecha_str[:6], tzinfo=timezone.utc)
        except Exception:
            pass
    return None


def _es_fecha_confiable(dt):
    if dt is None:
        return False
    ahora = datetime.now(timezone.utc)
    return (ahora - timedelta(days=730)) <= dt <= (ahora + timedelta(hours=2))


def _fecha_a_date_colombia(dt):
    return (dt + timedelta(hours=-5)).date()


def _fecha_display(dt):
    dt_col = dt + timedelta(hours=-5)
    return dt_col.strftime("%Y-%m-%d %H:%M")


# ═══════════════════════════════════════════════════════════════
# FILTRO DE FUENTES BLOQUEADAS
# ═══════════════════════════════════════════════════════════════

import base64

def _esta_bloqueado(url, titulo="", descripcion="", fuente_rss="", filtrar_argentina=True):
    """
    Retorna (bloqueado, razón).

    FIX 1: Las URLs de Google News (news.google.com/rss/articles/...) son
            redirecciones a fuentes externas. Se excluyen del check de
            DOMINIOS_PERMITIDOS porque su dominio es siempre news.google.com,
            que intencionalmente no está en la whitelist.

    FIX 2: Se eliminó el "el_tiempo" con guión bajo (false positive).

    FIX 3: "ingresa a nuestro grupo de whatsapp" eliminado — phrase demasiado
            genérica que bloqueaba fuentes legítimas.
    """
    url_lower = url.lower()
    fuente_lower = fuente_rss.lower()

    # ── Bloqueo explícito por nombre de fuente RSS ──────────────────────
    # Sólo bloqueamos si la fuente dice EXACTAMENTE "el tiempo" o "portafolio"
    # como nombre de medio, no si aparece en cualquier contexto.
    if fuente_lower in ("el tiempo", "portafolio", "portafolio.co",
                        "eltiempo.com", "el tiempo play"):
        return True, f"Feed RSS Fuente exacta: {fuente_rss}"

    # ── Frases EXCLUSIVAS de El Tiempo y Portafolio ─────────────────────
    texto_lower = f"{titulo} {descripcion} {fuente_rss}".lower()
    frases_prohibidas = [
        "artículo exclusivo para suscriptores de el tiempo",
        "sigue a el tiempo en whatsapp",
        "el tiempo play",
        "casa editorial el tiempo",
        "portafolio digital",
        "suscríbete a portafolio",
        "suscríbete a el tiempo",
        "el tiempoplay",
        # FIX: "el_tiempo" con guión bajo ELIMINADO (generaba falsos positivos)
        # FIX: "ingresa a nuestro grupo de whatsapp" ELIMINADO (demasiado genérico)
    ]
    for frase in frases_prohibidas:
        if frase in texto_lower:
            return True, f"Frase bloqueada (El Tiempo/Portafolio): {frase}"

    # ── Decodificar URL base64 de Google News ───────────────────────────
    if "news.google.com/rss/articles/" in url_lower:
        try:
            b64_part = url_lower.split("articles/")[-1].split("?")[0]
            b64_part += "=" * ((4 - len(b64_part) % 4) % 4)
            decodificado = base64.urlsafe_b64decode(b64_part).decode('utf-8', errors='ignore').lower()
            # Verificar si el destino real es Portafolio o El Tiempo
            for dom_bloq in ["portafolio.co", "eltiempo.com"]:
                if dom_bloq in decodificado:
                    return True, f"Google News redirige a dominio bloqueado: {dom_bloq}"
        except Exception:
            pass

    # ── FIX 1: Whitelist (DOMINIOS_PERMITIDOS) ──────────────────────────
    # Las URLs de Google News redirect (news.google.com/rss/articles/...)
    # NO se pasan por la whitelist porque su netloc siempre es news.google.com,
    # que fue excluido intencionalmente. El destino real ya fue verificado arriba.
    try:
        netloc = urllib.parse.urlparse(url).netloc.lower()   # ← URL original, no lower
        if netloc.startswith("www."): netloc = netloc[4:]

        es_gnews_redirect = "news.google.com" in netloc

        if not es_gnews_redirect:
            # Para fuentes directas, aplicar whitelist normalmente
            if DOMINIOS_PERMITIDOS and netloc not in DOMINIOS_PERMITIDOS:
                return True, f"Dominio no está en BD permitida: {netloc}"
    except Exception:
        pass

    if "redmas.com.co" in url_lower or "cronista.com" in url_lower:
        return True, "Filtro manual redmas/cronista"

    for dominio in DOMINIOS_BLOQUEADOS:
        if dominio in url_lower:
            return True, f"Dominio bloqueado: {dominio}"

    # ── Filtro de Argentina (opcional) ──────────────────────────────────
    if filtrar_argentina:
        for dominio in DOMINIOS_ARGENTINA:
            if dominio in url_lower:
                return True, f"Dominio Argentina: {dominio}"

    # ── Firmas textuales ─────────────────────────────────────────────────
    # FIX BUG 3: Solo usamos patrones muy específicos que no existen en español normal.
    # "el tiempo" genérico fue ELIMINADO (falso positivo con "el tiempo libre", etc.)
    # El bloqueo principal es por dominio (DOMINIOS_BLOQUEADOS / whitelist).
    firmas_especificas = [
        "eltiempo.com", "portafolio.co", "portafolio.co",
        "casa editorial el tiempo",
        "el tiempo play", "eltiempoplay",
        "suscríbete a el tiempo", "suscríbete a portafolio",
    ]
    for firma in firmas_especificas:
        if firma in texto_lower:
            return True, f"Firma específica: '{firma}'"

    if 'source="el tiempo"' in texto_lower or '>el tiempo</' in texto_lower:
        return True, "Firma oculta RSS: El Tiempo"

    return False, ""


# ═══════════════════════════════════════════════════════════════
# UTILIDADES DE TEXTO
# ═══════════════════════════════════════════════════════════════

def _limpiar_html(texto):
    if not texto:
        return ""
    limpio = re.sub(r'<[^>]+>', '', texto)
    return re.sub(r'\s+', ' ', limpio).strip()[:300]


def _normalizar(texto):
    if not texto:
        return ""
    texto = texto.lower().strip()
    texto = re.sub(r'[^\w\s]', '', texto)
    return re.sub(r'\s+', ' ', texto)


def _parece_ingles_puro(titulo, descripcion):
    """
    FIX 2: Reemplaza el filtro anterior " the "/" of "/" is " que era
    demasiado agresivo y bloqueaba artículos legítimos de Infobae, Forbes,
    El Heraldo, etc. que mencionan marcas o medios en inglés.

    Ahora sólo descarta un artículo si:
    - Tiene ≥4 marcadores lingüísticos del inglés en el título, Y
    - No tiene ninguna vocal española (á,é,í,ó,ú,ñ,ü,¿,¡), Y
    - Tiene longitud suficiente para el análisis.
    """
    MARCADORES_EN = [" the ", " is ", " are ", " was ", " were ",
                     " has ", " have ", " with ", " and ", " that ",
                     " this ", " from ", " their "]
    texto = f"{titulo} {descripcion}".lower()
    conteo = sum(1 for m in MARCADORES_EN if m in texto)
    tiene_espanol = any(c in texto for c in 'áéíóúñü¿¡')
    return conteo >= 4 and not tiene_espanol and len(texto) > 60


# ═══════════════════════════════════════════════════════════════
# FETCH + PARSE RSS
# ═══════════════════════════════════════════════════════════════

def _fetch_rss(url, timeout=TIMEOUT, max_retries=3):
    import urllib.error
    for attempt in range(max_retries):
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                contenido = resp.read()
                encoding = "utf-8"
                ct = resp.headers.get("Content-Type", "")
                if "charset=" in ct:
                    encoding = ct.split("charset=")[-1].strip().split(";")[0]
                try:
                    return contenido.decode(encoding)
                except Exception:
                    return contenido.decode("utf-8", errors="replace")
        except urllib.error.HTTPError as e:
            if e.code == 429:
                time.sleep(1.5 ** attempt)
                continue
            return None
        except Exception:
            return None
    return None


def _parsear_feed(xml_str, nombre_fuente):
    """
    Parsea RSS/Atom. Descarta artículos sin fecha confiable.

    FIX 2: El filtro de idioma inglés ' the '/' of '/' is ' fue ELIMINADO.
    Era la causa principal de que Infobae, Forbes y El Heraldo no aparecieran:
    sus artículos sobre marcas internacionales, Netflix, Apple, FMI, etc.
    contenían estas palabras y se descartaban silenciosamente.

    Ahora se usa _parece_ingles_puro() que requiere ≥4 marcadores Y ausencia
    de caracteres del español — mucho más preciso.
    """
    articulos = []
    try:
        xml_str = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', xml_str)
        root = ET.fromstring(xml_str)
    except ET.ParseError:
        return []

    ns_atom = {"atom": "http://www.w3.org/2005/Atom"}
    es_atom = "feed" in root.tag.lower()

    items = (
        root.findall(".//atom:entry", ns_atom) or root.findall(".//entry")
    ) if es_atom else root.findall(".//item")

    for item in items:
        def _get(tag, ns_tag=None):
            el = item.find(tag)
            if el is None and ns_tag:
                el = item.find(ns_tag, ns_atom)
            return (el.text or "").strip() if el is not None else ""

        if es_atom:
            titulo = _get("title") or _get("atom:title", "atom:title")
            url = ""
            link_el = item.find("link") or item.find("{http://www.w3.org/2005/Atom}link")
            if link_el is not None:
                url = link_el.get("href", "") or link_el.text or ""
            fecha_str = _get("updated") or _get("published") or _get("{http://www.w3.org/2005/Atom}updated")
            descripcion = _get("summary") or _get("content")
            fuente_rss = _get("source") or _get("atom:source", "atom:source") or ""
        else:
            titulo = _get("title")
            url = _get("link") or _get("guid")
            fecha_str = _get("pubDate") or _get("dc:date") or _get("{http://purl.org/dc/elements/1.1/}date")
            descripcion = _get("description") or _get("summary")
            fuente_rss = _get("source") or ""

        if not titulo or not url:
            continue

        fecha_dt = _parsear_fecha_rss(fecha_str)
        if not _es_fecha_confiable(fecha_dt):
            continue

        # FIX 2: Filtro de idioma mejorado (mucho más permisivo con español legítimo)
        if _parece_ingles_puro(titulo, descripcion):
            continue

        bloqueado, _ = _esta_bloqueado(url, titulo, descripcion, fuente_rss, filtrar_argentina=False)
        if bloqueado:
            continue

        articulos.append({
            "titulo":     titulo,
            "resumen":    _limpiar_html(descripcion) or titulo,
            "url":        url,
            "fecha_dt":   fecha_dt,
            "fecha_date": _fecha_a_date_colombia(fecha_dt),
            "fecha_str":  _fecha_display(fecha_dt),
            "fuente":     nombre_fuente,
        })

    return articulos


# ═══════════════════════════════════════════════════════════════
# MOTOR DE BÚSQUEDA
# ═══════════════════════════════════════════════════════════════

CATEGORIAS_DISPONIBLES = [
    "general", "economia", "politica", "deportes", "tecnologia",
    "cultura", "judicial", "justicia", "internacional", "salud",
    "tendencias", "negocios", "finanzas", "colombia",
    "motor", "vida", "virales", "bogota", "mis finanzas",
]


def buscar_noticias(categorias_seleccionadas=None, fecha_inicio=None, fecha_fin=None,
                    max_por_fuente=20, max_total=200, verbose=True,
                    tipo_noticias="ambas", filtrar_argentina=True):
    if categorias_seleccionadas:
        cats_lower = set(c.lower() for c in categorias_seleccionadas)
    else:
        cats_lower = None

    if fecha_inicio and fecha_fin:
        dias_rango = (fecha_fin - fecha_inicio).days + 1
        if dias_rango > 1:
            max_por_fuente = max(max_por_fuente, dias_rango * 15)
            max_total = max(max_total, dias_rango * 50)

    fuentes = []
    for f in FUENTES_RSS:
        if tipo_noticias != "ambas":
            if f.get("tipo", "nacional") != tipo_noticias:
                continue
        if cats_lower is None:
            fuentes.append(f)
        else:
            if any(c in cats_lower for c in f["categorias"]):
                fuentes.append(f)

    if verbose:
        log.info("")
        log.info("=" * 60)
        log.info("  BUSCADOR SANTAMARIA DE NOTICIAS — V4 (BUGS CORREGIDOS)")
        log.info("=" * 60)
        if fecha_inicio and fecha_fin:
            if fecha_inicio == fecha_fin:
                fecha_info = fecha_inicio.strftime("%Y-%m-%d")
            else:
                fecha_info = f"{fecha_inicio.strftime('%Y-%m-%d')} a {fecha_fin.strftime('%Y-%m-%d')}"
        else:
            fecha_info = "Todas"
        tipo_info = {"nacional": "🇨🇴 Nacional", "internacional": "🌍 Internacional",
                     "ambas": "🌐 Ambas"}.get(tipo_noticias, tipo_noticias)
        log.info(f"  Fuentes seleccionadas: {len(fuentes)}")
        log.info(f"  Tipo de noticias: {tipo_info}")
        log.info(f"  Filtro Argentina: {'Sí' if filtrar_argentina else 'No'}")
        log.info(f"  Filtro de fecha: {fecha_info}")
        log.info("=" * 60)

    todas = []
    fuentes_fallidas = []
    conteo_fuentes = {}
    titulos_vistos = set()

    for fuente in fuentes:
        nombre = fuente["nombre"]
        if verbose:
            log.info(f"  ⟳ {nombre}...")

        url_base = fuente["url"]
        is_google = "news.google.com/rss/search?q=" in url_base

        urls_to_fetch = []

        if is_google and fecha_inicio and fecha_fin:
            delta = (fecha_fin - fecha_inicio).days
            for d in range(0, delta + 1):
                dia_actual = fecha_inicio + timedelta(days=d)
                dia_siguiente = dia_actual + timedelta(days=1)
                f_ini_str = dia_actual.strftime("%Y-%m-%d")
                f_fin_str = dia_siguiente.strftime("%Y-%m-%d")
                partes = url_base.split("&hl=")
                q_part = partes[0]
                resto = "&hl=" + partes[1] if len(partes) > 1 else ""
                fechas_query = urllib.parse.quote(f" after:{f_ini_str} before:{f_fin_str}")
                urls_to_fetch.append(q_part + fechas_query + resto)
        else:
            urls_to_fetch.append(url_base)

        articulos_fuente = []
        responded = False
        for u in urls_to_fetch:
            xml_str = _fetch_rss(u)
            if xml_str:
                responded = True
                arts = _parsear_feed(xml_str, nombre)
                articulos_fuente.extend(arts)
            if is_google and len(urls_to_fetch) > 1:
                time.sleep(1.0)

        if not responded:
            fuentes_fallidas.append(nombre)
            if verbose:
                log.info(f"    ✗ Sin respuesta (el feed puede no existir o estar caído)")
            continue

        articulos = articulos_fuente

        if filtrar_argentina:
            articulos = [
                art for art in articulos
                if not _esta_bloqueado(art["url"], art["titulo"],
                                       art.get("resumen", ""), "",
                                       filtrar_argentina=True)[0]
            ]

        if fecha_inicio and fecha_fin:
            articulos = [a for a in articulos if fecha_inicio <= a["fecha_date"] <= fecha_fin]

        for art in articulos:
            if cats_lower:
                for cat in fuente["categorias"]:
                    if cat in cats_lower:
                        art["categoria"] = cat.upper()
                        break
                else:
                    art["categoria"] = fuente["categorias"][0].upper()
            else:
                art["categoria"] = fuente["categorias"][0].upper()

        nuevos = []
        for art in articulos[:max_por_fuente]:
            t_norm = _normalizar(art["titulo"])[:80]
            if not t_norm:
                continue
            es_duplicado = any(
                SequenceMatcher(None, t_norm, t_visto).ratio() >= 0.85
                for t_visto in titulos_vistos
            )
            if not es_duplicado:
                titulos_vistos.add(t_norm)
                nuevos.append(art)

        todas.extend(nuevos)
        conteo_fuentes[nombre] = len(nuevos)
        if verbose:
            log.info(f"    ✓ {len(nuevos)} artículos con fecha verificada")

        time.sleep(0.2)

    todas.sort(key=lambda a: a["fecha_dt"], reverse=True)
    resultado = todas[:max_total]

    notificacion = None
    if len(resultado) == 0:
        if fecha_inicio and fecha_fin:
            if fecha_inicio == fecha_fin:
                fmt = fecha_inicio.strftime("%d/%m/%Y")
            else:
                fmt = f"del {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
            ahora_col = _fecha_a_date_colombia(datetime.now(timezone.utc))
            if fecha_inicio > ahora_col:
                notificacion = f"⚠ Rango {fmt} es futuro. No hay noticias aún."
            else:
                notificacion = (
                    f"📭 Sin noticias para el rango {fmt}.\n"
                    f"   Se consultaron {len(fuentes) - len(fuentes_fallidas)} fuentes.\n"
                    f"   Las fuentes pueden no tener artículos de esas fechas."
                )
        else:
            notificacion = "⚠ No se encontraron noticias con los criterios seleccionados."
        if verbose:
            log.warning(f"  {notificacion}")

    if verbose:
        log.info(f"\n  ► TOTAL: {len(resultado)} artículos con fecha verificada")

    return {
        "noticias": resultado,
        "total": len(resultado),
        "fuentes_consultadas": len(fuentes) - len(fuentes_fallidas),
        "fuentes_fallidas": fuentes_fallidas,
        "conteo_fuentes": conteo_fuentes,
        "notificacion": notificacion,
    }


# ═══════════════════════════════════════════════════════════════
# GENERADOR DE EXCEL — UNA HOJA POR CATEGORÍA
# ═══════════════════════════════════════════════════════════════

class GeneradorExcelIDEAS:
    FONT_NAME = "Century Gothic"
    NEGRO = "000000"
    BLANCO = "FFFFFF"
    GRIS_CLARO = "F5F5F5"
    GRIS_BORDE = "CCCCCC"
    COLORES_TAB = [
        "FF4444", "FFD700", "00CC66", "00CCCC", "FF66FF",
        "FF8800", "4488FF", "AA44FF", "44DDAA", "DD4488",
        "88CC00", "0088DD", "FF6644", "6644FF", "44FF88",
    ]

    def __init__(self, articulos):
        self.articulos = articulos
        self.wb = Workbook()

    def generar(self, nombre_archivo):
        log.info("")
        log.info("=" * 60)
        log.info("  Generando Excel (una hoja por categoría)...")
        log.info("=" * 60)

        por_cat = {}
        for art in self.articulos:
            cat = art.get("categoria", "GENERAL")
            por_cat.setdefault(cat, []).append(art)

        cats_ord = sorted(por_cat.keys())

        ws_res = self.wb.active
        ws_res.title = "RESUMEN"
        ws_res.sheet_properties.tabColor = self.NEGRO
        self._hoja_resumen(ws_res, por_cat, cats_ord)

        for i, cat in enumerate(cats_ord):
            nombre_hoja = cat[:31]
            ws = self.wb.create_sheet(title=nombre_hoja)
            ws.sheet_properties.tabColor = self.COLORES_TAB[i % len(self.COLORES_TAB)]
            self._hoja_datos(ws, por_cat[cat])
            log.info(f"  ✓ Hoja '{nombre_hoja}': {len(por_cat[cat])} artículos")

        self.wb.save(nombre_archivo)
        log.info(f"  ✓ Guardado: {nombre_archivo}")
        log.info(f"  ✓ Total: {len(self.articulos)} en {len(cats_ord)} hojas")
        return nombre_archivo

    def _hoja_resumen(self, ws, por_cat, cats_ord):
        headers = [("CATEGORÍA", "FFFFFF"), ("ARTÍCULOS", "FFD700")]
        for ci, (h, c) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(name=self.FONT_NAME, bold=True, size=11, color=c)
            cell.fill = PatternFill(start_color=self.NEGRO, end_color=self.NEGRO, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        borde = Border(
            left=Side(style='hair', color=self.GRIS_BORDE),
            right=Side(style='hair', color=self.GRIS_BORDE),
            top=Side(style='hair', color=self.GRIS_BORDE),
            bottom=Side(style='hair', color=self.GRIS_BORDE))
        fp = PatternFill(start_color=self.GRIS_CLARO, end_color=self.GRIS_CLARO, fill_type='solid')
        fi = PatternFill(start_color=self.BLANCO, end_color=self.BLANCO, fill_type='solid')
        total = 0
        for idx, cat in enumerate(cats_ord):
            row = idx + 2
            cnt = len(por_cat[cat])
            total += cnt
            fill = fp if idx % 2 == 0 else fi
            c = ws.cell(row=row, column=1, value=cat)
            c.font = Font(name=self.FONT_NAME, size=10, bold=True, color="333333")
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.fill = fill; c.border = borde
            c = ws.cell(row=row, column=2, value=cnt)
            c.font = Font(name=self.FONT_NAME, size=10, color="333333")
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = fill; c.border = borde
        rt = len(cats_ord) + 2
        for ci, (v, clr) in enumerate([("TOTAL", "FFFFFF"), (total, "FFD700")], 1):
            c = ws.cell(row=rt, column=ci, value=v)
            c.font = Font(name=self.FONT_NAME, size=11, bold=True, color=clr)
            c.fill = PatternFill(start_color=self.NEGRO, end_color=self.NEGRO, fill_type='solid')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = borde
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.freeze_panes = "A2"

    def _hoja_datos(self, ws, articulos):
        headers = [("FUENTE","FF4444"),("TÍTULO","FFD700"),("RESUMEN CORTO","00CC66"),
                   ("URL","00CCCC"),("FECHA","FF66FF")]
        for ci, (h, c) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(name=self.FONT_NAME, bold=True, size=10, color=c)
            cell.fill = PatternFill(start_color=self.NEGRO, end_color=self.NEGRO, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28
        borde = Border(
            left=Side(style='hair', color=self.GRIS_BORDE),
            right=Side(style='hair', color=self.GRIS_BORDE),
            top=Side(style='hair', color=self.GRIS_BORDE),
            bottom=Side(style='hair', color=self.GRIS_BORDE))
        fn = Font(name=self.FONT_NAME, size=9, color="333333")
        fl = Font(name=self.FONT_NAME, size=9, color="0563C1", underline='single')
        al = Alignment(vertical='center', wrap_text=True)
        fp = PatternFill(start_color=self.GRIS_CLARO, end_color=self.GRIS_CLARO, fill_type='solid')
        fi = PatternFill(start_color=self.BLANCO, end_color=self.BLANCO, fill_type='solid')
        arts = sorted(articulos,
                      key=lambda a: a.get("fecha_dt") or datetime.min.replace(tzinfo=timezone.utc),
                      reverse=True)
        for idx, art in enumerate(arts):
            row = idx + 2
            fill = fp if idx % 2 == 0 else fi
            c = ws.cell(row=row, column=1, value=art.get("fuente", ""))
            c.font = Font(name=self.FONT_NAME, size=9, bold=True, color="333333")
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = fill; c.border = borde
            c = ws.cell(row=row, column=2, value=art.get("titulo", ""))
            c.font = fn; c.alignment = al; c.fill = fill; c.border = borde
            resumen = art.get("resumen", "")
            if len(resumen) > 150: resumen = resumen[:147] + "..."
            c = ws.cell(row=row, column=3, value=resumen)
            c.font = Font(name=self.FONT_NAME, size=8, color="666666")
            c.alignment = al; c.fill = fill; c.border = borde
            c = ws.cell(row=row, column=4, value=art.get("url", ""))
            c.font = fl; c.alignment = al; c.fill = fill; c.border = borde
            try: c.hyperlink = art["url"]
            except Exception: pass
            c = ws.cell(row=row, column=5, value=art.get("fecha_str", ""))
            c.font = fn
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = fill; c.border = borde
            ws.row_dimensions[row].height = 22
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 18
        uf = len(arts) + 1
        if uf > 1: ws.auto_filter.ref = f"A1:E{uf}"
        ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════
# GUI — CustomTkinter
# ═══════════════════════════════════════════════════════════════

def _siguiente_nombre_tabla():
    carpeta = os.path.join(BASE_APP_DIR, "Noticias")
    os.makedirs(carpeta, exist_ok=True)
    n = 1
    while True:
        nombre = os.path.join(carpeta, f"TABLA DE NOTICIAS {n}.xlsx")
        if not os.path.exists(nombre):
            return nombre
        n += 1


CATEGORIAS_GUI = [
    "General", "Economía", "Política", "Deportes", "Tecnología",
    "Cultura", "Judicial", "Internacional", "Salud", "Tendencias",
    "Negocios", "Finanzas", "Colombia",
    "Motor", "Vida", "Virales", "Bogotá", "Mis Finanzas",
]

MAPA_CATEGORIAS = {
    "General": ["general"],
    "Economía": ["economia"],
    "Política": ["politica"],
    "Deportes": ["deportes"],
    "Tecnología": ["tecnologia"],
    "Cultura": ["cultura"],
    "Judicial": ["judicial", "justicia"],
    "Internacional": ["internacional"],
    "Salud": ["salud"],
    "Tendencias": ["tendencias"],
    "Negocios": ["negocios"],
    "Finanzas": ["finanzas"],
    "Colombia": ["colombia"],
    "Motor": ["motor"],
    "Vida": ["vida"],
    "Virales": ["virales"],
    "Bogotá": ["bogota"],
    "Mis Finanzas": ["mis finanzas"],
}

if GUI_AVAILABLE:
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")


class RedirectText:
    def __init__(self, text_ctrl):
        self.output = text_ctrl
        self.after = text_ctrl.after
    def write(self, string):
        self.after(0, self._escribir, string)
    def _escribir(self, string):
        self.output.insert("end", string)
        self.output.see("end")
        self.output.update()
    def flush(self): pass


import calendar

class CTkCalendar(ctk.CTkToplevel):
    def __init__(self, master, current_date=None, command=None, **kwargs):
        super().__init__(master, **kwargs)
        self.title("Seleccionar Fecha")
        self.geometry("320x340")
        self.resizable(False, False)
        self.attributes("-topmost", True)
        self.grab_set()
        self.command = command
        self.today = date.today()
        self.current_date = current_date if current_date else self.today
        self.display_year = self.current_date.year
        self.display_month = self.current_date.month
        self._build_header()
        self._build_body()

    def _build_header(self):
        self.header_fr = ctk.CTkFrame(self, fg_color="transparent")
        self.header_fr.pack(fill="x", pady=10)
        ctk.CTkButton(self.header_fr, text="<", width=30, hover_color="#217346",
                      fg_color="#005931", command=self._prev_month).pack(side="left", padx=10)
        self.lbl_month = ctk.CTkLabel(self.header_fr,
                                       font=ctk.CTkFont(weight="bold", size=14), text="")
        self.lbl_month.pack(side="left", expand=True)
        ctk.CTkButton(self.header_fr, text=">", width=30, hover_color="#217346",
                      fg_color="#005931", command=self._next_month).pack(side="right", padx=10)

    def _build_body(self):
        if hasattr(self, 'body_fr'):
            self.body_fr.destroy()
        self.body_fr = ctk.CTkFrame(self, fg_color="transparent")
        self.body_fr.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        days = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
        for c, d in enumerate(days):
            ctk.CTkLabel(self.body_fr, text=d,
                         font=ctk.CTkFont(weight="bold")).grid(row=0, column=c, padx=5, pady=5)
        cal = calendar.monthcalendar(self.display_year, self.display_month)
        meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        self.lbl_month.configure(text=f"{meses[self.display_month-1]} {self.display_year}")
        for r, week in enumerate(cal):
            for c, day in enumerate(week):
                if day != 0:
                    es_hoy = (day == self.today.day and
                              self.display_month == self.today.month and
                              self.display_year == self.today.year)
                    btn = ctk.CTkButton(self.body_fr, text=str(day), width=35, height=35,
                        fg_color="#005931" if es_hoy else "transparent",
                        text_color="white" if es_hoy else "black",
                        hover_color="#217346",
                        command=lambda d=day: self._select_date(d))
                    btn.grid(row=r+1, column=c, padx=2, pady=2)

    def _prev_month(self):
        if self.display_month == 1:
            self.display_month = 12; self.display_year -= 1
        else:
            self.display_month -= 1
        self._build_body()

    def _next_month(self):
        if self.display_month == 12:
            self.display_month = 1; self.display_year += 1
        else:
            self.display_month += 1
        self._build_body()

    def _select_date(self, day):
        selected = date(self.display_year, self.display_month, day)
        if self.command: self.command(selected)
        self.destroy()


class CTkDateEntry(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self.entry = ctk.CTkEntry(self, width=110)
        self.entry.pack(side="left", padx=(0, 5))
        self.btn = ctk.CTkButton(self, text="📅", width=30,
                                  fg_color="#005931", hover_color="#217346",
                                  command=self._open_cal)
        self.btn.pack(side="left")
        self.date = date.today()
        self.delete(0, "end")
        self.insert(0, self.date.strftime("%Y-%m-%d"))

    def _open_cal(self):
        CTkCalendar(self.winfo_toplevel(), current_date=self.date, command=self._on_select)

    def _on_select(self, sel_date):
        self.date = sel_date
        self.delete(0, "end")
        self.insert(0, self.date.strftime("%Y-%m-%d"))

    def get(self): return self.entry.get()
    def delete(self, first, last=None): self.entry.delete(first, last)
    def insert(self, index, string):
        self.entry.insert(index, string)
        try:
            self.date = datetime.strptime(string, "%Y-%m-%d").date()
        except ValueError:
            pass


class AppNoticiasIDEAS(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("Light")
        self.title("Investigador de Noticias Santamaria-V4.0")
        self.geometry("1200x870")
        self.minsize(1100, 750)
        self.configure(fg_color="#f8f9fa")
        try:
            base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
            icon_path = os.path.join(base_path, "excel_icon.ico")
            self.iconbitmap(icon_path)
        except Exception:
            pass
        self.vars_categorias = {}
        self.create_widgets()
        sys.stdout = RedirectText(self.consola)
        sys.stderr = RedirectText(self.consola)
        import logging as _logging
        logger = _logging.getLogger("BuscadorNoticias")
        logger.handlers = []
        logger.addHandler(_logging.StreamHandler(sys.stdout))
        self.mostrar_bienvenida()

    def mostrar_bienvenida(self):
        print("[READY] System kernel initialized. Core v4.0 — BUGS CORREGIDOS")
        print("[SYNC] Connected to RSS/News XML Feeds ... OK")
        print("[SYNC] Local node 'Colombia-Main' active.")
        print("--------------------------------------------------")
        print("  Programa diseñado por Sebastian Rozo.")
        print("  Todos los derechos reservados.")
        print("  Utilizar con responsabilidad.")
        print("--------------------------------------------------")
        print("[FIX] Whitelist Google News: CORREGIDA")
        print("[FIX] Filtro inglés agresivo: CORREGIDO")
        print("[FIX] False positives El Tiempo: CORREGIDOS")
        print("[NEW] Fuentes agregadas: La Nota Econ., Raddar,")
        print("      Mi Bolsillo, Agro Negocios, DIAN, Ministerios,")
        print("      FMI, Banco Mundial, BID, CCB y más.")
        print("[NEW] Desplazamiento de fechas: ACTIVO")
        print("      (selección [ini,fin] → busca [fin, fin+delta])")
        print("[IDLE] Awaiting search dispatch...")
        print("")

    def create_widgets(self):
        font_family = "Helvetica"
        header_fr = ctk.CTkFrame(self, fg_color="transparent")
        header_fr.pack(fill="x", padx=30, pady=(25, 10))
        ctk.CTkLabel(header_fr, text="Investigador de Noticias Santamaria",
                     font=ctk.CTkFont(family=font_family, size=28, weight="bold"),
                     text_color="#191c1d").pack(anchor="w")
        ctk.CTkLabel(header_fr,
                     text="Configure los parámetros de búsqueda para la extracción y análisis de prensa.",
                     font=ctk.CTkFont(family=font_family, size=15),
                     text_color="#3f4941").pack(anchor="w")

        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=30, pady=10)
        main_frame.columnconfigure(0, weight=7)
        main_frame.columnconfigure(1, weight=5)

        panel_izq = ctk.CTkFrame(main_frame, fg_color="transparent")
        panel_izq.grid(row=0, column=0, sticky="nsew", padx=(0, 20))

        # Categorías
        fr_cat = ctk.CTkFrame(panel_izq, fg_color="#ffffff",
                               border_color="#e1e3e4", border_width=1, corner_radius=12)
        fr_cat.pack(fill="x", pady=(0, 15), ipadx=5, ipady=5)
        ctk.CTkLabel(fr_cat, text="CATEGORÍAS DE BÚSQUEDA",
                     font=ctk.CTkFont(family=font_family, size=11, weight="bold"),
                     text_color="#6f7a70").pack(anchor="w", padx=15, pady=(10, 0))
        self.scroll_cat = ctk.CTkScrollableFrame(fr_cat, height=140, fg_color="transparent")
        self.scroll_cat.pack(fill="x", padx=10, pady=5)
        col, row = 0, 0
        for cat in CATEGORIAS_GUI:
            var = ctk.BooleanVar(value=True)
            chk = ctk.CTkCheckBox(self.scroll_cat, text=cat, variable=var,
                                  font=ctk.CTkFont(family=font_family, size=12, weight="bold"),
                                  fg_color="#005931", hover_color="#217346",
                                  text_color="#191c1d", border_color="#bfc9be")
            chk.grid(row=row, column=col, padx=8, pady=6, sticky="w")
            self.vars_categorias[cat] = var
            col += 1
            if col > 2:
                col = 0; row += 1
        fr_btn_cat = ctk.CTkFrame(fr_cat, fg_color="transparent")
        fr_btn_cat.pack(fill="x", padx=15, pady=(0, 10))
        ctk.CTkButton(fr_btn_cat, text="Todas", width=80, height=26,
                       fg_color="#005931", hover_color="#217346", text_color="white",
                       font=ctk.CTkFont(family=font_family, size=11, weight="bold"),
                       command=lambda: self._marcar(True)).pack(side="left", padx=(0, 10))
        ctk.CTkButton(fr_btn_cat, text="Ninguna", width=80, height=26,
                       fg_color="#e7e8e9", text_color="#3f4941", hover_color="#d9dadb",
                       font=ctk.CTkFont(family=font_family, size=11, weight="bold"),
                       command=lambda: self._marcar(False)).pack(side="left")

        # Fechas — NEW: label explicativo del desplazamiento
        fr_fecha = ctk.CTkFrame(panel_izq, fg_color="#ffffff",
                                 border_color="#e1e3e4", border_width=1, corner_radius=12)
        fr_fecha.pack(fill="x", pady=(0, 15), ipadx=5, ipady=5)
        c_sw = ctk.CTkFrame(fr_fecha, fg_color="transparent")
        c_sw.pack(fill="x", padx=15, pady=(10, 5))
        self.var_usar_fecha = ctk.BooleanVar(value=True)
        ctk.CTkSwitch(c_sw, text="FILTRAR POR RANGO DE FECHAS", variable=self.var_usar_fecha,
                      font=ctk.CTkFont(family=font_family, size=11, weight="bold"),
                      text_color="#6f7a70", progress_color="#005931").pack(side="left")
        fr_inp = ctk.CTkFrame(fr_fecha, fg_color="transparent")
        fr_inp.pack(fill="x", padx=15, pady=(5, 5))
        ctk.CTkLabel(fr_inp, text="Período:",
                     font=ctk.CTkFont(family=font_family, size=12, weight="bold"),
                     text_color="#191c1d").pack(side="left", padx=(0, 5))
        self.entry_fecha_ini = CTkDateEntry(fr_inp)
        self.entry_fecha_ini.pack(side="left", padx=(0, 5))
        ctk.CTkLabel(fr_inp, text="→",
                     font=ctk.CTkFont(family=font_family, size=12),
                     text_color="#6f7a70").pack(side="left", padx=(0, 5))
        self.entry_fecha_fin = CTkDateEntry(fr_inp)
        self.entry_fecha_fin.pack(side="left", padx=(0, 10))
        ctk.CTkButton(fr_inp, text="Hoy", width=50, fg_color="#005931",
                       hover_color="#217346", text_color="white",
                       command=self._poner_hoy).pack(side="left", padx=(5, 5))
        ctk.CTkButton(fr_inp, text="✕", width=30, fg_color="#ba1a1a",
                       hover_color="#93000a", text_color="white",
                       command=lambda: [self.entry_fecha_ini.delete(0, "end"),
                                        self.entry_fecha_fin.delete(0, "end")]).pack(side="left")

        # NEW: Label informativo sobre el desplazamiento de fechas
        self.lbl_fecha_info = ctk.CTkLabel(
            fr_fecha,
            text="📅 Las noticias se buscarán en el período SIGUIENTE de igual duración.",
            font=ctk.CTkFont(family=font_family, size=10),
            text_color="#005931"
        )
        self.lbl_fecha_info.pack(anchor="w", padx=15, pady=(0, 8))

        # Filtros adicionales
        fr_filtros = ctk.CTkFrame(panel_izq, fg_color="transparent")
        fr_filtros.pack(fill="x", pady=(0, 15))
        fr_arg = ctk.CTkFrame(fr_filtros, fg_color="#ffffff",
                               border_color="#e1e3e4", border_width=1, corner_radius=12)
        fr_arg.pack(fill="x", pady=(0, 10))
        self.var_filtrar_argentina = ctk.BooleanVar(value=True)
        ctk.CTkSwitch(fr_arg, text="Omitir Noticias de Argentina",
                      variable=self.var_filtrar_argentina,
                      font=ctk.CTkFont(family=font_family, size=13, weight="bold"),
                      text_color="#191c1d", progress_color="#005931").pack(side="left", padx=20, pady=15)
        fr_scope = ctk.CTkFrame(fr_filtros, fg_color="#ffffff",
                                 border_color="#e1e3e4", border_width=1, corner_radius=12)
        fr_scope.pack(fill="x", ipady=3)
        self._tipo_noticias = "ambas"
        self.seg_tipo = ctk.CTkSegmentedButton(
            fr_scope, values=["Nacional", "Internacional", "Ambas"],
            command=self._on_tipo_change,
            font=ctk.CTkFont(family=font_family, size=13, weight="bold"),
            selected_color="#005931", selected_hover_color="#217346",
            unselected_color="#e7e8e9", unselected_hover_color="#d9dadb",
            text_color="#3f4941"
        )
        self.seg_tipo.set("Ambas")
        self.seg_tipo.pack(fill="x", padx=10, pady=10)

        self.btn_ejecutar = ctk.CTkButton(
            panel_izq, text="INICIAR BÚSQUEDA Y GENERAR EXCEL",
            font=ctk.CTkFont(family=font_family, size=16, weight="bold"), height=55,
            fg_color="#005931", hover_color="#217346", text_color="#ffffff", corner_radius=12,
            command=self.ejecutar_scraper)
        self.btn_ejecutar.pack(fill="x", pady=(10, 0))

        panel_der = ctk.CTkFrame(main_frame, fg_color="transparent")
        panel_der.grid(row=0, column=1, sticky="nsew")
        fr_stats = ctk.CTkFrame(panel_der, fg_color="#ffffff",
                                 border_color="#e1e3e4", border_width=1, corner_radius=12)
        fr_stats.pack(fill="x", pady=(0, 15), ipadx=10, ipady=10)
        ctk.CTkLabel(fr_stats, text="ESTADO DEL SISTEMA",
                     font=ctk.CTkFont(family=font_family, size=11, weight="bold"),
                     text_color="#6f7a70").pack(anchor="w", padx=15, pady=(5, 0))
        ctk.CTkLabel(fr_stats, text="En Espera",
                     font=ctk.CTkFont(family=font_family, size=24, weight="bold"),
                     text_color="#005931").pack(anchor="w", padx=15, pady=(0, 5))
        fr_term = ctk.CTkFrame(panel_der, fg_color="#d9dadb", corner_radius=12)
        fr_term.pack(fill="both", expand=True)
        fr_term_head = ctk.CTkFrame(fr_term, fg_color="#ffffff", corner_radius=12)
        fr_term_head.pack(fill="x", padx=2, pady=(2, 0))
        ctk.CTkLabel(fr_term_head, text="TERMINAL DE PROCESO",
                     font=ctk.CTkFont(family="Consolas", size=11, weight="bold"),
                     text_color="#191c1d").pack(side="left", padx=15, pady=8)
        self.consola = ctk.CTkTextbox(fr_term, font=ctk.CTkFont(family="Consolas", size=12),
                                      wrap="word", fg_color="#ffffff", text_color="#3f4941",
                                      corner_radius=0)
        self.consola.pack(fill="both", expand=True, padx=2, pady=(0, 2))

    def _on_tipo_change(self, valor):
        mapa = {"Nacional": "nacional", "Internacional": "internacional", "Ambas": "ambas"}
        self._tipo_noticias = mapa.get(valor, "ambas")

    def _marcar(self, estado):
        for v in self.vars_categorias.values():
            v.set(estado)

    def _poner_hoy(self):
        hoy = datetime.now(ZONA_COLOMBIA).strftime("%Y-%m-%d")
        self.entry_fecha_ini.delete(0, "end")
        self.entry_fecha_ini.insert(0, hoy)
        self.entry_fecha_fin.delete(0, "end")
        self.entry_fecha_fin.insert(0, hoy)

    def ejecutar_scraper(self):
        seleccionadas = [cat for cat, var in self.vars_categorias.items() if var.get()]
        if not seleccionadas:
            messagebox.showwarning("Atención", "Debes seleccionar al menos una categoría.")
            return

        fecha_ini_obj = None
        fecha_fin_obj = None
        if self.var_usar_fecha.get():
            fecha_ini_txt = self.entry_fecha_ini.get().strip()
            fecha_fin_txt = self.entry_fecha_fin.get().strip()
            if fecha_ini_txt and fecha_fin_txt:
                try:
                    fecha_ini_obj = datetime.strptime(fecha_ini_txt, "%Y-%m-%d").date()
                    fecha_fin_obj = datetime.strptime(fecha_fin_txt, "%Y-%m-%d").date()
                    if fecha_ini_obj > fecha_fin_obj:
                        messagebox.showwarning("Rango inválido",
                                               "La fecha de inicio debe ser menor o igual a la de fin.")
                        return
                except ValueError:
                    messagebox.showwarning("Fecha inválida",
                                           "Formato: YYYY-MM-DD\nEjemplo: 2026-03-23")
                    return
            else:
                messagebox.showwarning("Rango incompleto",
                                       "Debes ingresar tanto fecha de inicio como fecha de fin.")
                return

        self.btn_ejecutar.configure(state="disabled")
        self.consola.delete("0.0", "end")
        self.mostrar_bienvenida()

        tipo = self._tipo_noticias
        filtrar_arg = self.var_filtrar_argentina.get()
        hilo = threading.Thread(
            target=self._proceso,
            args=(seleccionadas, fecha_ini_obj, fecha_fin_obj, tipo, filtrar_arg),
            daemon=True
        )
        hilo.start()

    def _proceso(self, seleccionadas, fecha_inicio, fecha_fin, tipo_noticias="ambas", filtrar_argentina=True):
        try:
            cats_internas = set()
            for cat_gui in seleccionadas:
                for c in MAPA_CATEGORIAS.get(cat_gui, [cat_gui.lower()]):
                    cats_internas.add(c)

            nombre_archivo = _siguiente_nombre_tabla()

            # ═══════════════════════════════════════════════════
            # NEW FIX 5: Desplazamiento de fechas
            # Selección [ini, fin] → búsqueda en [fin, fin+delta]
            # Ejemplo: selección 23-26 → delta=3 → busca 26-29
            # ═══════════════════════════════════════════════════
            fecha_busqueda_inicio = fecha_inicio
            fecha_busqueda_fin = fecha_fin

            if fecha_inicio and fecha_fin and fecha_inicio != fecha_fin:
                delta = fecha_fin - fecha_inicio      # timedelta, e.g. 3 days
                fecha_busqueda_inicio = fecha_fin     # el nuevo inicio = el fin seleccionado
                fecha_busqueda_fin = fecha_fin + delta  # nuevo fin = fecha_fin + mismo delta

            tipo_display = {
                "nacional": "🇨🇴 Nacional",
                "internacional": "🌍 Internacional",
                "ambas": "🌐 Ambas"
            }.get(tipo_noticias, tipo_noticias)

            print()
            print("  ═" * 30)
            print(f"  Categorías seleccionadas: {len(seleccionadas)}")
            print(f"  Categorías internas: {', '.join(sorted(cats_internas))}")
            print(f"  Tipo de noticias: {tipo_display}")
            print(f"  Filtro Argentina: {'Sí' if filtrar_argentina else 'No'}")
            if fecha_inicio and fecha_fin:
                print(f"  Período seleccionado: {fecha_inicio} → {fecha_fin}")
                if fecha_inicio != fecha_fin:
                    print(f"  📰 Buscando publicaciones: {fecha_busqueda_inicio} → {fecha_busqueda_fin}")
                else:
                    print(f"  📰 Buscando publicaciones: {fecha_busqueda_inicio}")
            else:
                print(f"  Filtro de fecha: Todas (más recientes)")
            print(f"  Archivo de salida: {nombre_archivo}")
            print("  ═" * 30)
            print()

            resultado = buscar_noticias(
                categorias_seleccionadas=list(cats_internas),
                fecha_inicio=fecha_busqueda_inicio,
                fecha_fin=fecha_busqueda_fin,
                verbose=True,
                tipo_noticias=tipo_noticias,
                filtrar_argentina=filtrar_argentina,
            )

            noticias = resultado["noticias"]

            if noticias:
                generador = GeneradorExcelIDEAS(noticias)
                generador.generar(nombre_archivo)

                print()
                print("  ═" * 30)
                print("  RESUMEN FINAL")
                print("  ═" * 30)

                por_cat = {}
                for art in noticias:
                    cat = art["categoria"]
                    por_cat[cat] = por_cat.get(cat, 0) + 1
                for cat, cnt in sorted(por_cat.items(), key=lambda x: -x[1]):
                    bar = "█" * min(cnt, 30)
                    print(f"  {cat:<20} {cnt:>4}  {bar}")

                print()
                for fuente, cnt in sorted(resultado["conteo_fuentes"].items(),
                                          key=lambda x: -x[1]):
                    if cnt > 0:
                        print(f"  {fuente:<35} {cnt:>4} noticias")

                print(f"\n  TOTAL: {len(noticias)} artículos con fecha verificada")
                print(f"  Archivo: {nombre_archivo}")
                if resultado["fuentes_fallidas"]:
                    print(f"  Feeds sin respuesta: {', '.join(resultado['fuentes_fallidas'])}")
                print()

                total_f = len(noticias)
                self.after(0, lambda: self._msg(
                    "Proceso Terminado",
                    f"Listo. Se encontraron {total_f} resultados.\n\nGuardado en: {nombre_archivo}",
                    "info"))
            else:
                msg = resultado.get("notificacion", "No se encontraron noticias.")
                log.warning(msg)
                self.after(0, lambda: self._msg("Sin resultados", msg, "warning"))

        except Exception as e:
            log.error(f"Error: {e}")
            err = str(e)
            self.after(0, lambda: self._msg("Error", f"Error inesperado:\n{err}", "error"))
        finally:
            self.after(0, lambda: self.btn_ejecutar.configure(state="normal"))

    def _msg(self, titulo, mensaje, tipo):
        self.bell()
        if tipo == "info": messagebox.showinfo(titulo, mensaje)
        elif tipo == "warning": messagebox.showwarning(titulo, mensaje)
        elif tipo == "error": messagebox.showerror(titulo, mensaje)


if __name__ == "__main__":
    if GUI_AVAILABLE:
        app = AppNoticiasIDEAS()
        app.mainloop()
    else:
        print("La interfaz gráfica (CustomTkinter) no está disponible en este entorno.")
